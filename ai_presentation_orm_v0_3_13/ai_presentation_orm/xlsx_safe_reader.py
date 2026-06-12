from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

NS = {
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _col_to_idx(ref: str | None) -> int:
    match = re.match(r"([A-Z]+)", ref or "")
    out = 0
    for char in match.group(1) if match else "":
        out = out * 26 + ord(char) - 64
    return max(out - 1, 0)


class ZipSheet:
    def __init__(self, rows: list[list[object]]):
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = max((len(row) for row in rows), default=0)

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        end = min(max_row or self.max_row, self.max_row)
        for idx in range(max(min_row, 1) - 1, end):
            yield tuple(self._rows[idx])


class ZipWorkbook:
    """Small value-only XLSX reader for exports with invalid style XML."""

    def __init__(self, path: Path):
        self.path = path
        self._sheet_paths = {}
        self._shared = []
        self._cache = {}
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
            if "xl/sharedStrings.xml" in names:
                root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for item in root.findall("m:si", NS):
                    self._shared.append("".join(t.text or "" for t in item.findall(".//m:t", NS)))
            workbook = ET.fromstring(zf.read("xl/workbook.xml"))
            rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rel_map = {rel.get("Id"): rel.get("Target") for rel in rels.findall("pr:Relationship", NS)}
            for sheet in workbook.findall(".//m:sheet", NS):
                name = sheet.get("name") or "Sheet"
                rid = sheet.get(f"{{{NS['r']}}}id")
                target = (rel_map.get(rid) or "").lstrip("/")
                sheet_path = "xl/" + target
                if sheet_path not in names:
                    sheet_path = "xl/worksheets/" + Path(target).name
                self._sheet_paths[name] = sheet_path
        self.sheetnames = list(self._sheet_paths.keys())

    def __getitem__(self, sheet_name: str):
        if sheet_name not in self._sheet_paths:
            raise KeyError(sheet_name)
        if sheet_name not in self._cache:
            self._cache[sheet_name] = ZipSheet(self._read_rows(self._sheet_paths[sheet_name]))
        return self._cache[sheet_name]

    def _cell_value(self, cell):
        cell_type = cell.get("t")
        value = cell.find("m:v", NS)
        inline = cell.find("m:is", NS)
        if cell_type == "s" and value is not None and value.text not in (None, ""):
            idx = int(float(value.text))
            return self._shared[idx] if 0 <= idx < len(self._shared) else ""
        if cell_type == "inlineStr" and inline is not None:
            return "".join(t.text or "" for t in inline.findall(".//m:t", NS))
        if value is not None and value.text is not None:
            return value.text
        return ""

    def _read_rows(self, sheet_path: str) -> list[list[object]]:
        rows = []
        with zipfile.ZipFile(self.path) as zf:
            root = ET.fromstring(zf.read(sheet_path))
        for row_node in root.findall(".//m:sheetData/m:row", NS):
            values = []
            for cell in row_node.findall("m:c", NS):
                idx = _col_to_idx(cell.get("r"))
                while len(values) <= idx:
                    values.append("")
                values[idx] = self._cell_value(cell)
            while values and values[-1] == "":
                values.pop()
            rows.append(values)
        return rows


def open_workbook_safe(path: Path):
    try:
        return load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return ZipWorkbook(path)
