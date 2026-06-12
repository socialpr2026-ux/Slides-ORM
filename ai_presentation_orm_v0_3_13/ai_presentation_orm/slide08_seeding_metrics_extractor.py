from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
import re

from openpyxl import load_workbook

from .project_month_fact_extractor import extract_project_month_facts


MONTHS = [
    ("январ", "Январь", "январе", "январь"),
    ("феврал", "Февраль", "феврале", "февраль"),
    ("март", "Март", "марте", "март"),
    ("апрел", "Апрель", "апреле", "апрель"),
    ("ма", "Май", "мае", "май"),
    ("июн", "Июнь", "июне", "июнь"),
    ("июл", "Июль", "июле", "июль"),
    ("август", "Август", "августе", "август"),
    ("сентябр", "Сентябрь", "сентябре", "сентябрь"),
    ("октябр", "Октябрь", "октябре", "октябрь"),
    ("ноябр", "Ноябрь", "ноябре", "ноябрь"),
    ("декабр", "Декабрь", "декабре", "декабрь"),
]


def _cell(values: list, idx: int | None) -> str:
    if idx is None or idx >= len(values):
        return ""
    return str(values[idx] if values[idx] is not None else "").strip()


def _clean_int(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return max(int(round(float(value))), 0)
    text = re.sub(r"[^0-9,.-]", "", str(value)).replace(",", ".")
    try:
        parsed = int(round(float(text))) if text else 0
        return max(parsed, 0)
    except Exception:
        return 0


def _header_index(headers: list[str], *names: str) -> int | None:
    lowered = [str(item or "").strip().lower() for item in headers]
    for name in names:
        needle = name.lower()
        for idx, header in enumerate(lowered):
            if header == needle:
                return idx
    for name in names:
        needle = name.lower()
        for idx, header in enumerate(lowered):
            if needle in header:
                return idx
    return None


def _views_received_index(headers: list[str]) -> int | None:
    lowered = [str(item or "").strip().lower().replace("ё", "е") for item in headers]
    exact_names = ("просмотров получено", "просмотры")
    for exact in exact_names:
        for idx, header in enumerate(lowered):
            if header == exact:
                return idx
    for idx, header in enumerate(lowered):
        if "просмотров получено" in header:
            return idx
    return None


def _find_header(ws) -> tuple[int, list[str]]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
        headers = [str(value or "").strip() for value in row]
        lowered = [item.lower() for item in headers]
        has_type = any("тип размещения" in header or header == "формат" for header in lowered)
        has_platform = any("площадка" in header or "ссылка" in header for header in lowered)
        has_text = any(header == "текст" or "текст сообщ" in header or "коммент" in header for header in lowered)
        has_generic_month_table = has_platform and has_text and any(
            header == "дата" or header == "ник" or "просмотр" in header or header == "тип поста"
            for header in lowered
        )
        if (has_type and has_platform and has_text) or has_generic_month_table:
            return row_idx, headers
    raise ValueError("Project table header not found on the month sheet")


def _month_sheet(wb, month_sheet: str):
    if month_sheet and month_sheet in wb.sheetnames:
        return wb[month_sheet], month_sheet
    for name in wb.sheetnames:
        try:
            _find_header(wb[name])
            return wb[name], name
        except Exception:
            continue
    return wb[wb.sheetnames[0]], wb.sheetnames[0]


def _month_number(label: str) -> int | None:
    low = str(label or "").strip().lower().replace("ё", "е")
    for idx, (stem, _nom, _prep, _plain) in enumerate(MONTHS, start=1):
        if stem in low:
            return idx
    return None


def _sheet_sort_key(label: str) -> tuple[int, int] | None:
    month = _month_number(label)
    if month is None:
        return None
    match = re.search(r"(20\d{2})", str(label or ""))
    year = int(match.group(1)) if match else 0
    return year, month


def _sheet_month_label(label: str) -> str:
    month = _month_number(label)
    if month is None:
        text = str(label or "").strip()
        return text or "Период"
    return MONTHS[month - 1][1]


def _month_prepositional_phrase(period: str, sheet_name: str) -> str:
    month = _month_number(period) or _month_number(sheet_name)
    if month is None:
        return "в отчетном периоде"
    return "в " + MONTHS[month - 1][2]


def _month_genitive_phrase(period: str, sheet_name: str) -> str:
    month = _month_number(period) or _month_number(sheet_name)
    if month is None:
        return "отчетный период"
    return MONTHS[month - 1][3]


def _previous_month_sheet(wb, current_sheet_name: str):
    current_key = _sheet_sort_key(current_sheet_name)
    candidates = []
    for idx, name in enumerate(wb.sheetnames):
        if name == current_sheet_name:
            continue
        try:
            _find_header(wb[name])
        except Exception:
            continue
        key = _sheet_sort_key(name)
        if key is not None:
            candidates.append((key, idx, name))
    if current_key is not None:
        previous = [item for item in candidates if item[0] < current_key]
        if previous:
            return wb[sorted(previous, key=lambda item: item[0])[-1][2]]
    current_idx = wb.sheetnames.index(current_sheet_name) if current_sheet_name in wb.sheetnames else -1
    previous_by_order = [item for item in candidates if item[1] < current_idx]
    if previous_by_order:
        return wb[sorted(previous_by_order, key=lambda item: item[1])[-1][2]]
    return None


def _valid_month_sheet_candidates(wb) -> list[tuple[tuple[int, int] | None, int, str]]:
    candidates = []
    for idx, name in enumerate(wb.sheetnames):
        try:
            rows, _meta = _extract_rows(wb[name])
        except Exception:
            continue
        if not rows:
            continue
        candidates.append((_sheet_sort_key(name), idx, name))
    return candidates


def _latest_month_sheets(wb, limit: int = 2):
    candidates = _valid_month_sheet_candidates(wb)
    if not candidates:
        return []
    with_keys = [item for item in candidates if item[0] is not None]
    if len(with_keys) >= limit:
        selected = sorted(with_keys, key=lambda item: (item[0][0], item[0][1], item[1]))[-limit:]
    else:
        selected = sorted(candidates, key=lambda item: item[1])[-limit:]
    return [wb[name] for _key, _idx, name in selected]


def _platform_from_url(url: str) -> str:
    value = str(url or "").strip()
    low = value.lower()
    mapping = [
        ("youtube.com", "YouTube"),
        ("youtu.be", "YouTube"),
        ("rutube.ru", "Rutube"),
        ("vk.com", "ВКонтакте"),
        ("vk.ru", "ВКонтакте"),
        ("t.me", "Telegram"),
        ("telegram", "Telegram"),
        ("woman.ru", "Woman.ru"),
        ("dzen.ru", "Дзен"),
        ("otzovik.com", "Отзовик"),
        ("irecommend.ru", "IRecommend"),
        ("ozon.ru", "Ozon"),
        ("wildberries.ru", "Wildberries"),
        ("megapteka.ru", "Мегаптека"),
        ("apteka.ru", "Apteka.ru"),
        ("uteka.ru", "Ютека"),
        ("market.yandex", "Яндекс.Маркет"),
        ("rigla.ru", "Ригла"),
    ]
    for needle, label in mapping:
        if needle in low:
            return label
    match = re.search(r"https?://(?:www\.)?([^/]+)", low)
    if match:
        return match.group(1)
    return value[:40] or "Не указано"


def _valid_url(value: str) -> str:
    text = str(value or "").strip()
    return text if text.lower().startswith(("http://", "https://")) else ""


def _material_group(placement_type: str) -> str:
    text = str(placement_type or "").strip().lower()
    if "отзыв" in text:
        return "Отзывы"
    if "комментар" in text or "упомин" in text or "поддерж" in text or "реагир" in text or "мониторинг" in text or "обсужден" in text:
        return "Комментарии"
    return placement_type or "Другое"


def _section_placement_type(text: str) -> str:
    low = str(text or "").strip().lower().replace("ё", "е")
    if "отзыв" in low:
        return "Отзывы"
    if "топ" in low or "выдач" in low or "serm" in low:
        return "Комментарии на площадках в топ-20"
    if "реагирован" in low or "мониторинг" in low or "обсужд" in low or "встраив" in low:
        return "Комментарии в открытых обсуждениях"
    return ""


def _status_group(status: str) -> str:
    text = str(status or "").strip()
    return text or "Без статуса"


def _messages_word(value: int | float) -> str:
    number = abs(int(round(float(value))))
    if 11 <= number % 100 <= 14:
        return "материалов"
    last = number % 10
    if last == 1:
        return "материал"
    if 2 <= last <= 4:
        return "материала"
    return "материалов"


def _fact_group_counts(fact_by_type: dict) -> tuple[int, int]:
    reviews = 0
    comments = 0
    for name, value in (fact_by_type or {}).items():
        count = _clean_int(value)
        low = str(name or "").lower().replace("ё", "е")
        if "отзыв" in low:
            reviews += count
        elif "коммент" in low or "обсужд" in low or "топ" in low or "serm" in low:
            comments += count
    return reviews, comments


def _build_insight(data: dict) -> str:
    total = int(data.get("total_materials", 0) or 0)
    reviews = int(data.get("review_materials", 0) or 0)
    comments = int(data.get("comment_materials", 0) or 0)
    status_rows = data.get("status_rows") or []
    status_text = status_rows[0]["status"].lower() if len(status_rows) == 1 else "разные статусы"
    top_platforms = [row["platform"] for row in (data.get("platform_rows") or [])[:3]]
    platform_text = ", ".join(top_platforms) if top_platforms else "ключевые площадки"
    views_available = bool(data.get("views_metrics", {}).get("has_views"))
    if views_available:
        caveat = data["views_metrics"].get("metric_caveat") or ""
        views_part = f"По полю «{data['views_metrics'].get('metric_label', 'Просмотры')}» зафиксировано {data['views_metrics']['selected_views_total_formatted']} просмотров."
        if caveat:
            views_part += f" {caveat}"
    else:
        views_part = "Поля просмотров в проектной таблице не заполнены, поэтому метрика просмотров не выводится как факт."
    return (
        f"В проектной таблице за {data.get('month_genitive', 'отчетный период')} по бренду {data.get('project_brand') or 'бренд'} учтено {total} "
        f"{_messages_word(total)}: {reviews} отзывов и {comments} комментариев. "
        f"Все найденные материалы имеют статус «{status_text}». "
        f"Основные точки размещения: {platform_text}. {views_part}"
    )


def _extract_rows(ws) -> tuple[list[dict], dict]:
    header_row, headers = _find_header(ws)
    type_idx = _header_index(headers, "Тип размещения")
    platform_idx = _header_index(headers, "Площадка")
    product_idx = _header_index(headers, "Продукт")
    link_idx = _header_index(headers, "Ссылка на сообщение", "Ссылка")
    topic_idx = _header_index(headers, "Тема")
    text_idx = _header_index(headers, "Текст", "Текст сообщения")
    status_idx = _header_index(headers, "Согласование")
    date_idx = _header_index(headers, "Дата", "Дата публикации")
    author_idx = _header_index(headers, "Автор", "Ник")
    start_views_idx = _header_index(headers, "Просмотры темы на старте")
    received_views_idx = _views_received_index(headers)
    end_views_idx = _header_index(headers, "Просмотры в конце месяца")
    engagement_idx = _header_index(headers, "Вовлечение")
    post_type_idx = _header_index(headers, "Тип поста")

    rows = []
    current_section = ""
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = list(row)
        first_cell = _cell(values, 0)
        platform_raw = _cell(values, platform_idx)
        link = _cell(values, link_idx)
        topic = _cell(values, topic_idx)
        text = _cell(values, text_idx)
        date = _cell(values, date_idx)
        author = _cell(values, author_idx)
        post_type = _cell(values, post_type_idx)
        section = _section_placement_type(first_cell)
        if section and not any([link, topic, text, date, author, post_type]):
            current_section = section
            continue

        placement_type = _cell(values, type_idx) or current_section
        if not placement_type:
            continue
        if placement_type.strip().lower() == "тип размещения":
            continue
        if text and not any([platform_raw, link, topic, date, author, post_type]):
            continue
        message_link = _valid_url(link) or _valid_url(topic) or _valid_url(platform_raw)
        if not any([platform_raw, link, topic, text, message_link]):
            continue
        approval_status = _cell(values, status_idx) or ("Опубликовано" if message_link or text else "")
        rows.append({
            "row": row_num,
            "placement_type": placement_type,
            "material_group": _material_group(placement_type),
            "platform_raw": platform_raw,
            "platform": _platform_from_url(message_link or platform_raw),
            "product": _cell(values, product_idx),
            "message_link": message_link,
            "text": text,
            "text_length": len(text),
            "approval_status": _status_group(approval_status),
            "date": date,
            "author": author,
            "views_at_start": _clean_int(values[start_views_idx] if start_views_idx is not None and start_views_idx < len(values) else None),
            "received_views": _clean_int(values[received_views_idx] if received_views_idx is not None and received_views_idx < len(values) else None),
            "views_at_end_of_month": _clean_int(values[end_views_idx] if end_views_idx is not None and end_views_idx < len(values) else None),
            "engagement": _clean_int(values[engagement_idx] if engagement_idx is not None and engagement_idx < len(values) else None),
            "post_type": post_type,
        })
    return rows, {
        "header_row": header_row,
        "headers": headers,
        "received_views_label": "Просмотры" if received_views_idx is not None else "",
    }


def _choose_views_metric(rows: list[dict], preferred_key: str = "") -> dict:
    item = {"key": "received_views", "label": "Просмотры", "kind": "received"}
    total = sum(int(row.get(item["key"], 0) or 0) for row in rows)
    if total > 0:
        return {**item, "total": int(total), "available": True}
    return {**item, "total": 0, "available": False}


def extract_slide08_seeding_metrics(
    *,
    project_orm_path: Path,
    project_brand: str,
    month_sheet: str = "Май",
    period: str = "",
    source_system: str = "проектная таблица ORM",
) -> dict:
    wb = load_workbook(project_orm_path, data_only=True, read_only=False)
    explicit_month = str(month_sheet or "").strip()
    if explicit_month:
        ws, sheet_name = _month_sheet(wb, explicit_month)
        previous_ws = _previous_month_sheet(wb, sheet_name)
        latest_sheets = [item for item in [previous_ws, ws] if item is not None]
    else:
        latest_sheets = _latest_month_sheets(wb, limit=2)
        if latest_sheets:
            ws = latest_sheets[-1]
            sheet_name = ws.title
        else:
            ws, sheet_name = _month_sheet(wb, month_sheet)
    rows, sheet_meta = _extract_rows(ws)
    fact_data = {}
    try:
        fact_data = extract_project_month_facts(
            project_orm_path,
            month_sheet=sheet_name,
            period=period,
            source_system=source_system,
        )
    except Exception as exc:
        fact_data = {"status": "error", "error": str(exc)}
    fact_by_type = fact_data.get("campaign_publications_by_type") or fact_data.get("fact_rows_by_name") or {}
    authoritative_total = int(fact_data.get("total_fact", 0) or 0)
    if authoritative_total <= 0:
        authoritative_total = len(rows)
    fact_reviews, fact_comments = _fact_group_counts(fact_by_type)
    if not fact_reviews:
        fact_reviews = int(Counter(row["material_group"] for row in rows).get("Отзывы", 0))
    if not fact_comments:
        fact_comments = max(authoritative_total - fact_reviews, 0)
    selected_views = _choose_views_metric(rows)
    selected_views["label"] = "Просмотры"
    views_key = selected_views["key"]

    comparison_sheets = latest_sheets or [ws]

    by_type = Counter(row["placement_type"] for row in rows)
    by_group = Counter(row["material_group"] for row in rows)
    by_status = Counter(row["approval_status"] for row in rows)
    by_post_type = Counter(row["post_type"] or "Не указано" for row in rows)
    platform_counts: Counter[str] = Counter()
    platform_views: defaultdict[str, int] = defaultdict(int)
    platform_engagement: defaultdict[str, int] = defaultdict(int)
    platform_links: dict[str, str] = {}
    for row in rows:
        platform = row["platform"]
        platform_counts[platform] += 1
        platform_views[platform] += int(row.get(views_key, 0) or 0)
        platform_engagement[platform] += int(row["engagement"] or 0)
        platform_links.setdefault(platform, row.get("message_link") or row.get("platform_raw") or "")

    platform_rows = []
    if selected_views.get("available"):
        selected_platforms = sorted(platform_counts.items(), key=lambda item: (platform_views[item[0]], item[1]), reverse=True)[:9]
    else:
        selected_platforms = platform_counts.most_common(9)
    selected_names = {platform for platform, _count in selected_platforms}
    other_count = sum(count for platform, count in platform_counts.items() if platform not in selected_names)
    platform_items = list(selected_platforms)
    if other_count:
        platform_items.append(("Другие", other_count))
    for platform, count in platform_items:
        if platform == "Другие":
            other_names = {name for name in platform_counts if name not in selected_names}
            received_views = sum(platform_views[name] for name in other_names)
            engagement = sum(platform_engagement[name] for name in other_names)
            sample_link = ""
        else:
            received_views = int(platform_views[platform])
            engagement = int(platform_engagement[platform])
            sample_link = platform_links.get(platform, "")
        platform_rows.append({
            "platform": platform,
            "materials": int(count),
            "received_views": int(received_views),
            "engagement": int(engagement),
            "share": count / len(rows) if rows else 0,
            "sample_link": sample_link,
        })

    total_views = int(selected_views.get("total", 0) or 0)
    total_engagement = sum(int(row["engagement"] or 0) for row in rows)
    chart_rows = []
    for chart_ws in comparison_sheets:
        chart_rows_data, _chart_meta = _extract_rows(chart_ws)
        chart_views = _choose_views_metric(chart_rows_data, preferred_key=views_key)
        chart_rows.append({
            "label": _sheet_month_label(chart_ws.title),
            "sheet": chart_ws.title,
            "value": int(chart_views.get("total", 0) or 0),
        })
    raw_received_views_total = sum(int(row.get("received_views", 0) or 0) for row in rows)
    raw_end_views_total = sum(int(row.get("views_at_end_of_month", 0) or 0) for row in rows)
    raw_start_views_total = sum(int(row.get("views_at_start", 0) or 0) for row in rows)
    metric_caveat = ""
    if total_views <= 0:
        metric_caveat = "Поля «Просмотров получено» или «Просмотры» в выбранном листе не заполнены."
    data = {
        "source_file": project_orm_path.name,
        "source_system": source_system,
        "source_label": "проектная таблица ORM",
        "project_brand": project_brand,
        "period_raw": period,
        "sheet": sheet_name,
        "header_row": sheet_meta["header_row"],
        "month_prepositional": _month_prepositional_phrase(period, sheet_name),
        "month_genitive": _month_genitive_phrase(period, sheet_name),
        "total_materials": authoritative_total,
        "review_materials": fact_reviews,
        "comment_materials": fact_comments,
        "tabular_rows_with_platform_data": len(rows),
        "campaign_fact_summary": fact_data,
        "material_type_rows": [
            {"type": str(key), "materials": int(value), "share": int(value) / authoritative_total if authoritative_total else 0}
            for key, value in (fact_by_type or dict(by_type.most_common())).items()
        ],
        "material_group_rows": [
            {"group": "Отзывы", "materials": int(fact_reviews), "share": fact_reviews / authoritative_total if authoritative_total else 0},
            {"group": "Комментарии", "materials": int(fact_comments), "share": fact_comments / authoritative_total if authoritative_total else 0},
        ],
        "status_rows": [
            {"status": key, "materials": int(value), "share": value / len(rows) if rows else 0}
            for key, value in by_status.most_common()
        ],
        "post_type_rows": [
            {"post_type": key, "materials": int(value), "share": value / len(rows) if rows else 0}
            for key, value in by_post_type.most_common()
        ],
        "platform_rows": platform_rows,
        "views_metrics": {
            "received_views_total": total_views,
            "received_views_total_formatted": f"{total_views:,}".replace(",", " "),
            "selected_views_total": total_views,
            "selected_views_total_formatted": f"{total_views:,}".replace(",", " "),
            "metric_key": views_key,
            "metric_label": selected_views["label"],
            "metric_kind": selected_views["kind"],
            "metric_caveat": metric_caveat,
            "has_views": total_views > 0,
            "engagement_total": total_engagement,
            "rows_with_selected_views": sum(1 for row in rows if int(row.get(views_key, 0) or 0) > 0),
            "rows_with_received_views": sum(1 for row in rows if int(row.get("received_views", 0) or 0) > 0),
            "rows_with_engagement": sum(1 for row in rows if int(row["engagement"] or 0) > 0),
            "has_received_views": raw_received_views_total > 0,
            "raw_received_views_total": raw_received_views_total,
            "raw_end_views_total": raw_end_views_total,
            "raw_start_views_total": raw_start_views_total,
            "missing_views_note": "Поля просмотров в выбранном листе не заполнены." if total_views <= 0 else "",
        },
        "views_chart": {
            "metric_key": views_key,
            "metric_label": selected_views["label"],
            "comparison_mode": "month_over_month" if selected_views["kind"] == "received" else "current_only_with_caveat",
            "rows": chart_rows,
            "latest_month_sheets": [chart_ws.title for chart_ws in comparison_sheets],
        },
        "sample_rows": rows[:12],
        "methodology": {
            "material_filter": "Rows below the header with placement type and either platform/link or text, excluding section label rows.",
            "review_count": "Rows whose placement type contains 'отзыв'.",
            "comment_count": "Rows whose placement type contains 'комментар'.",
            "platform": "Normalized from placement URL/domain in the 'Площадка' column.",
            "views": "Uses received views only and labels the client-facing metric as 'Просмотры'. Views are not estimated from start/end counters.",
            "status": "Grouped by 'Согласование'.",
        },
    }
    data["insight_text"] = _build_insight(data)
    data["methodology_status"] = "ready" if rows else "blocked"
    return data
