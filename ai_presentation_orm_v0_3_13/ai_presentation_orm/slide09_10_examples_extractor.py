from __future__ import annotations

from collections import Counter
from pathlib import Path
import re

from openpyxl import load_workbook


MONTH_PREPOSITIONAL = [
    ("январ", "в январе"),
    ("феврал", "в феврале"),
    ("март", "в марте"),
    ("апрел", "в апреле"),
    ("ма", "в мае"),
    ("июн", "в июне"),
    ("июл", "в июле"),
    ("август", "в августе"),
    ("сентябр", "в сентябре"),
    ("октябр", "в октябре"),
    ("ноябр", "в ноябре"),
    ("декабр", "в декабре"),
]

COMMENT_SCREENSHOT_PLATFORM_PRIORITY = [
    "mom.life",
    "forum.baby.ru",
    "babyblog.ru",
    "Дзен",
    "YouTube",
    "Telegram",
    "ВКонтакте",
]

REVIEW_SCREENSHOT_PLATFORM_PRIORITY = [
    "Мегаптека",
    "vseotzyvy.ru",
    "otzyv.pro",
    "IRecommend",
    "Отзовик",
    "Ютека",
    "Apteka.ru",
]


def _cell(values: list, idx: int | None) -> str:
    if idx is None or idx >= len(values):
        return ""
    return str(values[idx] if values[idx] is not None else "").strip()


def _header_index(headers: list[str], *names: str) -> int | None:
    lowered = [str(item or "").strip().lower() for item in headers]
    for name in names:
        needle = name.lower()
        for idx, header in enumerate(lowered):
            if header == needle:
                return idx
    for name in names:
        needle = name.lower()
        for idx, header in enumerate(lowered):
            if needle in header:
                return idx
    return None


def _find_header(ws) -> tuple[int, list[str]]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
        headers = [str(value or "").strip() for value in row]
        lowered = [item.lower() for item in headers]
        has_type = any("тип размещения" in header or header == "формат" for header in lowered)
        has_platform = any("площадка" in header or "ссылка" in header for header in lowered)
        has_text = any(header == "текст" or "текст сообщ" in header or "коммент" in header for header in lowered)
        has_generic_month_table = has_platform and has_text and any(
            header == "дата" or header == "ник" or "просмотр" in header or header == "тип поста"
            for header in lowered
        )
        if (has_type and has_platform and has_text) or has_generic_month_table:
            return row_idx, headers
    raise ValueError("Project table header not found on the month sheet")


def _month_sheet(wb, month_sheet: str):
    if month_sheet and month_sheet in wb.sheetnames:
        return wb[month_sheet], month_sheet
    for name in wb.sheetnames:
        try:
            _find_header(wb[name])
            return wb[name], name
        except Exception:
            continue
    return wb[wb.sheetnames[0]], wb.sheetnames[0]


def _month_phrase(period: str, sheet_name: str) -> str:
    text = f"{period} {sheet_name}".lower().replace("ё", "е")
    for stem, phrase in MONTH_PREPOSITIONAL:
        if stem in text:
            return phrase
    return "в отчетном периоде"


def _platform_from_url(url: str) -> str:
    value = str(url or "").strip()
    low = value.lower()
    mapping = [
        ("youtube.com", "YouTube"),
        ("youtu.be", "YouTube"),
        ("rutube.ru", "Rutube"),
        ("vk.com", "ВКонтакте"),
        ("vk.ru", "ВКонтакте"),
        ("t.me", "Telegram"),
        ("telegram", "Telegram"),
        ("woman.ru", "Woman.ru"),
        ("dzen.ru", "Дзен"),
        ("otzovik.com", "Отзовик"),
        ("irecommend.ru", "IRecommend"),
        ("ozon.ru", "Ozon"),
        ("wildberries.ru", "Wildberries"),
        ("megapteka.ru", "Мегаптека"),
        ("apteka.ru", "Apteka.ru"),
        ("uteka.ru", "Ютека"),
        ("rigla.ru", "Ригла"),
    ]
    for needle, label in mapping:
        if needle in low:
            return label
    match = re.search(r"https?://(?:www\.)?([^/]+)", low)
    if match:
        return match.group(1)
    return value[:40] or "Не указано"


def _valid_url(value: str) -> str:
    text = str(value or "").strip()
    return text if text.lower().startswith(("http://", "https://")) else ""


def _screenshot_link_index(headers: list[str]) -> int | None:
    idx = _header_index(headers, "Ссылка на скриншот", "Скриншот", "Комментарий: разное")
    if idx is not None:
        return idx
    if len(headers) > 18:
        return 18
    return 9 if len(headers) > 9 else None


def _clean_text(value: str) -> str:
    text = str(value or "").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.replace(";", ",")


def _material_group(placement_type: str) -> str:
    text = str(placement_type or "").strip().lower()
    if "отзыв" in text:
        return "Отзывы"
    if "комментар" in text or "упомин" in text or "поддерж" in text or "реагир" in text or "мониторинг" in text or "обсужден" in text:
        return "Комментарии"
    return placement_type or "Другое"


def _section_placement_type(text: str) -> str:
    low = str(text or "").strip().lower().replace("ё", "е")
    if "отзыв" in low:
        return "Отзывы"
    if "топ" in low or "выдач" in low or "serm" in low:
        return "Комментарии на площадках в топ-20"
    if "реагирован" in low or "мониторинг" in low or "обсужд" in low or "встраив" in low:
        return "Комментарии в открытых обсуждениях"
    return ""


def _ru_word(value: int | float, one: str, few: str, many: str) -> str:
    number = abs(int(round(float(value))))
    if 11 <= number % 100 <= 14:
        return many
    last = number % 10
    if last == 1:
        return one
    if 2 <= last <= 4:
        return few
    return many


def _topic_from_text(text: str) -> str:
    lowered = text.lower()
    if any(token in lowered for token in ["врач", "педиатр", "рекомендов"]):
        return "Обсуждение рекомендаций"
    if any(token in lowered for token in ["помог", "эффект", "работал", "прошел", "прошёл"]):
        return "Обсуждение эффективности"
    if any(token in lowered for token in ["ребен", "ребён", "дет", "малыш"]):
        return "Обсуждение опыта применения"
    return "Обсуждение пользовательского опыта"


def _title_from_text(text: str, fallback: str) -> str:
    clean = _clean_text(text)
    first = re.split(r"(?<=[.!?])\s+", clean, maxsplit=1)[0]
    if len(first) < 18:
        first = clean[:80]
    return first[:86].rstrip(" ,.-") or fallback


def _pick_by_rows(rows: list[dict], row_numbers: list[int], limit: int) -> list[dict]:
    by_row = {int(row.get("row", 0)): row for row in rows}
    selected: list[dict] = []
    for row_number in row_numbers:
        row = by_row.get(row_number)
        if row and row not in selected:
            selected.append(row)
        if len(selected) >= limit:
            return selected
    return selected


def _fill_by_platform(selected: list[dict], candidates: list[dict], platforms: list[str], limit: int) -> list[dict]:
    used = {row["row"] for row in selected}
    for platform in platforms:
        for row in candidates:
            if row["row"] in used or row.get("platform") != platform:
                continue
            selected.append(row)
            used.add(row["row"])
            break
        if len(selected) >= limit:
            return selected
    for row in candidates:
        if row["row"] not in used:
            selected.append(row)
            used.add(row["row"])
        if len(selected) >= limit:
            return selected
    return selected


def _join_platforms(rows: list[dict], fallback: str) -> str:
    labels = []
    for row in rows:
        label = str(row.get("platform") or "").strip()
        if label and label not in labels:
            labels.append(label)
    if not labels:
        return fallback
    if len(labels) == 1:
        return labels[0]
    return ", ".join(labels[:-1]) + " и " + labels[-1]


def _row_score(row: dict) -> tuple[int, int, int]:
    text_len = len(str(row.get("text") or ""))
    readable_len = min(text_len, 900)
    has_link = 1 if row.get("message_link") else 0
    return has_link, readable_len, -abs(text_len - 450)


def _select_diverse_rows(rows: list[dict], limit: int, exclude_rows: set[int] | None = None) -> list[dict]:
    excluded = exclude_rows or set()
    candidates = [row for row in rows if row.get("text") and int(row.get("row", 0) or 0) not in excluded]
    selected: list[dict] = []
    used_rows: set[int] = set()
    platforms = [platform for platform, _count in Counter(row.get("platform") for row in candidates).most_common()]
    for platform in platforms:
        platform_rows = [row for row in candidates if row.get("platform") == platform and int(row.get("row", 0) or 0) not in used_rows]
        if not platform_rows:
            continue
        row = max(platform_rows, key=_row_score)
        selected.append(row)
        used_rows.add(int(row.get("row", 0) or 0))
        if len(selected) >= limit:
            return selected
    for row in sorted(candidates, key=_row_score, reverse=True):
        row_id = int(row.get("row", 0) or 0)
        if row_id in used_rows:
            continue
        selected.append(row)
        used_rows.add(row_id)
        if len(selected) >= limit:
            break
    return selected


def _screenshot_row_score(row: dict, platform_priority: list[str]) -> tuple[int, int, int, int, int, int, int, int]:
    text_len = len(str(row.get("text") or ""))
    text_low = str(row.get("text") or "").lower()
    link = str(row.get("message_link") or "").lower()
    platform = str(row.get("platform") or "")
    try:
        priority = len(platform_priority) - platform_priority.index(platform)
    except ValueError:
        priority = 0
    has_link = 1 if link.startswith(("http://", "https://")) else 0
    good_length = 1 if 70 <= text_len <= 650 else 0
    direct_page = 1 if any(token in link for token in ["review", "reviews", "comments", "goto", "post", "content"]) else 0
    clean_platform_page = 1 if platform == "otzyv.pro" and "/category/" in link else 0
    context_signal = 1 if any(token in text_low for token in ["проб", "приним", "курс", "реб", "врач", "аптек", "отзыв", "помог", "вопрос", "опыт"]) else 0
    length_fit = -abs(text_len - 280)
    not_too_long = 1 if text_len <= 950 else 0
    return has_link, priority, context_signal, clean_platform_page, good_length, direct_page, not_too_long, length_fit


def _select_screenshot_rows(
    rows: list[dict],
    limit: int,
    platform_priority: list[str],
    *,
    max_per_platform: int = 3,
    exclude_rows: set[int] | None = None,
) -> list[dict]:
    excluded = exclude_rows or set()
    candidates = [row for row in rows if row.get("text") and row.get("message_link") and int(row.get("row", 0) or 0) not in excluded]
    if not candidates:
        return []

    grouped: dict[str, list[dict]] = {}
    for row in sorted(candidates, key=lambda item: _screenshot_row_score(item, platform_priority), reverse=True):
        grouped.setdefault(str(row.get("platform") or ""), []).append(row)

    def platform_key(platform: str) -> tuple:
        top_row = grouped[platform][0]
        return _screenshot_row_score(top_row, platform_priority)

    platform_order = sorted(grouped, key=platform_key, reverse=True)
    effective_cap = max_per_platform
    selected: list[dict] = []
    used_rows: set[int] = set()
    platform_counts: Counter = Counter()

    while len(selected) < limit:
        changed = False
        for platform in platform_order:
            if platform_counts[platform] >= effective_cap:
                continue
            if selected and str(selected[-1].get("platform") or "") == platform and len(platform_order) > 1:
                continue
            for row in grouped[platform]:
                row_id = int(row.get("row", 0) or 0)
                if row_id in used_rows:
                    continue
                selected.append(row)
                used_rows.add(row_id)
                platform_counts[platform] += 1
                changed = True
                break
            if len(selected) >= limit:
                return selected
        if not changed:
            break

    for row in sorted(candidates, key=lambda item: _screenshot_row_score(item, platform_priority), reverse=True):
        row_id = int(row.get("row", 0) or 0)
        if row_id in used_rows:
            continue
        selected.append(row)
        used_rows.add(row_id)
        if len(selected) >= limit:
            break
    return selected


def _comment_cards(comment_rows: list[dict]) -> list[dict]:
    selected = _select_screenshot_rows(comment_rows, 18, COMMENT_SCREENSHOT_PLATFORM_PRIORITY, max_per_platform=4)
    cards = []
    for row in selected:
        cards.append({
            "kind": "social_comment",
            "row": row["row"],
            "platform": row["platform"],
            "placement_type": row["placement_type"],
            "context": _topic_from_text(row["text"]),
            "title": "Комментарий в обсуждении",
            "text": row["text"],
            "link": row["message_link"],
            "screenshot_link": row.get("screenshot_link", ""),
            "date": row.get("date", ""),
            "nick": row.get("nick", ""),
            "author": row.get("author", ""),
        })
    return cards


def _review_cards(review_rows: list[dict], project_brand: str) -> list[dict]:
    selected = _select_screenshot_rows(review_rows, 10, REVIEW_SCREENSHOT_PLATFORM_PRIORITY, max_per_platform=2)
    cards = []
    for row in selected:
        text = row.get("text", "")
        cards.append({
            "kind": "review_site",
            "row": row.get("row", 0),
            "platform": row.get("platform") or "отзывная площадка",
            "product": row.get("product") or project_brand,
            "title": _title_from_text(text, f"Отзыв о {project_brand}"),
            "text": text,
            "link": row.get("message_link", ""),
            "screenshot_link": row.get("screenshot_link", ""),
            "date": row.get("date", ""),
            "nick": row.get("nick", ""),
            "author": row.get("author", ""),
        })
    return cards


def _marketplace_cluster(rows: list[dict], project_brand: str) -> dict:
    selected = _select_diverse_rows(rows, 3)
    items = []
    for row in selected[:3]:
        items.append({
            "row": row["row"],
            "platform": row["platform"],
            "product": row.get("product") or project_brand,
            "text": row["text"],
            "link": row["message_link"],
            "screenshot_link": row.get("screenshot_link", ""),
        })
    has_purchase_reviews = any("покуп" in str(row.get("placement_type") or "").lower() for row in selected)
    return {
        "kind": "marketplace_reviews",
        "title": "Отзывы с покупкой" if has_purchase_reviews else "Отзывы на площадках",
        "platform": _join_platforms(selected, "отзывные площадки"),
        "items": items,
    }


def _review_site_card(rows: list[dict], project_brand: str, exclude_rows: set[int] | None = None) -> dict:
    selected = _select_diverse_rows(rows, 1, exclude_rows=exclude_rows)
    if not selected:
        selected = _select_diverse_rows(rows, 1)
    row = selected[0] if selected else {}
    text = row.get("text", "")
    return {
        "kind": "review_site",
        "row": row.get("row", 0),
        "platform": row.get("platform") or "отзывная площадка",
        "product": row.get("product") or project_brand,
        "title": _title_from_text(text, f"Отзыв о {project_brand}"),
        "text": text,
        "link": row.get("message_link", ""),
        "screenshot_link": row.get("screenshot_link", ""),
    }


def extract_slide09_10_examples(
    *,
    project_orm_path: Path,
    project_brand: str,
    month_sheet: str = "Май",
    period: str = "",
    source_system: str = "проектная таблица ORM",
) -> dict:
    wb = load_workbook(project_orm_path, data_only=True, read_only=False)
    ws, sheet_name = _month_sheet(wb, month_sheet)
    header_row, headers = _find_header(ws)
    type_idx = _header_index(headers, "Тип размещения")
    platform_idx = _header_index(headers, "Площадка")
    product_idx = _header_index(headers, "Продукт")
    link_idx = _header_index(headers, "Ссылка на сообщение", "Ссылка")
    topic_idx = _header_index(headers, "Тема")
    text_idx = _header_index(headers, "Текст", "Текст сообщения")
    date_idx = _header_index(headers, "Дата")
    nick_idx = _header_index(headers, "Ник")
    author_idx = _header_index(headers, "Автор")
    screenshot_idx = _screenshot_link_index(headers)

    rows = []
    current_section = ""
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = list(row)
        first_cell = _cell(values, 0)
        platform_raw = _cell(values, platform_idx)
        link = _cell(values, link_idx)
        topic = _cell(values, topic_idx)
        text = _clean_text(_cell(values, text_idx))
        date = _cell(values, date_idx)
        nick = _cell(values, nick_idx)
        author = _cell(values, author_idx)
        section = _section_placement_type(first_cell)
        if section and not any([link, topic, text, date, nick, author]):
            current_section = section
            continue

        placement_type = _cell(values, type_idx) or current_section
        if not placement_type:
            continue
        message_link = _valid_url(link) or _valid_url(topic) or _valid_url(platform_raw)
        if text and not any([platform_raw, link, topic, date, nick, author, message_link]):
            continue
        if not any([platform_raw, link, topic, text, message_link]):
            continue
        rows.append({
            "row": row_num,
            "placement_type": placement_type,
            "material_group": _material_group(placement_type),
            "platform_raw": platform_raw,
            "platform": _platform_from_url(message_link or platform_raw),
            "product": _cell(values, product_idx),
            "message_link": message_link,
            "screenshot_link": _valid_url(_cell(values, screenshot_idx)),
            "text": text,
            "date": date,
            "nick": nick,
            "author": author,
        })

    all_comment_rows = [row for row in rows if row["material_group"] == "Комментарии"]
    all_review_rows = [row for row in rows if row["material_group"] == "Отзывы"]
    comment_rows = [row for row in all_comment_rows if row.get("text")]
    review_rows = [row for row in all_review_rows if row.get("text")]
    purchased_review_rows = [row for row in review_rows if "покуп" in row.get("placement_type", "").lower()]
    review_site_rows = [row for row in review_rows if row not in purchased_review_rows]

    comment_platforms = Counter(row["platform"] for row in all_comment_rows)
    review_platforms = Counter(row["platform"] for row in all_review_rows)
    slide09_cards = _comment_cards(comment_rows)
    slide10_cards = _review_cards(review_rows, project_brand)
    month_phrase = _month_phrase(period, sheet_name)

    slide09 = {
        "slide": 9,
        "source_system": source_system,
        "source_label": "проектная таблица ORM",
        "project_brand": project_brand,
        "period_raw": period,
        "sheet": sheet_name,
        "title": "Примеры сообщений о бренде",
        "comment_count": len(all_comment_rows),
        "summary": (
            f"{month_phrase.capitalize()} в проектной таблице учтено {len(all_comment_rows)} "
            f"{_ru_word(len(all_comment_rows), 'комментарий', 'комментария', 'комментариев')} о бренде {project_brand} "
            "в обсуждениях, социальных сетях и материалах в топе выдачи. Примеры показывают, как авторы "
            "описывали пользовательский опыт, рекомендации, симптомы и ситуации применения продукта."
        ),
        "cards": slide09_cards,
        "platform_rows": [{"platform": key, "materials": int(value)} for key, value in comment_platforms.most_common()],
        "sample_rows": comment_rows[:12],
        "methodology_status": "ready" if slide09_cards else "blocked",
        "methodology": {
            "material_filter": "Rows from the month sheet whose placement type contains 'комментар'.",
            "examples": "Five deterministic examples selected from real project table rows, prioritizing browser-visible public pages and compact message text.",
        },
    }
    slide10 = {
        "slide": 10,
        "source_system": source_system,
        "source_label": "проектная таблица ORM",
        "project_brand": project_brand,
        "period_raw": period,
        "sheet": sheet_name,
        "title": "Примеры отзывов о бренде",
        "review_count": len(all_review_rows),
        "purchased_review_count": len(purchased_review_rows),
        "summary": (
            f"{month_phrase.capitalize()} в проектной таблице учтено {len(all_review_rows)} "
            f"{_ru_word(len(all_review_rows), 'отзыв', 'отзыва', 'отзывов')} о бренде {project_brand} "
            "на отзывных площадках и маркетплейсах. В примерах авторы описывали опыт курса, "
            "опыт применения, динамику симптомов и удобство формата продукта."
        ),
        "cards": slide10_cards,
        "platform_rows": [{"platform": key, "materials": int(value)} for key, value in review_platforms.most_common()],
        "sample_rows": review_rows[:12],
        "methodology_status": "ready" if review_rows else "blocked",
        "methodology": {
            "material_filter": "Rows from the month sheet whose placement type contains 'отзыв'.",
            "examples": "Five deterministic review examples selected from real project table rows, prioritizing browser-visible public review pages and compact text.",
        },
    }
    return {"slide09": slide09, "slide10": slide10}
