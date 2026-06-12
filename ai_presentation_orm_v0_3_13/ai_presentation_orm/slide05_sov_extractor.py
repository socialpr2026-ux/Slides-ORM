from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from .project_month_fact_extractor import extract_project_month_facts
from .slide03_data_extractor import (
    _aggregate_brand_rows,
    _brand_from_product_name,
    _clean_int,
    _filter_competitive_rows,
    _find_header,
    _find_metrics_sheet,
    _get_index,
    _is_thematic_metric_label,
    _period_label_from_content,
    _period_to_caption,
)
from .xlsx_safe_reader import open_workbook_safe


def _source_label(source_system: str) -> str:
    low = (source_system or "").lower()
    if "brand analytics" in low or low.strip() == "ba":
        return "сервис мониторинга Brand Analytics"
    if "медиалог" in low or "medialog" in low:
        return "сервис мониторинга Медиалогия"
    return "сервис мониторинга Медиалогия"


def _brand_rows(wb, project_brand: str, competitor_brands: list[str] | None = None) -> tuple[list[dict], dict]:
    ws, sheet_name = _find_metrics_sheet(wb)
    header_row, headers = _find_header(ws)
    object_idx = _get_index(headers, "Объект", "Тег", "Продукт", default=1)
    mentions_idx = _get_index(headers, "Сообщения")
    audience_idx = _get_index(headers, "Аудитория")
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        values = list(row)
        raw_brand = values[object_idx] if object_idx is not None and len(values) > object_idx else ""
        brand = str(raw_brand or "").strip()
        if not brand:
            continue
        if str(sheet_name).lower() == "продукты":
            brand = _brand_from_product_name(brand, project_brand)
        if brand.lower() in {"дата", "итого"}:
            break
        if brand.lower() == "всего":
            continue
        if _is_thematic_metric_label(brand):
            continue
        mentions = _clean_int(values[mentions_idx] if mentions_idx is not None and len(values) > mentions_idx else 0)
        if mentions <= 0:
            continue
        rows.append({
            "brand": brand,
            "mentions": mentions,
            "audience": _clean_int(values[audience_idx] if audience_idx is not None and len(values) > audience_idx else 0),
            "sov": 0,
        })
    if str(sheet_name).lower() == "продукты":
        rows = _aggregate_brand_rows(rows)
    rows = _filter_competitive_rows(rows, project_brand, competitor_brands)
    total = sum(item["mentions"] for item in rows)
    for item in rows:
        item["sov"] = item["mentions"] / total if total else 0
    project_row = next((item for item in rows if item["brand"].lower() == project_brand.lower()), {})
    return sorted(rows, key=lambda item: item["mentions"], reverse=True), project_row


def _cell(values: list, idx: int) -> str:
    if idx is None:
        return ""
    return str(values[idx] if idx < len(values) and values[idx] is not None else "").strip()


def _section_placement_type(text: str) -> str:
    low = str(text or "").strip().lower().replace("ё", "е")
    if "отзыв" in low:
        return "Отзывы"
    if "топ" in low or "выдач" in low or "serm" in low:
        return "Комментарии в топе выдачи"
    if "реагирован" in low or "мониторинг" in low or "обсужд" in low or "встраив" in low:
        return "Комментарии в обсуждениях"
    return ""


def _campaign_materials_from_month_sheet(project_orm_path: Path | None, month_sheet: str = "") -> dict:
    if not project_orm_path:
        return {
            "status": "missing_project_table",
            "sheet": month_sheet,
            "count": 0,
            "by_type": {},
            "by_status": {},
            "rows_with_text": 0,
        }
    wb = load_workbook(project_orm_path, data_only=True, read_only=False)
    sheet_name = month_sheet if month_sheet and month_sheet in wb.sheetnames else ""
    if not sheet_name and month_sheet:
        low_requested = month_sheet.lower().replace("ё", "е")
        sheet_name = next(
            (name for name in wb.sheetnames if low_requested in name.lower().replace("ё", "е")),
            "",
        )
    if not sheet_name:
        for candidate in ["Май", "Июнь", "Апрель", "Март"]:
            if candidate in wb.sheetnames:
                sheet_name = candidate
                break
    if not sheet_name:
        return {
            "status": "missing_month_sheet",
            "sheet": month_sheet,
            "count": 0,
            "by_type": {},
            "by_status": {},
            "rows_with_text": 0,
        }

    ws = wb[sheet_name]
    header_row = 0
    type_idx = site_idx = link_idx = topic_idx = text_idx = status_idx = date_idx = post_type_idx = None
    header_candidates = []
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
        vals = [str(v or "").strip() for v in row]
        lowered = [v.lower() for v in vals]
        has_type = "тип размещения" in lowered or "формат" in lowered
        has_platform = any("площадка" in value or "ссылка" in value for value in lowered)
        has_text = any("текст" in value for value in lowered)
        has_generic_month_table = has_platform and has_text and any(
            value == "дата" or value == "ник" or "просмотр" in value or value == "тип поста"
            for value in lowered
        )
        if (has_type and has_text) or has_generic_month_table:
            header_candidates.append((idx, lowered))
    if header_candidates:
        header_row, lowered = sorted(header_candidates, key=lambda item: ("согласование" not in item[1], item[0]))[0]
        type_idx = lowered.index("тип размещения") if "тип размещения" in lowered else lowered.index("формат") if "формат" in lowered else None
        site_idx = lowered.index("площадка") if "площадка" in lowered else None
        link_idx = next((i for i, value in enumerate(lowered) if "ссылка" in value and "скрин" not in value), None)
        topic_idx = lowered.index("тема") if "тема" in lowered else None
        text_idx = next((i for i, value in enumerate(lowered) if "текст" in value), None)
        status_idx = next((i for i, value in enumerate(lowered) if "согласование" in value), None)
        date_idx = next((i for i, value in enumerate(lowered) if "дата" in value), None)
        post_type_idx = lowered.index("тип поста") if "тип поста" in lowered else None
    if not header_row or text_idx is None:
        return {
            "status": "missing_material_columns",
            "sheet": sheet_name,
            "count": 0,
            "by_type": {},
            "by_status": {},
            "rows_with_text": 0,
        }

    by_type: Counter[str] = Counter()
    by_status: Counter[str] = Counter()
    rows_with_text = 0
    rows_with_date = 0
    samples = []
    current_section = ""
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = list(row)
        first_cell = _cell(values, 0)
        site = _cell(values, site_idx) if site_idx is not None else ""
        link = _cell(values, link_idx) if link_idx is not None else ""
        topic = _cell(values, topic_idx) if topic_idx is not None else ""
        text = _cell(values, text_idx)
        status = _cell(values, status_idx) if status_idx is not None else ""
        date = _cell(values, date_idx) if date_idx is not None else ""
        post_type = _cell(values, post_type_idx) if post_type_idx is not None else ""
        section = _section_placement_type(first_cell)
        if section and not any([link, topic, text, status, date, post_type]):
            current_section = section
            continue
        placement_type = _cell(values, type_idx) or current_section
        if not placement_type:
            continue
        if placement_type.strip().lower() == "тип размещения":
            continue
        if text and not any([site, link, topic, status, date, post_type]):
            continue
        if not any([site, link, topic, text]):
            continue
        by_type[placement_type] += 1
        if text:
            rows_with_text += 1
        if status:
            by_status[status] += 1
        if date:
            rows_with_date += 1
        if len(samples) < 8:
            samples.append({
                "row": row_num,
                "type": placement_type,
                "site": site[:120],
                "status": status,
                "date": date,
                "text_length": len(text),
            })

    return {
        "status": "ready",
        "sheet": sheet_name,
        "count": sum(by_type.values()),
        "by_type": dict(by_type),
        "by_status": dict(by_status),
        "rows_with_text": rows_with_text,
        "rows_with_date": rows_with_date,
        "samples": samples,
    }


def _campaign_facts_from_project_month(project_orm_path: Path | None, month_sheet: str = "", period: str = "") -> dict:
    if not project_orm_path:
        return {"status": "missing_project_table", "count": 0, "by_type": {}, "method": ""}
    try:
        facts = extract_project_month_facts(
            project_orm_path,
            month_sheet=month_sheet,
            period=period,
            source_system="проектная таблица ORM",
        )
        count = int(facts.get("campaign_publications_count") or facts.get("total_fact") or 0)
        if count:
            return {
                "status": "ready",
                "sheet": facts.get("sheet", month_sheet),
                "count": count,
                "by_type": facts.get("campaign_publications_by_type") or facts.get("fact_rows_by_name") or {},
                "method": facts.get("campaign_fact_method") or facts.get("method") or "project_month_fact",
                "source": facts,
            }
    except Exception as exc:
        return {"status": "error", "count": 0, "by_type": {}, "method": "", "error": str(exc)}
    return {"status": "missing_required", "count": 0, "by_type": {}, "method": ""}


def _ratings_summary(project_orm_path: Path | None) -> dict:
    if not project_orm_path:
        return {"status": "missing_project_table"}
    wb = load_workbook(project_orm_path, data_only=True, read_only=False)
    if "Рейтинги" not in wb.sheetnames:
        return {"status": "missing_ratings_sheet"}
    ws = wb["Рейтинги"]
    total_start_reviews = 0
    weighted_rating_sum = 0.0
    current_rows_filled = 0
    platforms_with_start = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        values = list(row)
        url = values[0] if values else None
        if not url:
            continue
        start_reviews = _clean_int(values[3] if len(values) > 3 else 0)
        try:
            start_rating = float(values[4] or 0) if len(values) > 4 else 0.0
        except Exception:
            start_rating = 0.0
        if start_reviews or start_rating:
            platforms_with_start += 1
            total_start_reviews += start_reviews
            weighted_rating_sum += start_reviews * start_rating
        if (len(values) > 5 and values[5] is not None) or (len(values) > 6 and values[6] is not None):
            current_rows_filled += 1
    return {
        "status": "ready",
        "platforms_with_start": platforms_with_start,
        "total_start_reviews": total_start_reviews,
        "weighted_start_rating": weighted_rating_sum / total_start_reviews if total_start_reviews else 0,
        "current_rows_filled": current_rows_filled,
    }


def _pie_rows(rows: list[dict], project_brand: str, max_items: int = 6) -> list[dict]:
    selected = rows[:max_items]
    project = next((item for item in rows if item["brand"].lower() == project_brand.lower()), None)
    if project and all(item["brand"] != project["brand"] for item in selected):
        selected = selected[: max_items - 1] + [project]
    selected_names = {item["brand"] for item in selected}
    other_mentions = sum(item["mentions"] for item in rows if item["brand"] not in selected_names)
    total = sum(item["mentions"] for item in rows)
    out = []
    for item in selected:
        out.append({
            "brand": item["brand"],
            "mentions": item["mentions"],
            "sov": item["mentions"] / total if total else 0,
        })
    if other_mentions > 0:
        out.append({"brand": "Прочие", "mentions": other_mentions, "sov": other_mentions / total if total else 0})
    return out


def extract_slide05_sov(
    *,
    analytics_path: Path,
    project_brand: str,
    source_system: str = "",
    project_orm_path: Path | None = None,
    campaign_month_sheet: str = "",
    campaign_materials_override: int | None = None,
    competitor_brands: list[str] | None = None,
) -> dict:
    wb = open_workbook_safe(analytics_path)
    content_ws = wb["Содержание"] if "Содержание" in wb.sheetnames else wb[wb.sheetnames[0]]
    period_raw = _period_label_from_content(content_ws)
    month_prep, month_caption, month_title = _period_to_caption(period_raw)
    rows, project_row = _brand_rows(wb, project_brand, competitor_brands)
    total_mentions = sum(item["mentions"] for item in rows)
    project_mentions = int(project_row.get("mentions", 0) or 0)
    project_sov = project_mentions / total_mentions if total_mentions else 0

    campaign_materials = _campaign_facts_from_project_month(project_orm_path, campaign_month_sheet, period_raw)
    if not int(campaign_materials.get("count", 0) or 0):
        campaign_materials = _campaign_materials_from_month_sheet(project_orm_path, campaign_month_sheet)
    ratings = _ratings_summary(project_orm_path)
    if campaign_materials_override is not None:
        campaign_count = max(int(campaign_materials_override), 0)
        method = "campaign_materials_override"
        method_note = "Тестовая настройка: количество материалов кампании задано вручную."
    else:
        campaign_count = int(campaign_materials.get("count", 0) or 0)
        method = campaign_materials.get("method") or "project_month_sheet_material_rows"
        method_note = "Кампания рассчитана как количество публикаций в месячном листе проектной ORM-таблицы."

    verified_split = campaign_count > 0 and total_mentions > 0
    attributed_campaign_mentions = campaign_count
    organic_mentions = max(project_mentions - attributed_campaign_mentions, 0) if verified_split else project_mentions
    organic_total_mentions = total_mentions - attributed_campaign_mentions
    if organic_total_mentions <= 0:
        organic_total_mentions = total_mentions
    organic_sov = organic_mentions / organic_total_mentions if organic_total_mentions else 0
    lift_pp = (project_sov - organic_sov) * 100 if verified_split else None
    comparison_rows = []
    for row in rows:
        current_mentions = int(row.get("mentions", 0) or 0)
        row_organic_mentions = organic_mentions if row["brand"].lower() == project_brand.lower() else current_mentions
        comparison_rows.append({
            "brand": row["brand"],
            "current_mentions": current_mentions,
            "current_sov": current_mentions / total_mentions if total_mentions else 0,
            "organic_mentions": row_organic_mentions,
            "organic_sov": row_organic_mentions / organic_total_mentions if organic_total_mentions else 0,
        })

    return {
        "source_file": analytics_path.name,
        "source_system": source_system or "Медиалогия",
        "source_label": _source_label(source_system),
        "project_brand": project_brand,
        "period_raw": period_raw,
        "month_prepositional": month_prep,
        "month_caption": month_caption,
        "month_title": month_title,
        "total_mentions_competitive_set": total_mentions,
        "project_brand_row": project_row,
        "project_mentions": project_mentions,
        "project_sov": project_sov,
        "project_rank": next((idx + 1 for idx, item in enumerate(rows) if item["brand"].lower() == project_brand.lower()), None),
        "pie_rows": _pie_rows(rows, project_brand),
        "campaign_split": {
            "status": "verified" if verified_split else "missing_required",
            "method": method if verified_split else "",
            "method_note": method_note if verified_split else "",
            "campaign_materials": campaign_count,
            "campaign_publications_count": campaign_count,
            "campaign_publications_by_type": campaign_materials.get("by_type") or {},
            "campaign_fact_method": method if verified_split else "",
            "campaign_fact_source_sheet": campaign_materials.get("sheet", ""),
            "campaign_required": True,
            "campaign_vs_organic_status": "verified" if verified_split else "missing_required",
            "campaign_materials_override": campaign_materials_override,
            "attributed_campaign_mentions": attributed_campaign_mentions if verified_split else 0,
            "campaign_sov": attributed_campaign_mentions / total_mentions if total_mentions else 0,
            "organic_mentions": organic_mentions,
            "organic_total_mentions": organic_total_mentions,
            "organic_sov": organic_sov,
            "organic_sov_same_denominator": organic_mentions / total_mentions if total_mentions else 0,
            "sov_lift_pp": lift_pp,
            "caveat": (
                "Для настройки слайда использовано заданное количество материалов кампании."
                if campaign_materials_override is not None
                else "Split рассчитан по материалам месячного листа проектной таблицы."
            ) if verified_split else "Нет данных по материалам кампании; причинный вывод не строится.",
        },
        "campaign_materials_summary": campaign_materials,
        "campaign_publications_count": campaign_count,
        "campaign_publications_by_type": campaign_materials.get("by_type") or {},
        "campaign_fact_method": method if verified_split else "",
        "campaign_fact_source_sheet": campaign_materials.get("sheet", ""),
        "campaign_vs_organic_status": "verified" if verified_split else "missing_required",
        "ratings_summary": ratings,
        "sov_comparison_rows": comparison_rows,
        "all_brand_rows": rows,
        "methodology_status": "ready" if verified_split else "ready_with_caveat",
    }
