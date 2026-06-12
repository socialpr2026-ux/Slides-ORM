from __future__ import annotations

from collections import Counter
from pathlib import Path
import re

from openpyxl import load_workbook

from .media_plan_extractor import STANDARD_ROW_ORDER, SERVICE_ROW_NAME, standardize_format_name


def _text(value) -> str:
    return str(value or "").strip()


def _lower(value) -> str:
    return _text(value).lower().replace("ё", "е")


def _clean_int(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return max(int(round(float(value))), 0)
    text = re.sub(r"[^0-9,.-]", "", str(value)).replace(",", ".")
    if not text:
        return 0
    try:
        return max(int(round(float(text))), 0)
    except Exception:
        return 0


def _first_number(text: str) -> int | None:
    cleaned = re.sub(r"топ\s*[-–]?\s*20", "топ", _lower(text))
    match = re.search(r"(?<!\d)(\d[\d\s]*)(?!\d)", cleaned)
    if not match:
        return None
    return int(match.group(1).replace(" ", ""))


def _header_index(headers: list[str], *needles: str) -> int | None:
    lowered = [_lower(item) for item in headers]
    for needle in needles:
        low_needle = _lower(needle)
        for idx, header in enumerate(lowered):
            if header == low_needle:
                return idx
    for needle in needles:
        low_needle = _lower(needle)
        for idx, header in enumerate(lowered):
            if low_needle in header:
                return idx
    return None


def _find_header(ws) -> tuple[int, list[str]]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 40), values_only=True), start=1):
        headers = [_text(value) for value in row]
        lowered = [_lower(value) for value in headers]
        has_type = any("тип размещ" in value or value == "формат" for value in lowered)
        has_platform = any("площадк" in value or "ссылка" in value for value in lowered)
        has_text = any(value == "текст" or "текст сообщ" in value or "коммент" in value for value in lowered)
        has_generic_month_table = has_platform and has_text and any(
            value == "дата" or value == "ник" or "просмотр" in value or value == "тип поста"
            for value in lowered
        )
        if (has_type and has_platform and has_text) or has_generic_month_table:
            return row_idx, headers
    raise ValueError("Project month fact table header not found")


def _sheet_with_project_table(wb, requested: str = ""):
    if requested and requested in wb.sheetnames:
        return wb[requested], requested
    if requested:
        requested_low = _lower(requested)
        for name in wb.sheetnames:
            name_low = _lower(name)
            if requested_low and (requested_low in name_low or name_low in requested_low):
                try:
                    _find_header(wb[name])
                    return wb[name], name
                except Exception:
                    continue
    for name in wb.sheetnames:
        try:
            _find_header(wb[name])
            return wb[name], name
        except Exception:
            continue
    raise ValueError("Project month fact sheet not found")


def _pre_header_summary_text(ws, header_row: int) -> str:
    lines = []
    for row in ws.iter_rows(min_row=1, max_row=max(header_row - 1, 1), values_only=True):
        row_values = [_text(value) for value in row if _text(value)]
        if row_values:
            lines.append("\n".join(row_values))
    return "\n".join(lines)


def _parse_summary_counts(summary_text: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for raw_line in re.split(r"[\r\n]+", summary_text or ""):
        line = raw_line.strip()
        low = _lower(line)
        if not low:
            continue
        if "отзыв" in low:
            plus_match = re.search(r"(?<!\d)(\d+)\s*\+\s*(\d+)(?!\d)", low)
            purchase_terms = any(term in low for term in ["покуп", "выкуп", "аптек", "apteka"])
            if plus_match:
                counts["Отзывы (отзовики, аптеки) на топ-площадках без покупки"] = int(plus_match.group(1))
                counts["Отзывы с покупкой"] = int(plus_match.group(2))
            else:
                value = _first_number(low)
                if value is None:
                    continue
                target = "Отзывы с покупкой" if purchase_terms else "Отзывы (отзовики, аптеки) на топ-площадках без покупки"
                counts[target] = value
            continue
        if "фото" in low and "коммент" in low:
            value = _first_number(low)
            if value is not None:
                counts["Комментарии с нативным фото продукта"] = value
            continue
        if "топ" in low or "serm" in low or "выдач" in low:
            value = _first_number(low)
            if value is not None:
                counts["Комментарии на площадках в топ-20"] = value
            continue
        if "встраив" in low or "форум" in low or "соцсет" in low or "обсужд" in low or "коммент" in low:
            value = _first_number(low)
            if value is not None:
                counts["Комментарии в открытых обсуждениях (симптоматика, диагнозы, препараты)"] = value
    return {key: value for key, value in counts.items() if value is not None}


def _classify_row(placement_type: str, context: str) -> str:
    low_type = _lower(placement_type)
    low_context = _lower(context)
    if "отзыв" in low_type:
        if any(term in low_context for term in ["покуп", "выкуп", "аптек", "apteka"]):
            return "Отзывы с покупкой"
        return "Отзывы (отзовики, аптеки) на топ-площадках без покупки"
    if "фото" in low_context and "коммент" in low_context:
        return "Комментарии с нативным фото продукта"
    if "коммент" in low_type or "коммент" in low_context or "упомин" in low_type or "поддерж" in low_type:
        if any(term in low_context for term in ["топ-20", "топ 20", "топ", "serm", "выдач"]):
            return "Комментарии на площадках в топ-20"
        return "Комментарии в открытых обсуждениях (симптоматика, диагнозы, препараты)"
    return standardize_format_name(placement_type)


def _section_label(text: str) -> str:
    low = _lower(text)
    if "отзыв" in low:
        return STANDARD_ROW_ORDER[0]
    if "топ" in low or "выдач" in low or "serm" in low:
        return STANDARD_ROW_ORDER[2]
    if "реагирован" in low or "мониторинг" in low or "обсужд" in low or "встраив" in low:
        return STANDARD_ROW_ORDER[3]
    return ""


def _is_repeated_header_row(values: list, headers: list[str]) -> bool:
    matches = 0
    meaningful = 0
    for idx, value in enumerate(values[: len(headers)]):
        cell = _lower(value)
        if not cell:
            continue
        meaningful += 1
        header = _lower(headers[idx]) if idx < len(headers) else ""
        if header and (
            cell == header
            or (len(cell) >= 4 and header.startswith(cell))
            or (len(header) >= 4 and cell.startswith(header))
        ):
            matches += 1
    return meaningful >= 2 and matches >= 2


def _extract_tabular_counts(ws, header_row: int, headers: list[str]) -> dict[str, int]:
    type_idx = _header_index(headers, "Тип размещения", "Формат")
    platform_idx = _header_index(headers, "Площадка", "Ссылка")
    link_idx = _header_index(headers, "Ссылка на сообщение", "Ссылка")
    topic_idx = _header_index(headers, "Тема")
    text_idx = _header_index(headers, "Текст", "Текст сообщения")
    photo_idx = _header_index(headers, "Комментарий: фото", "Фото")
    post_type_idx = _header_index(headers, "Тип поста")

    counts: Counter[str] = Counter()
    current_section = ""
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = list(row)
        first_cell = _text(values[0]) if values else ""
        if _is_repeated_header_row(values, headers):
            continue
        placement_type = _text(values[type_idx]) if type_idx is not None and type_idx < len(values) else ""
        platform = _text(values[platform_idx]) if platform_idx is not None and platform_idx < len(values) else ""
        link = _text(values[link_idx]) if link_idx is not None and link_idx < len(values) else ""
        topic = _text(values[topic_idx]) if topic_idx is not None and topic_idx < len(values) else ""
        text = _text(values[text_idx]) if text_idx is not None and text_idx < len(values) else ""
        photo = _text(values[photo_idx]) if photo_idx is not None and photo_idx < len(values) else ""
        post_type = _text(values[post_type_idx]) if post_type_idx is not None and post_type_idx < len(values) else ""

        section = _section_label(first_cell)
        if section and not any([link, topic, text, photo]):
            current_section = section
            continue

        if not any([platform, link, topic, text, photo]):
            continue
        if text and not any([placement_type, platform, link, topic, photo, post_type]):
            continue
        context = " ".join([placement_type, platform, link, topic, text, photo, post_type, current_section])
        if current_section in {STANDARD_ROW_ORDER[2], STANDARD_ROW_ORDER[3]}:
            label = current_section
        elif current_section and (not placement_type or placement_type.upper() in {"ОС", "ЦС"}):
            label = current_section
        else:
            label = _classify_row(placement_type or current_section, context)
        if label != SERVICE_ROW_NAME:
            counts[label] += 1
    return dict(counts)


def _ordered_fact_rows(counts: dict[str, int], method: str) -> list[dict]:
    rows = []
    seen = set()
    for name in STANDARD_ROW_ORDER:
        if name == SERVICE_ROW_NAME or name not in counts:
            continue
        seen.add(name)
        rows.append({"name": name, "value": int(counts[name]), "method": method})
    for name, value in counts.items():
        if name not in seen and name != SERVICE_ROW_NAME:
            rows.append({"name": name, "value": int(value), "method": method})
    return rows


def extract_project_month_facts(
    path: Path,
    *,
    month_sheet: str = "",
    period: str = "",
    source_system: str = "проектная таблица",
) -> dict:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws, sheet_name = _sheet_with_project_table(wb, month_sheet)
    header_row, headers = _find_header(ws)
    summary_text = _pre_header_summary_text(ws, header_row)
    summary_counts = _parse_summary_counts(summary_text)
    tabular_counts = _extract_tabular_counts(ws, header_row, headers)
    counts = tabular_counts or summary_counts
    method = "tabular_row_count" if tabular_counts else "pre_header_summary"
    fact_rows = _ordered_fact_rows(counts, method)
    total = sum(_clean_int(row.get("value")) for row in fact_rows)
    fact_rows_by_name = {row["name"]: row["value"] for row in fact_rows}
    warnings = []
    if summary_counts and tabular_counts:
        summary_total = sum(summary_counts.values())
        tabular_total = sum(tabular_counts.values())
        if summary_total and tabular_total and summary_total != tabular_total:
            warnings.append(
                f"Summary total ({summary_total}) differs from tabular row count ({tabular_total}); tabular project rows are used as authoritative."
            )
    return {
        "source_file": path.name,
        "source_system": source_system,
        "period_raw": period,
        "sheet": sheet_name,
        "header_row": header_row,
        "fact_rows": fact_rows,
        "fact_rows_by_name": fact_rows_by_name,
        "total_fact": total,
        "campaign_publications_count": total,
        "campaign_publications_by_type": fact_rows_by_name,
        "campaign_fact_method": method,
        "campaign_fact_source_sheet": sheet_name,
        "campaign_vs_organic_status": "verified" if fact_rows else "missing_required",
        "summary_text_present": bool(summary_text.strip()),
        "method": method,
        "warnings": warnings,
        "status": "ready" if fact_rows else "missing_required",
        "methodology": {
            "summary": "Counts filled project ORM table rows for the requested month as the authoritative publication fact.",
            "fallback": "Uses numeric pre-header monthly summary lines only when tabular project rows cannot be extracted.",
            "universality": "No brand or client names are used in classification.",
        },
    }
