from __future__ import annotations

import re
import zipfile
from collections import Counter
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from .slide03_data_extractor import (
    _clean_int,
    _find_header_row_by_columns,
    _idx,
    _period_label_from_content,
    _period_to_caption,
    _rows_from_table,
)


GENERIC_STOP_WORDS = {
    "отзыв", "отзывы", "оценка", "оценки", "день", "раз", "ответ", "ответы",
    "посол", "лета", "цена", "руб", "рублей", "товар", "товары", "заказ",
    "купить", "покупка", "скидка", "скидки", "плюс", "минус", "доставка",
    "капсулы", "форт", "взрослый", "баки", "сет", "шт", "мг", "мл",
    "село", "город", "страна",
}

LOW_SIGNAL_THEMES = {"прочее", "юмор", "знакомство и общение"}
CONTEXT_KEEP_WORDS = {"микрофлора", "кишечник", "пробиотик", "пребиотик", "синбиотик", "врач", "препарат", "капсула", "дети"}
PROMO_FRAGMENTS = {"скид", "бонус", "заказ", "доставка", "polza", "рубл"}
THEMATIC_NOISE_FRAGMENTS = {
    "голосован", "выборы", "опрос", "трамп", "телевизор", "свекров", "книг",
    "путешеств", "мудра", "раковин", "горловк",
}
LOR_CONTEXT_MARKERS = {
    "ангина", "отит", "аденоид", "тонзиллит", "тонзилит", "ларингит",
    "фарингит", "горло", "носоглот", "миндалина", "голос", "осип", "кашель",
}
LOR_RAIL_CONTEXT_RULES = [
    {
        "phrase": "Голос, осиплость и нагрузка на связки",
        "keywords": ["голос", "осип", "охрип", "связк"],
        "exclude_keywords": ["голосован", "трамп", "опрос"],
    },
    {
        "phrase": "Боль в горле, першение и температура",
        "keywords": ["горло", "перш", "температура", "ангина"],
        "exclude_keywords": ["горловк"],
    },
    {
        "phrase": "Кашель и раздражение горла",
        "keywords": ["кашель", "раздраж", "сухость"],
    },
    {
        "phrase": "Ларингит, ангина и тонзиллит",
        "keywords": ["ларингит", "фарингит", "ангина", "тонзиллит", "тонзилит"],
    },
    {
        "phrase": "Носоглотка, аденоиды и ЛОР-осложнения",
        "keywords": ["носоглот", "аденоид", "отит", "ухо", "лора", "лор-", "насморк"],
    },
    {
        "phrase": "Консультация врача и выбор лечения",
        "keywords": ["врач", "лечение", "лора", "назнач"],
    },
    {
        "phrase": "Детские ЛОР-обращения и аденоиды",
        "keywords": ["ребенок", "ребёнок", "дети", "детск", "аденоид"],
    },
]

VASCULAR_CONTEXT_MARKERS = {
    "кровообращ", "кровоток", "микроциркуляц", "сосуд", "капилляр", "кровоснабж",
    "пентоксифиллин", "трентал", "вазонит", "ног", "судорог", "онемен", "диабет",
    "атеросклер", "ишеми", "гематокрит", "эритроцит", "сетчатк", "вен", "тромб",
}

GUT_CONTEXT_MARKERS = {
    "кишечник", "микрофлор", "пробиот", "пребиот", "синбиот", "живот",
    "пищевар", "запор", "стул", "диаре", "антибиотик", "нутриент",
}

VASCULAR_CONTEXT_RULES = [
    {
        "label": "Кровообращение и микроциркуляция",
        "keywords": ["кровообращ", "кровоток", "микроциркуляц", "кровоснабж", "капилляр", "сосуд"],
        "rail": "Кровообращение, микроциркуляция и состояние сосудов",
    },
    {
        "label": "Боли, судороги и онемение ног",
        "keywords": ["ног", "стоп", "икр", "судорог", "онемен", "хромот", "трофическ", "язв"],
        "rail": "Боли, судороги, онемение и дискомфорт в ногах",
    },
    {
        "label": "Диабет и сосудистые осложнения",
        "keywords": ["диабет", "ангиопат", "нейропат", "атеросклер", "ишеми", "инсульт", "сердц"],
        "rail": "Диабет, ишемия и сосудистые осложнения",
    },
    {
        "label": "Разжижение крови и показатели анализа",
        "keywords": ["разжиж", "эритроцит", "гематокрит", "густ", "тромб", "кровь", "коагул"],
        "rail": "Разжижение крови, гематокрит и риск тромбов",
    },
    {
        "label": "Зрение, слух и периферический кровоток",
        "keywords": ["зрен", "сетчат", "глаз", "слух", "уш", "шум в уш", "вестибул"],
        "rail": "Зрение, слух и периферическое кровоснабжение",
    },
    {
        "label": "Беременность, ЭКО и перенос",
        "keywords": ["беремен", "эко", "перенос", "зачат", "плод", "цикл", "овуляц"],
        "rail": "Беременность, ЭКО и поддержка кровотока",
    },
]

GUT_CONTEXT_RULES = [
    {
        "label": "Кишечник и пищеварительный дискомфорт",
        "keywords": ["кишечник", "живот", "пищевар", "запор", "стул", "диаре", "вздут"],
        "rail": "Кишечник, стул и пищеварительный дискомфорт",
    },
    {
        "label": "Микрофлора и пробиотики",
        "keywords": ["микрофлор", "пробиот", "пребиот", "синбиот", "бактери"],
        "rail": "Микрофлора, пробиотики и восстановление баланса",
    },
    {
        "label": "Прием после антибиотиков",
        "keywords": ["антибиотик", "после антибиот", "восстанов"],
        "rail": "Прием после антибиотиков и восстановление",
    },
    {
        "label": "Детское и семейное применение",
        "keywords": ["дети", "ребен", "ребён", "детск", "семейн", "педиатр"],
        "rail": "Детское и семейное применение",
    },
]

GENERIC_MESSAGE_CONTEXT_RULES = [
    {
        "label": "Назначение врача и схема приема",
        "keywords": ["врач", "назнач", "рекоменд", "курс", "дозиров", "принимать", "пить", "таблет"],
        "rail": "Назначение врача, дозировка и схема приема",
    },
    {
        "label": "Выбор препарата и сравнение вариантов",
        "keywords": ["что лучше", "выбор", "сравн", "аналог", "замен", "эффект", "помог"],
        "rail": "Выбор препарата, аналоги и ожидание эффекта",
    },
]

PROBLEM_CONTEXT_RULES = [
    {
        "label": "Осиплость и нагрузка на голос",
        "keywords": ["голос", "осип", "охрип", "связк"],
        "exclude_keywords": ["голосован", "трамп", "опрос"],
        "rail": "Голос, осиплость и нагрузка на связки",
    },
    {
        "label": "Боль и першение в горле",
        "keywords": ["горло", "перш", "глотан", "температура"],
        "exclude_keywords": ["горловк"],
        "rail": "Боль в горле, першение и температура",
    },
    {
        "label": "Кашель и раздражение горла",
        "keywords": ["кашель", "раздраж", "сухость"],
        "rail": "Кашель и раздражение горла",
    },
    {
        "label": "Ларингит и фарингит",
        "keywords": ["ларингит", "фарингит"],
        "rail": "Ларингит, фарингит и воспаление горла",
    },
    {
        "label": "Ангина с температурой",
        "keywords": ["ангина"],
        "rail": "Ангина, боль в горле и температура",
    },
    {
        "label": "Ушная боль и отит",
        "keywords": ["отит", "ухо"],
        "rail": "Ушная боль, отит и ЛОР-осложнения",
    },
    {
        "label": "Аденоиды и носоглотка",
        "keywords": ["носоглот", "аденоид", "насморк"],
        "rail": "Носоглотка, аденоиды и ЛОР-осложнения",
    },
    {
        "label": "Тонзиллит и миндалины",
        "keywords": ["тонзиллит", "тонзилит", "миндалина"],
        "rail": "Хронический тонзиллит и миндалины",
    },
    {
        "label": "Консультация и выбор лечения",
        "keywords": ["врач", "лечение", "назнач", "лора", "лор-", "рекоменд"],
        "rail": "Консультация врача и выбор лечения",
    },
    {
        "label": "Детские ЛОР-обращения",
        "keywords": ["ребенок", "ребёнок", "дети", "детск", "педиатр"],
        "rail": "Детские ЛОР-обращения и аденоиды",
    },
    {
        "label": "Запах изо рта и полость рта",
        "keywords": ["рот", "запах", "галитоз"],
        "rail": "Полость рта и запах изо рта",
    },
    {
        "label": "Поддержка местного иммунитета",
        "keywords": ["иммунитет", "бактерия"],
        "rail": "Поддержка местного иммунитета",
    },
    {
        "label": "Профилактика ЛОР-инфекций",
        "keywords": ["инфекция", "простуда", "вирус"],
        "rail": "Профилактика ЛОР-инфекций",
    },
    {
        "label": "Пробиотики и синбиотики",
        "keywords": ["пробиотик", "пребиотик", "синбиотик", "мульти-пробиотик"],
        "rail": "Пробиотики и синбиотики в пользовательском выборе",
    },
    {
        "label": "Курс и формат приема",
        "keywords": ["капсула", "капсулы"],
        "rail": "Формат капсул и курс приема",
    },
    {
        "label": "Кишечник и пищеварение",
        "keywords": ["кишечник", "жкт", "живот", "пищевар", "диар", "запор"],
        "rail": "Сценарии кишечника и пищеварительного дискомфорта",
    },
    {
        "label": "Детское применение",
        "keywords": ["дети", "ребенок", "ребёнок", "детск", "педиатр"],
        "rail": "Детское применение и семейные сценарии",
    },
    {
        "label": "Микрофлора кишечника",
        "keywords": ["микрофлора", "микробиота"],
        "rail": "Восстановление микрофлоры кишечника",
    },
    {
        "label": "Выбор и ожидание эффекта",
        "keywords": ["вопрос", "выбор", "сравн", "как принимать", "что лучше"],
        "rail": "Вопросы выбора, приема и ожидаемого эффекта",
    },
    {
        "label": "Рекомендации врача",
        "keywords": ["врач", "назнач", "рекоменд", "гастроэнтеролог"],
        "rail": "Рекомендации врача и аптечный выбор",
    },
    {
        "label": "Схема приема препарата",
        "keywords": ["прием", "приём", "препарат", "курс", "дозиров"],
        "rail": "Курс приема и понятные инструкции",
    },
    {
        "label": "Восстановление после лечения",
        "keywords": ["антибиотик", "после лечения", "восстанов"],
        "rail": "Восстановление после лечения и антибиотиков",
    },
]

NS = {
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _norm(value: object) -> str:
    text = str(value or "").replace("\xa0", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def _display_label(value: object) -> str:
    text = _cell_text(value)
    return text[:1].upper() + text[1:] if text else ""


def _label_word_count(value: object) -> int:
    return len([token for token in re.split(r"[\s,;/()«»\"®+._-]+", _norm(value)) if token])


def _is_low_signal_context_name(value: object) -> bool:
    low = _norm(value)
    if not low:
        return True
    if low in GENERIC_STOP_WORDS:
        return True
    return any(fragment in low for fragment in THEMATIC_NOISE_FRAGMENTS)


def _term_matches_rule(term_norm: str, rule: dict) -> bool:
    if any(fragment in term_norm for fragment in rule.get("exclude_keywords", [])):
        return False
    return any(keyword in term_norm for keyword in rule.get("keywords", []))


def _marker_score(corpus_norm: str, markers: set[str]) -> int:
    return sum(corpus_norm.count(marker) for marker in markers)


def _active_context_rules(corpus_text: str) -> tuple[list[dict], str]:
    corpus = _norm(corpus_text)
    scores = {
        "vascular": _marker_score(corpus, VASCULAR_CONTEXT_MARKERS),
        "lor": _marker_score(corpus, LOR_CONTEXT_MARKERS),
        "gut": _marker_score(corpus, GUT_CONTEXT_MARKERS),
    }
    profile = max(scores, key=scores.get)
    if scores[profile] < 4:
        return GENERIC_MESSAGE_CONTEXT_RULES, "generic"
    if profile == "vascular":
        return VASCULAR_CONTEXT_RULES + GENERIC_MESSAGE_CONTEXT_RULES, profile
    if profile == "gut":
        return GUT_CONTEXT_RULES + GENERIC_MESSAGE_CONTEXT_RULES, profile
    lor_rules = [
        rule for rule in PROBLEM_CONTEXT_RULES
        if any(marker in _norm(f"{rule.get('label', '')} {rule.get('rail', '')} {' '.join(rule.get('keywords', []))}") for marker in LOR_CONTEXT_MARKERS)
    ]
    return (lor_rules or PROBLEM_CONTEXT_RULES[:10]) + GENERIC_MESSAGE_CONTEXT_RULES, profile


def _cell_text(value: object) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def _col_to_idx(ref: str | None) -> int:
    match = re.match(r"([A-Z]+)", ref or "")
    out = 0
    for char in match.group(1) if match else "":
        out = out * 26 + ord(char) - 64
    return max(out - 1, 0)


class _ZipSheet:
    def __init__(self, rows: list[list[object]]):
        self._rows = rows
        self.max_row = len(rows)

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        end = min(max_row or self.max_row, self.max_row)
        for idx in range(max(min_row, 1) - 1, end):
            yield tuple(self._rows[idx])


class _ZipWorkbook:
    """Small no-style XLSX reader for exports with invalid style XML."""

    def __init__(self, path: Path):
        self.path = path
        self._sheet_paths = {}
        self._shared = []
        self._cache = {}
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
            if "xl/sharedStrings.xml" in names:
                root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for item in root.findall("m:si", NS):
                    self._shared.append("".join(t.text or "" for t in item.findall(".//m:t", NS)))
            workbook = ET.fromstring(zf.read("xl/workbook.xml"))
            rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rel_map = {rel.get("Id"): rel.get("Target") for rel in rels.findall("pr:Relationship", NS)}
            for sheet in workbook.findall(".//m:sheet", NS):
                name = sheet.get("name") or "Sheet"
                rid = sheet.get(f"{{{NS['r']}}}id")
                target = (rel_map.get(rid) or "").lstrip("/")
                sheet_path = "xl/" + target
                if sheet_path not in names:
                    sheet_path = "xl/worksheets/" + Path(target).name
                self._sheet_paths[name] = sheet_path
        self.sheetnames = list(self._sheet_paths.keys())

    def __getitem__(self, sheet_name: str):
        if sheet_name not in self._sheet_paths:
            raise KeyError(sheet_name)
        if sheet_name not in self._cache:
            self._cache[sheet_name] = _ZipSheet(self._read_rows(self._sheet_paths[sheet_name]))
        return self._cache[sheet_name]

    def _cell_value(self, cell):
        cell_type = cell.get("t")
        value = cell.find("m:v", NS)
        inline = cell.find("m:is", NS)
        if cell_type == "s" and value is not None and value.text not in (None, ""):
            idx = int(float(value.text))
            return self._shared[idx] if 0 <= idx < len(self._shared) else ""
        if cell_type == "inlineStr" and inline is not None:
            return "".join(t.text or "" for t in inline.findall(".//m:t", NS))
        if value is not None and value.text is not None:
            return value.text
        return ""

    def _read_rows(self, sheet_path: str) -> list[list[object]]:
        rows = []
        with zipfile.ZipFile(self.path) as zf:
            root = ET.fromstring(zf.read(sheet_path))
        for row_node in root.findall(".//m:sheetData/m:row", NS):
            values = []
            for cell in row_node.findall("m:c", NS):
                idx = _col_to_idx(cell.get("r"))
                while len(values) <= idx:
                    values.append("")
                values[idx] = self._cell_value(cell)
            while values and values[-1] == "":
                values.pop()
            rows.append(values)
        return rows


def _open_workbook(path: Path | None):
    if not path:
        return None
    try:
        return load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return _ZipWorkbook(path)


def _content_period(wb, fallback: str = "") -> str:
    if not wb:
        return fallback
    content_ws = wb["Содержание"] if "Содержание" in wb.sheetnames else wb[wb.sheetnames[0]]
    return _period_label_from_content(content_ws) or fallback


def _collect_brand_noise(wbs: Iterable, project_brand: str = "") -> set[str]:
    noise = set()
    for value in [project_brand]:
        if value:
            noise.add(_norm(value))
    for wb in wbs:
        if not wb:
            continue
        for sheet_name in ["Показатели брендов", "Показатели"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            header_row, headers = _find_header_row_by_columns(ws, ["Сообщения"], max_rows=20)
            if not header_row:
                continue
            object_idx = _idx(headers, "Объект", default=1)
            for row in ws.iter_rows(min_row=header_row + 1, max_row=min(ws.max_row, header_row + 80), values_only=True):
                values = list(row)
                label = values[object_idx] if object_idx is not None and object_idx < len(values) else None
                label_norm = _norm(label)
                if not label_norm or label_norm == "итого":
                    continue
                noise.add(label_norm)
                for token in re.split(r"[\s,;/()«»\"®+._-]+", label_norm):
                    if len(token) >= 4:
                        noise.add(token)
    return noise


def _filter_rows(rows: list[dict], brand_noise: set[str], max_take: int = 8) -> list[dict]:
    out = []
    seen = set()
    for row in sorted(rows, key=lambda item: item.get("messages", 0), reverse=True):
        name = str(row.get("name") or "").strip()
        low = _norm(name)
        if not low or low in seen:
            continue
        if _is_low_signal_context_name(name):
            continue
        if re.fullmatch(r"\d+([.,]\d+)?", low) or re.fullmatch(r"\d{1,2}[.]\d{1,2}[.]\d{4}", low):
            continue
        if any(fragment in low for fragment in PROMO_FRAGMENTS):
            continue
        if low in GENERIC_STOP_WORDS or (low in brand_noise and low not in CONTEXT_KEEP_WORDS):
            continue
        tokens = [t for t in re.split(r"[\s,;/()«»\"®+._-]+", low) if t]
        if tokens and all(t in GENERIC_STOP_WORDS or (t in brand_noise and t not in CONTEXT_KEEP_WORDS) for t in tokens):
            continue
        messages = _clean_int(row.get("messages"))
        if messages <= 0:
            continue
        out.append({"name": name, "messages": messages, "audience": _clean_int(row.get("audience")), "source": row.get("source", "")})
        seen.add(low)
        if len(out) >= max_take:
            break
    return out


def _sheet_rows(wb, sheet_name: str, required: list[str], source_label: str, *, max_take: int = 20, numeric_col: str = "Сообщения") -> list[dict]:
    if not wb or sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header_row, headers = _find_header_row_by_columns(ws, required, max_rows=120)
    if not header_row:
        return []
    name_candidates = [
        item for item in required
        if item.lower() not in {"сообщения", "аудитория", numeric_col.lower()}
    ]
    name_candidates.extend(["Слово", "Слова", "Тег", "Теги", "Сюжеты", "Сюжет", "Тематика", "Площадка", "Объект"])
    name_idx = _first_idx(headers, *name_candidates, default=1 if len(headers) > 1 else 0)
    msg_idx = _first_idx(headers, numeric_col, "Сообщения", default=None)
    aud_idx = _first_idx(headers, "Аудитория", default=None)
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        values = list(row)
        name = values[name_idx] if name_idx is not None and name_idx < len(values) else None
        if not name or _norm(name) == "итого":
            continue
        messages = _clean_int(values[msg_idx] if msg_idx is not None and msg_idx < len(values) else 0)
        audience = _clean_int(values[aud_idx] if aud_idx is not None and aud_idx < len(values) else 0)
        if messages == 0 and audience == 0:
            continue
        rows.append({"name": str(name).strip(), "messages": messages, "audience": audience, "source": source_label})
    return sorted(rows, key=lambda x: (x["messages"], x["audience"]), reverse=True)[:max_take]


def _raw_popular_words(wb, source_label: str) -> list[dict]:
    return _sheet_rows(wb, "Популярные слова", ["Слово", "Сообщения"], source_label, max_take=30)


def _raw_subjects(wb, source_label: str) -> list[dict]:
    return _sheet_rows(wb, "Сюжеты", ["Сюжеты", "Сообщения"], source_label, max_take=12)


def _thematic_total_from_tags(wb) -> int:
    """Brand Analytics keeps the thematic-query total in the Tags sheet."""
    rows = []
    rows.extend(_sheet_rows(wb, "Теги", ["Тег", "Сообщения"], "tags:Теги", max_take=60))
    rows.extend(_sheet_rows(wb, "Теги", ["Теги", "Сообщения"], "tags:Теги", max_take=60))
    for row in rows:
        name = _norm(row.get("name"))
        if "тематические запрос" in name:
            return _clean_int(row.get("messages"))
    return 0


def _is_lor_problem_field(items: list[dict]) -> bool:
    chunks = []
    for item in items:
        chunks.append(str(item.get("name") or ""))
        chunks.extend(str(term or "") for term in item.get("supporting_terms") or [])
    text = _norm(" ".join(chunks))
    return any(marker in text for marker in LOR_CONTEXT_MARKERS)


def _make_lor_context_rail(source_terms: list[dict]) -> list[str]:
    scored = []
    for rule in LOR_RAIL_CONTEXT_RULES:
        total = 0
        matched = set()
        for row in source_terms:
            name = _norm(row.get("name"))
            if not name:
                continue
            if _is_low_signal_context_name(name):
                continue
            if _term_matches_rule(name, rule):
                total += _clean_int(row.get("messages"))
                matched.add(name)
        if total > 0 and matched:
            scored.append((total, rule["phrase"]))
    scored.sort(key=lambda item: item[0], reverse=True)

    out = []
    seen = set()
    for _, phrase in scored:
        key = _norm(phrase)
        if key not in seen:
            out.append(phrase)
            seen.add(key)
        if len(out) >= 5:
            break
    return out


def _first_idx(headers, *names, default=None):
    wanted = [str(name or "").strip().lower() for name in names if str(name or "").strip()]
    for name in wanted:
        for idx, header in enumerate(headers):
            header_low = str(header or "").strip().lower()
            if header_low == name or name in header_low:
                return idx
    return default


def _total_from_indicators(wb) -> int:
    if not wb:
        return 0
    for sheet_name in ["Показатели", "Сводные данные"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True):
            vals = list(row)
            labels = [_norm(v) for v in vals]
            if "всего" in labels and "сообщения" in labels:
                msg_idx = labels.index("сообщения")
                total_idx = labels.index("всего")
                if msg_idx + 1 < len(vals):
                    return _clean_int(vals[msg_idx + 1])
                if total_idx + 1 < len(vals):
                    return _clean_int(vals[total_idx + 1])
            if labels and labels[0] == "сообщения" and len(vals) > 1:
                return _clean_int(vals[1])
            if "сообщения" in labels:
                idx = labels.index("сообщения")
                if idx + 1 < len(vals):
                    total = _clean_int(vals[idx + 1])
                    if total:
                        return total
        header_row, headers = _find_header_row_by_columns(ws, ["Сообщения"], max_rows=20)
        msg_idx = _idx(headers, "Сообщения", default=None)
        if header_row and msg_idx is not None:
            for row in ws.iter_rows(min_row=header_row + 1, max_row=min(ws.max_row, header_row + 10), values_only=True):
                values = list(row)
                first = next((_norm(v) for v in values if _norm(v)), "")
                if first == "всего" and msg_idx < len(values):
                    return _clean_int(values[msg_idx])
    return 0


def _source_label(source_system: str) -> str:
    low = _norm(source_system)
    if "brand analytics" in low or low == "ba":
        return "сервис мониторинга Brand Analytics"
    return "сервис мониторинга Медиалогия SM"


def _brand_matrix_values(wb, sheet_name: str, project_brand: str, *, label_col: int = 1, stop_at_total: bool = True, max_data_rows: int = 24) -> list[dict]:
    if not wb or sheet_name not in wb.sheetnames or not project_brand:
        return []
    ws = wb[sheet_name]
    project_col = None
    header_row = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True), start=1):
        values = list(row)
        for col_idx, value in enumerate(values):
            if _norm(value) == _norm(project_brand):
                project_col = col_idx
                header_row = idx
                break
        if project_col is not None:
            break
    if project_col is None or header_row is None:
        return []
    out = []
    scanned = 0
    started = False
    section_stops = {"аудитория", "просмотры", "вовлечённость", "вовлеченность", "см индекс"}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        scanned += 1
        if scanned > max_data_rows:
            break
        values = list(row)
        label = _cell_text(values[label_col] if len(values) > label_col else "")
        if not label:
            if started:
                break
            continue
        if _norm(label) in section_stops:
            break
        value = _clean_int(values[project_col] if len(values) > project_col else 0)
        if value <= 0:
            continue
        started = True
        raw_value = values[project_col] if len(values) > project_col else 0
        out.append({"name": label, "messages": float(raw_value) if _norm(label) == "средняя оценка" else value})
        if stop_at_total and _norm(label) in {"всего", "итого"}:
            break
    return out


def _brand_mentions_context(wb, project_brand: str) -> dict:
    if not wb or not project_brand:
        return {}
    rows = []
    for sheet_name in ["Показатели брендов", "Показатели"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row, headers = _find_header_row_by_columns(ws, ["Объект", "Сообщения"], max_rows=40)
        if not header_row:
            continue
        object_idx = _idx(headers, "Объект", default=1)
        msg_idx = _idx(headers, "Сообщения", default=None)
        aud_idx = _idx(headers, "Аудитория", default=None)
        if msg_idx is None:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            values = list(row)
            brand = str(values[object_idx] if object_idx is not None and len(values) > object_idx else "" or "").strip()
            if not brand:
                continue
            if _norm(brand) == "итого":
                break
            messages = _clean_int(values[msg_idx] if len(values) > msg_idx else 0)
            if messages <= 0:
                continue
            rows.append({
                "brand": brand,
                "messages": messages,
                "audience": _clean_int(values[aud_idx] if aud_idx is not None and len(values) > aud_idx else 0),
            })
        if rows:
            break
    total = sum(item["messages"] for item in rows)
    sorted_rows = sorted(rows, key=lambda item: item["messages"], reverse=True)
    project = next((item for item in sorted_rows if _norm(item["brand"]) == _norm(project_brand)), {})
    rank = next((idx + 1 for idx, item in enumerate(sorted_rows) if _norm(item["brand"]) == _norm(project_brand)), None)
    if not project:
        return {}
    return {
        "brand_mentions": project.get("messages", 0),
        "brand_audience": project.get("audience", 0),
        "competitive_total_mentions": total,
        "brand_sov": project.get("messages", 0) / total if total else 0,
        "brand_rank": rank,
        "leader": sorted_rows[0] if sorted_rows else {},
    }


def _simple_table_after_header(ws, header_name: str) -> dict:
    header_row = None
    headers = []
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=True), start=1):
        values = list(row)
        if any(_norm(value) == _norm(header_name) for value in values):
            header_row = idx
            headers = values
            break
    if not header_row:
        return {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=min(ws.max_row, header_row + 8), values_only=True):
        values = list(row)
        if _norm(values[1] if len(values) > 1 else "") not in {"всего", "итого"}:
            continue
        out = {}
        for idx, header in enumerate(headers):
            key = str(header or "").strip()
            if key:
                out[key] = values[idx] if idx < len(values) else None
        return out
    return {}


def _own_brand_indicators(wb) -> dict:
    if not wb or "Показатели" not in wb.sheetnames:
        return {}
    ws = wb["Показатели"]
    sentiment = _simple_table_after_header(ws, "Позитивные")
    msg_types = _simple_table_after_header(ws, "Посты")
    out = {}
    if sentiment:
        out["positive_messages"] = _clean_int(sentiment.get("Позитивные"))
        out["neutral_messages"] = _clean_int(sentiment.get("Нейтральные"))
        out["negative_messages"] = _clean_int(sentiment.get("Негативные"))
        out["positive_share"] = float(sentiment.get("% позитивных") or 0)
        out["negative_share"] = float(sentiment.get("% негативных") or 0)
    if msg_types:
        out["reviews_messages"] = _clean_int(msg_types.get("Отзывы"))
        out["comments_messages"] = _clean_int(msg_types.get("Комментарии"))
        out["posts_messages"] = _clean_int(msg_types.get("Посты"))
    return out


def _message_sheet_summary(wb, sheet_name: str) -> dict:
    if not wb or sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    header_row, headers = _find_header_row_by_columns(ws, ["Тональность"], max_rows=30)
    if not header_row:
        return {}
    text_idx = _idx(headers, "Текст", default=None)
    tone_idx = _idx(headers, "Тональность", default=None)
    type_idx = _idx(headers, "Тип источника", default=None)
    source_idx = _idx(headers, "Источник", default=None)
    total = 0
    tones = Counter()
    source_types = Counter()
    sources = Counter()
    keywords = Counter()
    keyword_rules = {
        "витамин D / усвоение": ["витамин", "усво"],
        "антибиотики / восстановление": ["антибиот", "восстанов"],
        "запор / стул": ["запор", "стул"],
        "нет эффекта": ["нет результата", "не помог", "не повлиял", "не почувств"],
        "упаковка / доставка": ["упаков", "вскрыт", "просроч", "транспортиров"],
        "аллергия / переносимость": ["аллерг", "дискомфорт", "отрава"],
    }
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        values = list(row)
        text = _cell_text(values[text_idx] if text_idx is not None and len(values) > text_idx else "")
        tone = _cell_text(values[tone_idx] if tone_idx is not None and len(values) > tone_idx else "").lower()
        source_type = _cell_text(values[type_idx] if type_idx is not None and len(values) > type_idx else "")
        source = _cell_text(values[source_idx] if source_idx is not None and len(values) > source_idx else "")
        if not text and not tone and not source:
            continue
        total += 1
        if tone:
            tones[tone] += 1
        if source_type:
            source_types[source_type] += 1
        if source:
            sources[source] += 1
        low_text = text.lower()
        for label, fragments in keyword_rules.items():
            if any(fragment in low_text for fragment in fragments):
                keywords[label] += 1
    return {
        "sheet": sheet_name,
        "total_rows": total,
        "tones": dict(tones.most_common()),
        "source_types": dict(source_types.most_common(5)),
        "sources": dict(sources.most_common(5)),
        "keyword_contexts": dict(keywords.most_common(6)),
    }


def _initiated_summary(wb) -> dict:
    if not wb or "Инициированные сообщения" not in wb.sheetnames:
        return {}
    ws = wb["Инициированные сообщения"]
    header_row, headers = _find_header_row_by_columns(ws, ["Тип размещения", "Ссылка на сообщение"], max_rows=30)
    if not header_row:
        return {}
    type_idx = _idx(headers, "Тип размещения", default=None)
    product_idx = _idx(headers, "Продукт", default=None)
    total = 0
    by_type = Counter()
    products = Counter()
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        values = list(row)
        placement_type = _cell_text(values[type_idx] if type_idx is not None and len(values) > type_idx else "")
        product = _cell_text(values[product_idx] if product_idx is not None and len(values) > product_idx else "")
        if not placement_type:
            continue
        total += 1
        by_type[placement_type] += 1
        if product:
            products[product] += 1
    return {"total": total, "by_type": dict(by_type.most_common()), "products": dict(products.most_common(5))}


def _rating_summary(wb, project_brand: str) -> dict:
    rows = _brand_matrix_values(wb, "Оценки", project_brand, stop_at_total=False, max_data_rows=12)
    if not rows:
        return {}
    values = {item["name"]: item["messages"] for item in rows}
    total = values.get("Итого", 0)
    return {
        "total_reviews_with_rating": total,
        "five_star": values.get("Оценка 5", 0),
        "one_two_star": values.get("Оценка 1", 0) + values.get("Оценка 2", 0),
        "average_rating": next((float(item["messages"]) for item in rows if item["name"] == "Средняя оценка"), 0.0),
        "five_star_share": values.get("Оценка 5", 0) / total if total else 0,
    }


def _insight_signals(wb, project_brand: str) -> dict:
    if not wb:
        return {}
    platform_types = [item for item in _brand_matrix_values(wb, "Типы площадок ", project_brand) if _norm(item["name"]) not in {"всего", "итого"}]
    platforms = [item for item in _brand_matrix_values(wb, "Площадки ", project_brand) if _norm(item["name"]) not in {"всего", "итого"}]
    positive = _message_sheet_summary(wb, "Сообщения позитив")
    neutral_negative = _message_sheet_summary(wb, "Сообщения нейтрал+негатив")
    vitamin_d = _message_sheet_summary(wb, "Упоминания с витамином Д")
    initiated = _initiated_summary(wb)
    brand_mentions = _brand_mentions_context(wb, project_brand)
    if initiated and brand_mentions.get("brand_mentions"):
        initiated["organic_mentions"] = max(brand_mentions["brand_mentions"] - initiated.get("total", 0), 0)
    return {
        "brand_mentions": brand_mentions,
        "own_brand": _own_brand_indicators(wb),
        "platform_types": sorted(platform_types, key=lambda item: item["messages"], reverse=True)[:5],
        "platforms": sorted(platforms, key=lambda item: item["messages"], reverse=True)[:5],
        "ratings": _rating_summary(wb, project_brand),
        "positive_messages": positive,
        "neutral_negative_messages": neutral_negative,
        "vitamin_d_messages": vitamin_d,
        "initiated": initiated,
    }


def _make_context_phrases(chart_terms: list[dict], themes: list[dict], tags: list[dict], subjects: list[dict]) -> list[str]:
    phrases = []
    if themes:
        for item in themes:
            if _norm(item.get("name")) not in LOW_SIGNAL_THEMES:
                phrases.append(str(item["name"]))
        for item in subjects:
            phrases.append(str(item["name"]))
        for item in chart_terms:
            phrases.append(str(item["name"]))
    else:
        for item in chart_terms:
            phrases.append(str(item["name"]))
        for item in subjects:
            phrases.append(str(item["name"]))
    for item in tags:
        if _norm(item.get("name")) not in LOW_SIGNAL_THEMES:
            phrases.append(str(item["name"]))

    out = []
    seen = set()
    for phrase in phrases:
        key = _norm(phrase)
        if _is_low_signal_context_name(phrase):
            continue
        if re.fullmatch(r"\d+([.,]\d+)?", key) or re.fullmatch(r"\d{1,2}[.]\d{1,2}[.]\d{4}", key):
            continue
        if any(fragment in key for fragment in PROMO_FRAGMENTS):
            continue
        if key and key not in seen:
            out.append(phrase)
            seen.add(key)
        if len(out) >= 5:
            break
    return out


def _make_problem_context_model(source_terms: list[dict]) -> tuple[list[dict], list[str]]:
    """Convert raw frequent words into reference-style problem contexts.

    The source rows are keyword mentions, not mutually exclusive clusters. The
    grouped value therefore remains a context strength indicator and is not used
    as an additive share of the whole infopole.
    """
    grouped = []
    used_labels = set()
    assigned_terms = set()
    active_rules, _profile = _active_context_rules(" ".join(str(row.get("name") or "") for row in source_terms))
    for rule in active_rules:
        matched_terms = []
        seen_terms = set()
        for row in source_terms:
            name = str(row.get("name") or "").strip()
            low = _norm(name)
            if not low:
                continue
            if _is_low_signal_context_name(name) or low in assigned_terms:
                continue
            if _term_matches_rule(low, rule):
                key = (low, row.get("source", ""))
                if key not in seen_terms:
                    matched_terms.append(row)
                    seen_terms.add(key)
        value = sum(_clean_int(item.get("messages")) for item in matched_terms)
        if value <= 0:
            continue
        grouped.append({
            "name": rule["label"],
            "messages": value,
            "audience": sum(_clean_int(item.get("audience")) for item in matched_terms),
            "source": "semantic_group:" + ",".join(str(item.get("source", "")) for item in matched_terms if item.get("source")),
            "supporting_terms": [str(item.get("name") or "").strip() for item in matched_terms if item.get("name")],
            "context_note": "Контекст сгруппирован из частотных слов; значения показывают силу темы и не суммируются между собой.",
            "rail_phrase": rule["rail"],
        })
        used_labels.add(_norm(rule["label"]))
        assigned_terms.update(_norm(item.get("name")) for item in matched_terms if _norm(item.get("name")))

    grouped = sorted(grouped, key=lambda item: item.get("messages", 0), reverse=True)
    if len(grouped) < 6:
        for row in sorted(source_terms, key=lambda item: item.get("messages", 0), reverse=True):
            label = str(row.get("name") or "").strip()
            low = _norm(label)
            if not label or low in used_labels or low in assigned_terms:
                continue
            if _is_low_signal_context_name(label):
                continue
            if _label_word_count(label) < 2:
                continue
            label = _display_label(label)
            grouped.append({
                "name": label,
                "messages": _clean_int(row.get("messages")),
                "audience": _clean_int(row.get("audience")),
                "source": row.get("source", ""),
                "supporting_terms": [label],
                "context_note": "Сохранен как прямой контекст из исходной выгрузки.",
            })
            used_labels.add(_norm(label))
            assigned_terms.add(low)
            if len(grouped) >= 6:
                break

    rail = []
    seen = set()
    for item in grouped:
        phrase = item.get("rail_phrase") or item.get("name")
        key = _norm(phrase)
        if key and key not in seen:
            rail.append(phrase)
            seen.add(key)
        if len(rail) >= 5:
            break
    return grouped[:8], rail


def _message_key(values: list[object], indexes: Iterable[int | None], fallback: int) -> str:
    for idx in indexes:
        if idx is None or idx >= len(values):
            continue
        value = _cell_text(values[idx])
        if value:
            return value
    return f"row:{fallback}"


def _nonempty_message_text(values: list[object], indexes: Iterable[int | None]) -> str:
    chunks = []
    for idx in indexes:
        if idx is None or idx >= len(values):
            continue
        text = _cell_text(values[idx])
        if text:
            chunks.append(text)
    return " ".join(chunks).strip()


def _make_exclusive_message_context_model(raw_wb, brand_noise: set[str]) -> tuple[list[dict], list[str], dict]:
    """Build slide 04 contexts from raw BA messages without double-counting.

    Each source message is assigned to the first matching analytical rule. This
    keeps the chart additive: one row of the raw "Messages" sheet can contribute
    to one context only, even if the text contains several keywords.
    """
    diagnostics = {
        "status": "missing_raw_messages",
        "source_sheet": "",
        "unique_message_count": 0,
        "assigned_message_count": 0,
        "unmatched_message_count": 0,
        "category_count": 0,
    }
    if not raw_wb:
        return [], [], diagnostics
    sheet_name = next((name for name in raw_wb.sheetnames if _norm(name) == "сообщения"), "")
    if not sheet_name:
        return [], [], diagnostics
    ws = raw_wb[sheet_name]
    header_row, headers = _find_header_row_by_columns(ws, ["Текст"], max_rows=40)
    if not header_row:
        header_row, headers = _find_header_row_by_columns(ws, ["Текст сообщения"], max_rows=40)
    if not header_row:
        diagnostics["status"] = "missing_raw_message_header"
        diagnostics["source_sheet"] = sheet_name
        return [], [], diagnostics

    title_idx = _first_idx(headers, "Заголовок", default=None)
    text_idx = _first_idx(headers, "Текст", "Текст сообщения", default=None)
    subject_idx = _first_idx(headers, "Сюжет", default=None)
    thematic_idx = _first_idx(headers, "Тематические запросы", default=None)
    hash_idx = _first_idx(headers, "Hash сообщения", "Хэш сообщения", "Hash", default=None)
    id_idx = _first_idx(headers, "ID сообщения", "Id сообщения", "ID", default=None)
    url_idx = _first_idx(headers, "Url", "URL", "Ссылка", "Ссылка на сообщение", default=None)
    audience_idx = _first_idx(headers, "Аудитория", default=None)
    search_indexes = [title_idx, text_idx, subject_idx, thematic_idx]
    key_indexes = [hash_idx, id_idx, url_idx]

    message_records = []
    seen_messages = set()
    unmatched_messages = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True), start=header_row + 1):
        values = list(row)
        message_text = _nonempty_message_text(values, search_indexes)
        if len(message_text) < 8:
            continue
        key = _message_key(values, key_indexes, row_num)
        if key in seen_messages:
            continue
        seen_messages.add(key)
        low = _norm(message_text)
        if low in brand_noise and low not in CONTEXT_KEEP_WORDS:
            unmatched_messages += 1
            continue
        message_records.append({
            "values": values,
            "low": low,
            "thematic": _cell_text(values[thematic_idx] if thematic_idx is not None and thematic_idx < len(values) else ""),
            "subject": _cell_text(values[subject_idx] if subject_idx is not None and subject_idx < len(values) else ""),
        })

    active_rules, context_profile = _active_context_rules(" ".join(record["low"] for record in message_records))
    buckets = {
        _norm(rule["label"]): {
            "rule": rule,
            "messages": 0,
            "audience": 0,
            "examples": Counter(),
        }
        for rule in active_rules
    }
    assigned_messages = 0

    for record in message_records:
        values = record["values"]
        low = record["low"]
        matched_rule = None
        for rule in active_rules:
            if _term_matches_rule(low, rule):
                matched_rule = rule
                break
        if not matched_rule:
            unmatched_messages += 1
            continue
        bucket = buckets[_norm(matched_rule["label"])]
        bucket["messages"] += 1
        bucket["audience"] += _clean_int(values[audience_idx] if audience_idx is not None and audience_idx < len(values) else 0)
        example = record["thematic"] or record["subject"]
        if example and not _is_low_signal_context_name(example):
            bucket["examples"][example] += 1
        assigned_messages += 1

    grouped = []
    for bucket in buckets.values():
        rule = bucket["rule"]
        messages = int(bucket["messages"])
        if messages <= 0:
            continue
        supporting_terms = [term for term, _ in bucket["examples"].most_common(5)] or [rule["label"]]
        grouped.append({
            "name": rule["label"],
            "messages": messages,
            "audience": int(bucket["audience"]),
            "source": f"message_rows:{sheet_name}",
            "supporting_terms": supporting_terms,
            "context_note": "Контекст построен по строкам сообщений: каждое сообщение отнесено только к одному ведущему блоку.",
            "rail_phrase": rule["rail"],
        })

    grouped = sorted(grouped, key=lambda item: item.get("messages", 0), reverse=True)[:8]
    rail = []
    seen_rail = set()
    for item in grouped:
        phrase = item.get("rail_phrase") or item.get("name")
        key = _norm(phrase)
        if key and key not in seen_rail:
            rail.append(phrase)
            seen_rail.add(key)
        if len(rail) >= 5:
            break

    diagnostics.update({
        "status": "ready" if len(grouped) >= 6 else "insufficient_categories",
        "source_sheet": sheet_name,
        "unique_message_count": len(seen_messages),
        "assigned_message_count": assigned_messages,
        "unmatched_message_count": unmatched_messages,
        "category_count": len(grouped),
        "context_profile": context_profile,
    })
    return grouped, rail, diagnostics


def _enrich_context_rail(base_contexts: list[str], signals: dict) -> list[str]:
    contexts = []
    if signals.get("ratings", {}).get("total_reviews_with_rating"):
        contexts.append("Эффект и переносимость по отзывам")
    contexts.extend(base_contexts)
    out = []
    seen = set()
    for item in contexts:
        key = _norm(item)
        if key and key not in seen:
            out.append(item)
            seen.add(key)
        if len(out) >= 5:
            break
    return out


def extract_problem_field(
    *,
    analytical_path: Path | None = None,
    raw_path: Path | None = None,
    report_path: Path | None = None,
    project_brand: str = "",
    source_system: str = "",
    project_period: str = "",
) -> dict:
    """Build the slide 04 data model.

    Source precedence follows the plan: Medialogia analytical report first, raw
    export as an April-compatible supplement, and report workbook only for brand
    noise / fallback metadata.
    """
    analytical_wb = _open_workbook(analytical_path)
    raw_wb = _open_workbook(raw_path)
    report_wb = _open_workbook(report_path)

    period_raw = _content_period(analytical_wb, "") or _content_period(raw_wb, "") or _content_period(report_wb, "") or project_period
    month_prep, month_caption, month_title = _period_to_caption(period_raw)
    brand_noise = _collect_brand_noise([analytical_wb, raw_wb, report_wb], project_brand)

    words = _sheet_rows(analytical_wb, "Слова", ["Слова", "Сообщения"], "analytical:Слова", max_take=30)
    themes = _sheet_rows(analytical_wb, "Тематики", ["Тематика", "Сообщения"], "analytical:Тематики", max_take=12)
    tags = _sheet_rows(analytical_wb, "Теги", ["Теги", "Сообщения"], "analytical:Теги", max_take=12)
    raw_words = _raw_popular_words(raw_wb, "raw:Популярные слова")
    raw_tags = _sheet_rows(raw_wb, "Теги", ["Тег", "Сообщения"], "raw:Теги", max_take=12)
    raw_subjects = _raw_subjects(raw_wb, "raw:Сюжеты")

    chart_terms = _filter_rows(words, brand_noise, max_take=8)
    if len(chart_terms) < 6:
        chart_terms = _filter_rows(chart_terms + raw_words + raw_subjects, brand_noise, max_take=8)
    if len(chart_terms) < 6:
        chart_terms = _filter_rows(chart_terms + themes + tags + raw_tags, brand_noise, max_take=8)

    semantic_source_terms = _filter_rows(
        words + raw_words + themes + raw_subjects + tags + raw_tags,
        brand_noise,
        max_take=80,
    ) or chart_terms
    exclusive_chart_terms, exclusive_contexts, exclusive_diagnostics = _make_exclusive_message_context_model(raw_wb, brand_noise)
    chart_method = "semantic_keyword_groups"
    if len(exclusive_chart_terms) >= 4:
        display_chart_terms = exclusive_chart_terms
        right_contexts = exclusive_contexts
        chart_method = "mutually_exclusive_message_rows"
    else:
        semantic_chart_terms, semantic_contexts = _make_problem_context_model(semantic_source_terms)
        if len(semantic_chart_terms) >= 4:
            display_chart_terms = semantic_chart_terms
            right_contexts = semantic_contexts
        else:
            display_chart_terms = chart_terms
            right_contexts = _make_context_phrases(chart_terms, _filter_rows(themes, brand_noise, 8), _filter_rows(tags + raw_tags, brand_noise, 8), raw_subjects)
    if _is_lor_problem_field(display_chart_terms):
        right_contexts = _make_lor_context_rail(semantic_source_terms) or right_contexts
    analytical_total = _total_from_indicators(analytical_wb)
    raw_total = _total_from_indicators(raw_wb)
    thematic_total = _thematic_total_from_tags(raw_wb) or _thematic_total_from_tags(analytical_wb)
    report_total = _total_from_indicators(report_wb)
    total_messages = thematic_total or max([value for value in [analytical_total, raw_total] if value] or [report_total])
    signals = _insight_signals(report_wb, project_brand) or _insight_signals(analytical_wb, project_brand)
    if len(display_chart_terms) < 8 and signals:
        vitamin_messages = _clean_int((signals.get("vitamin_d_messages") or {}).get("keyword_contexts", {}).get("витамин D / усвоение"))
        if vitamin_messages > 0 and all(_norm(item.get("name")) != "витамин d и усвоение" for item in display_chart_terms):
            display_chart_terms = display_chart_terms + [{
                "name": "Витамин D и усвоение",
                "messages": vitamin_messages,
                "audience": 0,
                "source": "message_texts:Упоминания с витамином Д",
                "supporting_terms": ["витамин", "усвоение"],
                "context_note": "Контекст выделен по текстам сообщений; категории могут пересекаться.",
            }]
    right_contexts = _enrich_context_rail(right_contexts, signals) if signals else right_contexts
    if _is_lor_problem_field(display_chart_terms):
        right_contexts = _make_lor_context_rail(semantic_source_terms) or right_contexts

    caveats = []
    if chart_method == "mutually_exclusive_message_rows":
        caveats.append("Контексты графика взаимоисключающие: одно сообщение относится только к одному ведущему блоку.")
    else:
        caveats.append("Контекстные группы построены по частотным словам и текстам сообщений; группы могут пересекаться.")
    if len(display_chart_terms) < 6:
        caveats.append("Найдено меньше 6 устойчивых контекстов; слайд требует ручной проверки.")

    return {
        "source_files": {
            "analytical": str(analytical_path) if analytical_path else "",
            "raw": str(raw_path) if raw_path else "",
            "report": str(report_path) if report_path else "",
        },
        "source_system": source_system or "Медиалогия",
        "source_label": _source_label(source_system),
        "project_brand": project_brand,
        "period_raw": period_raw,
        "month_prepositional": month_prep,
        "month_caption": month_caption,
        "month_title": month_title,
        "total_topic_messages": total_messages,
        "chart_terms": display_chart_terms,
        "raw_chart_terms": chart_terms,
        "chart_method": chart_method,
        "chart_unique_source_sheet": exclusive_diagnostics.get("source_sheet", ""),
        "chart_unique_message_count": exclusive_diagnostics.get("unique_message_count", 0),
        "chart_assigned_message_count": exclusive_diagnostics.get("assigned_message_count", 0),
        "chart_unmatched_message_count": exclusive_diagnostics.get("unmatched_message_count", 0),
        "chart_exclusive_diagnostics": exclusive_diagnostics,
        "right_contexts": right_contexts,
        "themes": _filter_rows(themes, brand_noise, 8),
        "tags": _filter_rows(tags + raw_tags, brand_noise, 8),
        "raw_subjects": raw_subjects[:5],
        "insight_signals": signals,
        "caveats": caveats,
        "methodology_status": "ready_with_caveat" if caveats else "ready",
    }
