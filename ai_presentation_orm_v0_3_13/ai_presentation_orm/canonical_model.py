from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .project_month_fact_extractor import extract_project_month_facts
from .slide03_data_extractor import (
    _aggregate_brand_rows,
    _brand_from_product_name,
    _brand_name_matches,
    _canonical_brand_name,
    _clean_int,
    _clean_float,
    _filter_competitive_rows,
    _find_header as _find_metrics_header,
    _find_metrics_sheet,
    _get_index,
    _is_thematic_metric_label,
    _norm_brand_name,
    _period_label_from_content,
    _period_to_caption,
)
from .slide06_tonality_extractor import (
    _extract_project_negative_count as _extract_orm_negative,
    _apply_project_negative_override,
)
from .slide07_sources_extractor import (
    _source_display,
    _map_source_type,
    _merge_source_rows_by_label,
)
from .slide08_seeding_metrics_extractor import _month_number
from .xlsx_safe_reader import open_workbook_safe

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Legacy entities used by readiness_checker
CANONICAL_ENTITIES = [
    "project", "media_plan", "plan_fact", "brand_mentions",
    "sov", "themes", "sentiment", "sources", "seeding",
    "examples", "ratings", "conclusions", "methodology",
]

SLIDE_REQUIRED_ENTITIES = {
    "cover": ["project"],
    "current_report_plan_fact": ["project", "media_plan", "plan_fact"],
    "brand_mentions_sov": ["brand_mentions", "sov"],
    "problem_field": ["themes"],
    "tonality": ["sentiment"],
    "sources": ["sources"],
    "seeding_results": ["seeding"],
    "message_examples": ["examples"],
    "ratings_reviews_cards": ["ratings"],
    "conclusions": ["conclusions"],
    "unknown": [],
}

SOURCE_TYPE_ORDER = ["Соцсети", "Мессенджеры", "Отзывы", "Видео", "Другое", "Блоги", "Форумы"]
TONE_KEYS = ("positive", "neutral", "negative")

# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------


def _lower(value) -> str:
    return str(value or "").strip().lower().replace("ё", "е")


def _safe_pct(part: int, total: int) -> float:
    return part / total if total else 0.0


def _fmt_int(value: int) -> str:
    return f"{int(value):,}".replace(",", " ")


def _date_key(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    text = str(value or "").replace("\xa0", " ").strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except Exception:
            pass
    return text


def _normalize_url(url: str) -> str:
    text = str(url or "").strip().lower()
    text = re.sub(r"https?://(?:www\.)?", "", text)
    text = text.split("?")[0].split("#")[0].rstrip("/")
    return text


def _platform_from_url(url: str, fallback: str = "") -> str:
    text = str(url or "").strip()
    low = text.lower()
    mapping = [
        ("youtube.com", "YouTube"),
        ("youtu.be", "YouTube"),
        ("rutube.ru", "Rutube"),
        ("vk.com", "ВКонтакте"),
        ("vk.ru", "ВКонтакте"),
        ("t.me", "Telegram"),
        ("telegram.org", "Telegram"),
        ("dzen.ru", "Дзен"),
        ("ozon.ru", "Ozon"),
        ("wildberries.ru", "Wildberries"),
        ("market.yandex", "Яндекс.Маркет"),
        ("facebook.com", "Facebook"),
        ("instagram.com", "Instagram"),
        ("ok.ru", "Одноклассники"),
        ("tiktok.com", "TikTok"),
        ("mom.life", "Mom.Life"),
        ("babyblog.ru", "Babyblog.ru"),
        ("sprosivracha.com", "Sprosivracha.com"),
        ("woman.ru", "Woman.ru"),
        ("otzovik.com", "Отзовик"),
        ("irecommend.ru", "IRecommend"),
        ("apteka.ru", "Apteka.ru"),
        ("megapteka.ru", "Мегаптека"),
        ("uteka.ru", "Ютека"),
        ("prodoctorov.ru", "ПроДокторов"),
        ("rigla.ru", "Ригла"),
    ]
    for needle, label in mapping:
        if needle in low:
            return label
    match = re.search(r"https?://(?:www\.)?([^/]+)", low)
    if match:
        return match.group(1)
    return text[:40] or fallback or "Не указано"


def _normalize_source(source: str) -> str:
    return _source_display(source)


def _normalize_platform_type(raw_type: str) -> str:
    low = _lower(raw_type)
    if "видео" in low or "youtube" in low:
        return "Соцсети"
    if "чат" in low:
        return "Мессенджеры"
    if "блог" in low:
        return "Блоги"
    if "форум" in low:
        return "Форумы"
    if "сми" in low or "медиа" in low or "новост" in low:
        return "СМИ"
    return _map_source_type(raw_type)


def _normalize_message_type(value: str) -> str:
    low = _lower(value)
    if "отзыв" in low:
        return "review"
    if "коммент" in low or "обсужд" in low or "поддерж" in low:
        return "comment"
    if "форум" in low:
        return "forum"
    if "соцсет" in low or "микроблог" in low:
        return "social"
    if "блог" in low:
        return "blog"
    if "мессенджер" in low:
        return "messenger"
    if "медиа" in low or "сми" in low or "новост" in low:
        return "media"
    if "маркетплейс" in low or "отзов" in low:
        return "marketplace"
    return "other"


# ---------------------------------------------------------------------------
# Source exclusion
# ---------------------------------------------------------------------------


def _build_exclusion_note(source_exclusions: list[str]) -> dict:
    if not source_exclusions:
        return {
            "sources": [],
            "removed_messages_count": 0,
            "affected_scopes": [],
        }
    return {
        "sources": list(source_exclusions),
        "removed_messages_count": 0,
        "affected_scopes": ["competitive_sov", "topics", "tonality", "platforms"],
    }


# ---------------------------------------------------------------------------
# ORM campaign scope
# ---------------------------------------------------------------------------


def _extract_orm_campaign(project_orm_path: Path, month_sheet: str, period: str) -> dict:
    try:
        facts = extract_project_month_facts(
            project_orm_path,
            month_sheet=month_sheet,
            period=period,
            source_system="проектная таблица ORM",
        )
        total = int(facts.get("total_fact", 0) or 0)
        by_type = facts.get("campaign_publications_by_type") or facts.get("fact_rows_by_name") or {}
        reviews = 0
        comments = 0
        for name, value in by_type.items():
            low = _lower(name)
            cnt = _clean_int(value)
            if "отзыв" in low:
                reviews += cnt
            elif cnt > 0:
                comments += cnt
        if not reviews:
            for row in (facts.get("fact_rows") or []):
                if "отзыв" in _lower(row.get("name", "")):
                    reviews += _clean_int(row.get("value"))
        if not comments:
            comments = max(total - reviews, 0)

        views_total = 0
        try:
            wb = load_workbook(project_orm_path, data_only=True, read_only=False)
            if month_sheet and month_sheet in wb.sheetnames:
                ws = wb[month_sheet]
            else:
                for name in wb.sheetnames:
                    if month_sheet and month_sheet.lower() in name.lower():
                        ws = wb[name]
                        break
                else:
                    ws = wb[wb.sheetnames[0]]
            from .slide08_seeding_metrics_extractor import _extract_rows, _choose_views_metric
            rows, _meta = _extract_rows(ws)
            views_info = _choose_views_metric(rows)
            views_total = int(views_info.get("total", 0) or 0)
        except Exception:
            pass

        return {
            "total": total,
            "by_type": {
                "comments": comments,
                "reviews": reviews,
            },
            "views_total": views_total,
            "source_sheet": facts.get("sheet", month_sheet),
            "fact_method": facts.get("campaign_fact_method") or facts.get("method") or "project_orm_rows_with_content",
            "fact_rows_by_name": by_type,
            "status": "ready" if total > 0 else "missing_required",
        }
    except Exception as exc:
        return {
            "total": 0,
            "by_type": {"comments": 0, "reviews": 0},
            "views_total": 0,
            "source_sheet": month_sheet,
            "fact_method": "error",
            "fact_rows_by_name": {},
            "status": "error",
            "error": str(exc),
        }


# ---------------------------------------------------------------------------
# Comparative report extraction (the source of truth for brand data)
# ---------------------------------------------------------------------------


def _extract_comparative_brand_data(comp_path: Path, project_brand: str, competitor_brands: list[str]) -> dict:
    """Extract per-brand metrics from the comparative report's Показатели sheet.

    The comparative report has rows: Объект | Сообщения | СМ Индекс | Аудитория | ...
    and a Тональность section later in the same sheet.
    """
    result = {
        "by_brand": {},
        "by_brand_audience": {},
        "by_brand_views": {},
        "by_brand_engagement": {},
        "by_brand_sov": {},
        "by_brand_tonality": {},
        "total_mentions": 0,
        "has_data": False,
    }

    if not comp_path or not comp_path.exists():
        return result

    try:
        wb = load_workbook(comp_path, data_only=True, read_only=False)
    except Exception:
        return result

    if "Показатели" not in wb.sheetnames:
        return result

    ws = wb["Показатели"]
    header_found = False
    object_idx = mentions_idx = audience_idx = views_idx = engagement_idx = sov_idx = None
    header_row = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), 1):
        vals = [str(v or "").strip().lower() for v in row]
        if "объект" in vals and "сообщения" in vals:
            header_found = True
            header_row = r_idx
            for i, v in enumerate([str(x or "").strip() for x in row]):
                vl = v.lower()
                if vl == "объект":
                    object_idx = i
                elif vl == "сообщения":
                    mentions_idx = i
                elif vl in ("аудитория", "аудитория блога"):
                    audience_idx = i
                elif vl in ("просмотры",):
                    views_idx = i
                elif vl in ("вовлечённость", "вовлеченность"):
                    engagement_idx = i
                elif vl in ("share of voice", "sov", "доля голоса"):
                    sov_idx = i
            break

    if not header_found or object_idx is None or mentions_idx is None:
        return result

    brand_mentions = {}
    brand_audience = {}
    brand_views = {}
    brand_engagement = {}
    brand_sov = {}

    tonality_data = {}  # brand -> {positive, neutral, negative}
    main_section = True  # Are we still in the main Метрики section?

    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True), header_row + 1):
        vals = list(row)
        obj_raw = str(vals[object_idx] or "").strip() if object_idx < len(vals) else ""
        if not obj_raw:
            continue
        obj_low = obj_raw.lower()

        # Section header detection - end of main Метрики section
        if obj_low in ("итого", "всего"):
            if main_section:
                # End of main metrics
                main_section = False
            continue

        if obj_low in ("тональность",):
            # Тональность section: columns are Объект, Позитивные, Нейтральные, Негативные...
            for tr_idx, trow in enumerate(ws.iter_rows(min_row=r_idx + 2, max_row=ws.max_row, values_only=True), r_idx + 2):
                tv = list(trow)
                tobj = str(tv[object_idx] or "").strip() if object_idx < len(tv) else ""
                if not tobj or tobj.lower() in ("итого", "всего", "типы сообщений", "источники сообщений", ""):
                    if tobj.lower() in ("типы сообщений", "источники сообщений"):
                        break
                    continue
                # In tonality section: mentions_idx = Позитивные, +1 = Нейтральные, +2 = Негативные
                tpos = _clean_int(tv[mentions_idx] if mentions_idx is not None and mentions_idx < len(tv) else 0)
                tneu = _clean_int(tv[mentions_idx + 1] if mentions_idx is not None and mentions_idx + 1 < len(tv) else 0)
                tneg = _clean_int(tv[mentions_idx + 2] if mentions_idx is not None and mentions_idx + 2 < len(tv) else 0)
                if tpos or tneu or tneg:
                    tonality_data[tobj] = {"positive": tpos, "neutral": tneu, "negative": tneg}
            continue

        # Skip non-metric sections
        if not main_section:
            continue

        mentions = _clean_int(vals[mentions_idx] if mentions_idx < len(vals) else 0)
        if mentions <= 0:
            continue

        brand_mentions[obj_raw] = mentions
        if audience_idx is not None and audience_idx < len(vals):
            brand_audience[obj_raw] = _clean_int(vals[audience_idx])
        if views_idx is not None and views_idx < len(vals):
            brand_views[obj_raw] = _clean_int(vals[views_idx])
        if engagement_idx is not None and engagement_idx < len(vals):
            brand_engagement[obj_raw] = _clean_int(vals[engagement_idx])
        if sov_idx is not None and sov_idx < len(vals):
            raw_sov = str(vals[sov_idx] or "").strip().replace(",", ".")
            try:
                brand_sov[obj_raw] = float(raw_sov) / 100 if raw_sov else 0.0
            except Exception:
                brand_sov[obj_raw] = 0.0

    # Normalize brand names
    norm_mentions = {}
    norm_audience = {}
    norm_views = {}
    norm_engagement = {}
    norm_sov = {}
    norm_tonality = {}

    for raw_name, count in brand_mentions.items():
        norm = _canonical_brand_name(raw_name, project_brand, competitor_brands)
        norm_mentions[norm] = norm_mentions.get(norm, 0) + count
        if raw_name in brand_audience:
            norm_audience[norm] = norm_audience.get(norm, 0) + brand_audience[raw_name]
        if raw_name in brand_views:
            norm_views[norm] = norm_views.get(norm, 0) + brand_views[raw_name]
        if raw_name in brand_engagement:
            norm_engagement[norm] = norm_engagement.get(norm, 0) + brand_engagement[raw_name]
        if raw_name in brand_sov:
            norm_sov[norm] = brand_sov[raw_name]

    for raw_name, tone in tonality_data.items():
        norm = _canonical_brand_name(raw_name, project_brand, competitor_brands)
        if norm not in norm_tonality:
            norm_tonality[norm] = {"positive": 0, "neutral": 0, "negative": 0}
        norm_tonality[norm]["positive"] += tone["positive"]
        norm_tonality[norm]["neutral"] += tone["neutral"]
        norm_tonality[norm]["negative"] += tone["negative"]

    total = sum(norm_mentions.values())
    result.update({
        "by_brand": norm_mentions,
        "by_brand_audience": norm_audience,
        "by_brand_views": norm_views,
        "by_brand_engagement": norm_engagement,
        "by_brand_sov": norm_sov,
        "by_brand_tonality": norm_tonality,
        "total_mentions": total,
        "has_data": total > 0,
    })
    return result


# ---------------------------------------------------------------------------
# BA (Brand Analytics) export fallback — Теги sheet
# ---------------------------------------------------------------------------


def _extract_ba_brand_data(ba_path: Path, project_brand: str, competitor_brands: list[str]) -> dict:
    """Extract per-brand metrics from Brand Analytics export Теги sheet.

    BA exports have a Теги (Tags) sheet instead of a Показатели sheet.
    Header columns: Теги/Показатель | Сообщения | Аудитория | Вовлеченность | Лояльность | Позитив | Нейтрально | Негатив
    Data rows contain brand names with per-brand mention/audience/tonality counts.
    """
    result = {
        "by_brand": {},
        "by_brand_audience": {},
        "by_brand_views": {},
        "by_brand_engagement": {},
        "by_brand_sov": {},
        "by_brand_tonality": {},
        "total_mentions": 0,
        "has_data": False,
    }

    if not ba_path or not ba_path.exists():
        return result

    wb = open_workbook_safe(ba_path)
    if "Теги" not in wb.sheetnames:
        return result

    try:
        ws = wb["Теги"]
        header_row = None
        tag_col = mentions_col = audience_col = engagement_col = pos_col = neu_col = neg_col = None

        for r_idx in range(1, min(ws.max_row + 1, 20)):
            row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
            vals = [str(v or "").strip().lower() for v in row]
            if "тег" in vals or "показатель" in vals:
                if "сообщения" in vals:
                    header_row = r_idx
                    raw_vals = [str(v or "").strip() for v in row]
                    for i, v in enumerate(raw_vals):
                        vl = v.lower()
                        if vl in ("тег", "показатель", "теги / показатель", "теги/показатель"):
                            tag_col = i
                        elif vl == "сообщения":
                            mentions_col = i
                        elif vl in ("аудитория", "аудитория блога"):
                            audience_col = i
                        elif vl in ("вовлеченность", "вовлеченность"):
                            engagement_col = i
                        elif vl in ("позитив", "позитивные", "положительные"):
                            pos_col = i
                        elif vl in ("нейтрально", "нейтральные"):
                            neu_col = i
                        elif vl in ("негатив", "негативные"):
                            neg_col = i
                    break

        if header_row is None or tag_col is None or mentions_col is None:
            return result

        brand_mentions = {}
        brand_audience = {}
        brand_engagement = {}
        brand_tonality = {}

        for r_idx in range(header_row + 1, min(ws.max_row + 1, header_row + 50)):
            row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
            vals = [str(v or "").strip() for v in row]
            if tag_col >= len(vals):
                continue
            tag = vals[tag_col]
            if not tag:
                continue
            tag_lower = tag.lower()

            # Stop at daily breakdown (date column) or итого/всего summary
            if tag_lower in ("дата", "итого", "всего", ""):
                break

            # Skip non-brand rows (thematic queries, generic labels)
            skip_keywords = ["тематические запросы", "не считаем в sov", "всего", "итого"]
            if any(kw in tag_lower for kw in skip_keywords):
                continue

            mentions = _clean_int(vals[mentions_col] if mentions_col is not None and mentions_col < len(vals) else 0)
            if mentions <= 0:
                continue

            brand_mentions[tag] = mentions
            if audience_col is not None and audience_col < len(vals):
                brand_audience[tag] = _clean_int(vals[audience_col])
            if engagement_col is not None and engagement_col < len(vals):
                brand_engagement[tag] = _clean_int(vals[engagement_col])
            if pos_col is not None and neu_col is not None and neg_col is not None:
                pos = _clean_int(vals[pos_col] if pos_col < len(vals) else 0)
                neu = _clean_int(vals[neu_col] if neu_col < len(vals) else 0)
                neg = _clean_int(vals[neg_col] if neg_col < len(vals) else 0)
                if pos or neu or neg:
                    brand_tonality[tag] = {"positive": pos, "neutral": neu, "negative": neg}

        # Normalize brand names
        norm_mentions = {}
        norm_audience = {}
        norm_engagement = {}
        norm_tonality = {}

        for raw_name, count in brand_mentions.items():
            norm = _canonical_brand_name(raw_name, project_brand, competitor_brands)
            norm_mentions[norm] = norm_mentions.get(norm, 0) + count
            if raw_name in brand_audience:
                norm_audience[norm] = norm_audience.get(norm, 0) + brand_audience[raw_name]
            if raw_name in brand_engagement:
                norm_engagement[norm] = norm_engagement.get(norm, 0) + brand_engagement[raw_name]

        for raw_name, tone in brand_tonality.items():
            norm = _canonical_brand_name(raw_name, project_brand, competitor_brands)
            if norm not in norm_tonality:
                norm_tonality[norm] = {"positive": 0, "neutral": 0, "negative": 0}
            norm_tonality[norm]["positive"] += tone["positive"]
            norm_tonality[norm]["neutral"] += tone["neutral"]
            norm_tonality[norm]["negative"] += tone["negative"]

        total = sum(norm_mentions.values())
        result.update({
            "by_brand": norm_mentions,
            "by_brand_audience": norm_audience,
            "by_brand_views": {},
            "by_brand_engagement": norm_engagement,
            "by_brand_sov": {},
            "by_brand_tonality": norm_tonality,
            "total_mentions": total,
            "has_data": total > 0,
        })
    except Exception:
        pass

    return result


# ---------------------------------------------------------------------------
# Analytics message scope (from main analytics Сообщения sheet)
# ---------------------------------------------------------------------------


def _extract_analytics_messages(analytics_path: Path) -> dict:
    """Extract message-level data from the main analytics Сообщения sheet."""
    wb = open_workbook_safe(analytics_path)
    period_raw = ""
    if "Содержание" in wb.sheetnames:
        period_raw = _period_label_from_content(wb["Содержание"])

    result = {
        "source_file": analytics_path.name,
        "period_raw": period_raw,
        "total_messages": 0,
        "by_source": {},
        "by_platform_type": {},
        "by_topic": {},
        "message_cards": [],
    }

    if "Сообщения" not in wb.sheetnames:
        return result

    try:
        ws = wb["Сообщения"]
        # Find header row
        header_row = None
        headers = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            vals = [str(v or "").strip() for v in row]
            if "Ссылка на сообщение" in vals and "Текст сообщения" in vals:
                header_row = r_idx
                headers = vals
                break
            # Also handle BA Medialogia export format (Url + Текст)
            if "Url" in vals and "Текст" in vals and "Тональность" in vals:
                header_row = r_idx
                headers = vals
                break

        if header_row is None:
            return result

        # Column indices
        def ci(*names):
            for i, h in enumerate(headers):
                hl = h.lower().strip()
                for n in names:
                    if hl == n.lower().strip():
                        return i
            return None

        url_idx = ci("Ссылка на сообщение", "Url", "URL", "Ссылка")
        text_idx = ci("Текст сообщения", "Текст")
        site_idx = ci("Площадка", "Источник")
        type_idx = ci("Тип площадки", "Тип источника")
        tone_idx = ci("Тональность")
        date_idx = ci("Время публикации", "Дата")
        obj_idx = ci("Объекты", "Роль объекта")
        tag_idx = ci("Теги")
        community_idx = ci("Где пишет", "Место публикации")
        audience_idx = ci("Аудитория блога", "Аудитория")

        site_counts = Counter()
        type_counts = Counter()
        topic_counts = Counter()
        message_cards = []

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            vals = list(row)

            # Get objects
            objects_raw = str(vals[obj_idx] or "").strip() if obj_idx is not None and obj_idx < len(vals) else ""

            site_raw = str(vals[site_idx] if site_idx is not None and site_idx < len(vals) else "").strip()
            type_raw = str(vals[type_idx] if type_idx is not None and type_idx < len(vals) else "").strip()
            tone_raw = str(vals[tone_idx] if tone_idx is not None and tone_idx < len(vals) else "").strip().lower()
            url_raw = str(vals[url_idx] if url_idx is not None and url_idx < len(vals) else "").strip()
            text_raw = str(vals[text_idx] if text_idx is not None and text_idx < len(vals) else "").strip()
            date_raw = str(vals[date_idx] if date_idx is not None and date_idx < len(vals) else "").strip()

            tone = ""
            if "позит" in tone_raw or "полож" in tone_raw:
                tone = "positive"
            elif "негат" in tone_raw or "отриц" in tone_raw:
                tone = "negative"
            elif "нейтр" in tone_raw:
                tone = "neutral"

            if not tone:
                continue

            source_norm = _normalize_source(site_raw)
            type_norm = _normalize_platform_type(type_raw)

            site_counts[source_norm] += 1
            type_counts[type_norm] += 1

            card = {
                "date": _date_key(date_raw),
                "source": source_norm,
                "platform_type": type_norm,
                "url": _normalize_url(url_raw),
                "domain": _normalize_url(url_raw).split("/")[0] if url_raw else "",
                "author": "",
                "text_norm": text_raw[:200],
                "brand_norm": "",
                "message_type": _normalize_message_type(type_raw),
                "views": 0,
                "tone": tone,
                "objects_raw": objects_raw,
            }
            message_cards.append(card)

        result.update({
            "total_messages": len(message_cards),
            "by_source": dict(site_counts.most_common(30)),
            "by_platform_type": dict(type_counts),
            "message_cards": message_cards,
        })
    except Exception as exc:
        result["error"] = str(exc)

    return result


# ---------------------------------------------------------------------------
# Comparative platform scope (from Типы площадок sheet)
# ---------------------------------------------------------------------------


def _extract_comparative_platform_scope(comp_path: Path, project_brand: str) -> dict:
    """Extract platform type breakdown from comparative report."""
    result = {
        "total": 0,
        "by_platform_type": {},
        "platform_type_rows": [],
        "source_rows": [],
        "other_count": 0,
    }

    if not comp_path or not comp_path.exists():
        return result

    try:
        wb = load_workbook(comp_path, data_only=True, read_only=False)
    except Exception:
        return result

    if "Типы площадок" not in wb.sheetnames:
        return result

    ws = wb["Типы площадок"]
    # Find header: Объект | type1 | type2 | ...
    header_row = None
    headers = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        vals = [str(v or "").strip() for v in row]
        if "объект" in [v.lower() for v in vals]:
            header_row = r_idx
            headers = vals
            break

    if header_row is None:
        return result

    # Find target brand column
    brand_col = None
    brand_low = _lower(project_brand)
    for i, h in enumerate(headers):
        if brand_low == _lower(h) or brand_low in _lower(h) or _lower(h) in brand_low:
            brand_col = i
            break

    if brand_col is None:
        return result

    # Also get total (Итого) column
    total_col = None
    for i, h in enumerate(headers):
        if _lower(h) == "итого":
            total_col = i
            break

    type_counts = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = [str(v or "").strip() for v in row]
        type_name = vals[0] if vals else ""
        if not type_name or _lower(type_name) in ("всего", "итого", ""):
            continue
        count = _clean_int(vals[brand_col] if brand_col < len(vals) else 0)
        if count > 0:
            type_counts[type_name] = count

    total = sum(type_counts.values())
    platform_type_rows = []
    for tname in SOURCE_TYPE_ORDER:
        cnt = type_counts.pop(tname, 0)
        if cnt > 0:
            platform_type_rows.append({"type": tname, "messages": cnt, "share": cnt / total if total else 0})
    for tname, cnt in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
        platform_type_rows.append({"type": tname, "messages": cnt, "share": cnt / total if total else 0})

    # Also try to read Площадки sheet for source-level data
    if "Площадки" in wb.sheetnames:
        try:
            ws2 = wb["Площадки"]
            hrow2 = None
            for r_idx, row in enumerate(ws2.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                vals = [str(v or "").strip() for v in row]
                if "источник" in [v.lower() for v in vals]:
                    hrow2 = r_idx
                    headers2 = vals
                    break
            if hrow2 is not None:
                src_col = None
                for i, h in enumerate(headers2):
                    if _lower(h) == "источник":
                        src_col = i
                        break
                if src_col is not None:
                    source_rows = []
                    for row in ws2.iter_rows(min_row=hrow2 + 1, max_row=ws2.max_row, values_only=True):
                        vals = [str(v or "").strip() for v in row]
                        src = vals[src_col] if src_col < len(vals) else ""
                        if not src or _lower(src) in ("всего", "итого", ""):
                            continue
                        count = _clean_int(vals[brand_col] if brand_col < len(vals) else 0)
                        if count > 0:
                            source_rows.append({"source": src, "label": _source_display(src), "messages": count})
                    source_rows = sorted(source_rows, key=lambda x: x["messages"], reverse=True)
                    top10 = source_rows[:10]
                    top_sum = sum(r["messages"] for r in top10)
                    other = max(total - top_sum, 0)
                    for r in top10:
                        r["share"] = r["messages"] / total if total else 0
                    if other > 0:
                        top10.append({"source": "other", "label": "Другие", "messages": other, "share": other / total if total else 0})
                    result["source_rows"] = top10
                    result["other_count"] = other
        except Exception:
            pass

    result.update({
        "total": total,
        "by_platform_type": {r["type"]: r["messages"] for r in platform_type_rows},
        "platform_type_rows": platform_type_rows,
    })
    return result


# ---------------------------------------------------------------------------
# ORM-to-analytics reconciliation
# ---------------------------------------------------------------------------


def _build_match_card(orm_row: dict, index: int) -> dict:
    return {
        "index": index,
        "date": _date_key(orm_row.get("date", "")),
        "source": _normalize_source(orm_row.get("platform_raw", "")),
        "platform_type": _normalize_platform_type(orm_row.get("placement_type", "")),
        "url": _normalize_url(orm_row.get("message_link", "")),
        "domain": _normalize_url(orm_row.get("message_link", "")).split("/")[0] if orm_row.get("message_link") else "",
        "platform": orm_row.get("platform", ""),
        "author": orm_row.get("author", ""),
        "text_norm": (orm_row.get("text") or "")[:200],
        "brand_norm": "",
        "message_type": _normalize_message_type(orm_row.get("placement_type", "")),
        "views": _clean_int(orm_row.get("received_views", 0)),
    }


def _resolve_analytics_platform(card: dict) -> str:
    """Resolve a comparable platform name from an analytics card.

    Uses the URL's domain (via _platform_from_url) as the primary source,
    falling back to the source field normalized via _normalize_source.
    """
    url = card.get("url", "")
    source = card.get("source", "")
    from_url = _platform_from_url(url) if url else ""
    if from_url and from_url != "Не указано":
        return from_url
    from_source = _platform_from_url(source) if source else ""
    if from_source and from_source != "Не указано":
        return from_source
    return _normalize_source(source) if source else ""


def _date_close(d1: str, d2: str, max_days: int = 3) -> bool:
    if not d1 or not d2:
        return True
    try:
        return abs((datetime.strptime(d1, "%d.%m.%Y") - datetime.strptime(d2, "%d.%m.%Y")).days) <= max_days
    except Exception:
        return True


def _text_overlap(t1: str, t2: str) -> float:
    if not t1 or not t2:
        return 0.0
    w1 = {w for w in t1.lower().split() if len(w) > 3}
    w2 = {w for w in t2.lower().split() if len(w) > 3}
    if not w1 or not w2:
        return 0.0
    return len(w1 & w2) / max(len(w1 | w2), 1)


def _reconcile_orm_with_analytics(
    orm_rows: list[dict],
    analytics_cards: list[dict],
    brand_object_filters: list[str] | None = None,
) -> dict:
    """Match ORM placements against analytics messages.

    Multi-strategy matching in priority order (first match wins per ORM row):
      1. exact normalized URL
      2. same domain + close date (3d) + text overlap > 25%
      3. same platform + close date (3d) + text overlap > 20%
      4. same domain + text overlap > 15% (no date)
      5. same platform + text overlap > 10% (no date)
      6. same domain + close date (3d) + card objects_raw matches brand filters
      7. same platform + close date (3d) + card objects_raw matches brand filters
      8. same domain + close date (3d)
      9. same platform + close date (3d)

    Args:
        brand_object_filters: optional list of substrings; if provided, an
            analytics card is eligible for loose matching (steps 6-7) only
            when its ``objects_raw`` contains at least one filter (case-insensitive).

    Returns matched, unmatched ORM, and organic analytics messages.
    """
    orm_matched_indices: set[int] = set()
    analytics_matched_indices: set[int] = set()
    match_quality: dict[str, int] = {
        "exact_url": 0, "domain_date_text": 0, "platform_date_text": 0,
        "domain_text_weak": 0, "platform_text_weak": 0,
        "domain_date_filtered": 0, "platform_date_filtered": 0,
        "domain_date": 0, "platform_date": 0,
    }

    # Build ORM match cards once
    orm_cards = [_build_match_card(r, i) for i, r in enumerate(orm_rows)]

    # Build analytics lookup structures
    analytics_by_url: dict[str, list[int]] = defaultdict(list)
    analytics_by_domain: dict[str, list[int]] = defaultdict(list)
    analytics_by_platform: dict[str, list[int]] = defaultdict(list)
    analytics_is_filtered: list[bool] = [False] * len(analytics_cards)

    brand_filters_lower = [f.lower() for f in (brand_object_filters or [])]

    for idx, card in enumerate(analytics_cards):
        url = card.get("url", "")
        if url:
            analytics_by_url[url].append(idx)
        domain = card.get("domain", "")
        if domain:
            analytics_by_domain[domain].append(idx)
        platform = _resolve_analytics_platform(card)
        if platform:
            analytics_by_platform[platform].append(idx)
        # Check brand object filters
        if brand_filters_lower:
            obj_raw = (card.get("objects_raw") or "").lower()
            if any(f in obj_raw for f in brand_filters_lower):
                analytics_is_filtered[idx] = True

    # -------------------------------------------------------------------
    # Helper: try to match a single ORM card against a set of candidate
    # analytics indices with a given predicate.  Returns (matched_ai, None)
    # or (None, None).
    # -------------------------------------------------------------------
    def _try_match(oi, candidates, predicate):
        for ai in candidates:
            if ai in analytics_matched_indices:
                continue
            if predicate(oi, ai):
                orm_matched_indices.add(oi)
                analytics_matched_indices.add(ai)
                return ai
        return None

    # -----------------------------------------------------------------------
    # Priority 1: exact normalized URL match
    # -----------------------------------------------------------------------
    for oi, oc in enumerate(orm_cards):
        if oi in orm_matched_indices:
            continue
        url = oc["url"]
        if url and url in analytics_by_url:
            _try_match(oi, analytics_by_url[url],
                       lambda o, a: True)  # any match is good
            if oi in orm_matched_indices:
                match_quality["exact_url"] += 1

    # -----------------------------------------------------------------------
    # Priority 2: same domain + close date + text overlap > 25%
    # -----------------------------------------------------------------------
    for oi, oc in enumerate(orm_cards):
        if oi in orm_matched_indices:
            continue
        domain = oc["domain"]
        if not domain:
            continue
        _try_match(oi, analytics_by_domain.get(domain, []),
                   lambda o, a: (_date_close(orm_cards[o]["date"], analytics_cards[a].get("date", ""), 3)
                                 and _text_overlap(orm_cards[o]["text_norm"], analytics_cards[a].get("text_norm", "")) > 0.25))
        if oi in orm_matched_indices:
            match_quality["domain_date_text"] += 1

    # -----------------------------------------------------------------------
    # Priority 3: same platform + close date + text overlap > 20%
    # -----------------------------------------------------------------------
    for oi, oc in enumerate(orm_cards):
        if oi in orm_matched_indices:
            continue
        platform = oc.get("platform", "")
        if not platform:
            continue
        _try_match(oi, analytics_by_platform.get(platform, []),
                   lambda o, a: (_date_close(orm_cards[o]["date"], analytics_cards[a].get("date", ""), 3)
                                 and _text_overlap(orm_cards[o]["text_norm"], analytics_cards[a].get("text_norm", "")) > 0.20))
        if oi in orm_matched_indices:
            match_quality["platform_date_text"] += 1

    # -----------------------------------------------------------------------
    # Priority 4: same domain + text overlap > 15% (no date)
    # -----------------------------------------------------------------------
    for oi, oc in enumerate(orm_cards):
        if oi in orm_matched_indices:
            continue
        domain = oc["domain"]
        if not domain:
            continue
        _try_match(oi, analytics_by_domain.get(domain, []),
                   lambda o, a: _text_overlap(orm_cards[o]["text_norm"], analytics_cards[a].get("text_norm", "")) > 0.15)
        if oi in orm_matched_indices:
            match_quality["domain_text_weak"] += 1

    # -----------------------------------------------------------------------
    # Priority 5: same platform + text overlap > 10% (no date)
    # -----------------------------------------------------------------------
    for oi, oc in enumerate(orm_cards):
        if oi in orm_matched_indices:
            continue
        platform = oc.get("platform", "")
        if not platform:
            continue
        _try_match(oi, analytics_by_platform.get(platform, []),
                   lambda o, a: _text_overlap(orm_cards[o]["text_norm"], analytics_cards[a].get("text_norm", "")) > 0.10)
        if oi in orm_matched_indices:
            match_quality["platform_text_weak"] += 1

    # -----------------------------------------------------------------------
    # Priority 6: same domain + close date + objects_raw matches brand filter
    # -----------------------------------------------------------------------
    if brand_filters_lower:
        for oi, oc in enumerate(orm_cards):
            if oi in orm_matched_indices:
                continue
            domain = oc["domain"]
            if not domain:
                continue
            _try_match(oi,
                       [ai for ai in analytics_by_domain.get(domain, []) if analytics_is_filtered[ai]],
                       lambda o, a: _date_close(orm_cards[o]["date"], analytics_cards[a].get("date", ""), 3))
            if oi in orm_matched_indices:
                match_quality["domain_date_filtered"] += 1

    # -----------------------------------------------------------------------
    # Priority 7: same platform + close date + objects_raw matches brand filter
    # -----------------------------------------------------------------------
    if brand_filters_lower:
        for oi, oc in enumerate(orm_cards):
            if oi in orm_matched_indices:
                continue
            platform = oc.get("platform", "")
            if not platform:
                continue
            _try_match(oi,
                       [ai for ai in analytics_by_platform.get(platform, []) if analytics_is_filtered[ai]],
                       lambda o, a: _date_close(orm_cards[o]["date"], analytics_cards[a].get("date", ""), 3))
            if oi in orm_matched_indices:
                match_quality["platform_date_filtered"] += 1

    # -----------------------------------------------------------------------
    # Priority 8: same domain + close date (no text, no filter)
    # -----------------------------------------------------------------------
    for oi, oc in enumerate(orm_cards):
        if oi in orm_matched_indices:
            continue
        domain = oc["domain"]
        if not domain:
            continue
        _try_match(oi, analytics_by_domain.get(domain, []),
                   lambda o, a: _date_close(orm_cards[o]["date"], analytics_cards[a].get("date", ""), 3))
        if oi in orm_matched_indices:
            match_quality["domain_date"] += 1

    # -----------------------------------------------------------------------
    # Priority 9: same platform + close date (no text, no filter)
    # -----------------------------------------------------------------------
    for oi, oc in enumerate(orm_cards):
        if oi in orm_matched_indices:
            continue
        platform = oc.get("platform", "")
        if not platform:
            continue
        _try_match(oi, analytics_by_platform.get(platform, []),
                   lambda o, a: _date_close(orm_cards[o]["date"], analytics_cards[a].get("date", ""), 3))
        if oi in orm_matched_indices:
            match_quality["platform_date"] += 1

    unmatched_orm_count = len(orm_rows) - len(orm_matched_indices)
    organic_analytics_indices = [i for i in range(len(analytics_cards)) if i not in analytics_matched_indices]

    return {
        "orm_total": len(orm_rows),
        "matched_in_analytics": len(orm_matched_indices),
        "unmatched_added": unmatched_orm_count,
        "organic_analytics_messages": len(organic_analytics_indices),
        "match_quality": match_quality,
        "orm_matched_indices": sorted(orm_matched_indices),
        "analytics_matched_indices": sorted(analytics_matched_indices),
        "organic_analytics_indices": organic_analytics_indices,
    }


# ---------------------------------------------------------------------------
# Competitive SOV
# ---------------------------------------------------------------------------


def _compute_competitive_sov(
    comp_brand_data: dict,
    project_brand: str,
    competitor_brands: list[str],
    orm_campaign_total: int,
    target_organic_reconciled: int = 0,
) -> dict:
    """Compute SOV using comparative report brand data + ORM campaign.

    Args:
        target_organic_reconciled: organic count from ORM-analytics reconciliation.
            If 0, falls back to raw analytics count (no reconciliation).
    """
    by_brand = comp_brand_data.get("by_brand", {})
    by_brand_sov = comp_brand_data.get("by_brand_sov", {})

    # Target brand from analytics
    target_analytics = 0
    for name, count in by_brand.items():
        if _brand_name_matches(name, project_brand):
            target_analytics = count
            break

    # Use reconciled organic if available, otherwise raw analytics
    target_organic = target_organic_reconciled if target_organic_reconciled > 0 else target_analytics
    target_campaign = orm_campaign_total
    target_with_campaign = target_organic + target_campaign
    target_without_campaign = target_organic

    # Competitors: use ONLY the brands explicitly listed in competitor_brands
    # Do NOT add every non-target brand from the comparative report
    competitors_list = []
    competitor_total = 0
    competitor_norm = {_norm_brand_name(cb) for cb in (competitor_brands or [])}

    for name, count in by_brand.items():
        if _brand_name_matches(name, project_brand):
            continue
        if competitor_norm and _norm_brand_name(name) not in competitor_norm:
            continue  # skip brands not in the competitor list
        competitors_list.append({
            "brand": name,
            "mentions": count,
        })
        competitor_total += count

    # Add zero-mention entries for competitors not found in data
    found_competitor_norms = {_norm_brand_name(c["brand"]) for c in competitors_list}
    for cb in (competitor_brands or []):
        if _norm_brand_name(cb) not in found_competitor_norms and not _brand_name_matches(cb, project_brand):
            competitors_list.append({"brand": cb, "mentions": 0})

    total_with_campaign = target_with_campaign + competitor_total
    total_without_campaign = target_without_campaign + competitor_total

    sov_with_campaign = target_with_campaign / total_with_campaign if total_with_campaign else 0
    sov_without_campaign = target_without_campaign / total_without_campaign if total_without_campaign else 0
    sov_lift_pp = (sov_with_campaign - sov_without_campaign) * 100

    return {
        "target": {
            "brand": project_brand,
            "organic": target_organic,
            "campaign": target_campaign,
            "with_campaign": target_with_campaign,
            "without_campaign": target_without_campaign,
            "sov_with_campaign": round(sov_with_campaign, 4),
            "sov_without_campaign": round(sov_without_campaign, 4),
            "sov_lift_pp": round(sov_lift_pp, 2),
        },
        "competitors": competitors_list,
        "total_with_campaign": total_with_campaign,
        "total_without_campaign": total_without_campaign,
    }


# ---------------------------------------------------------------------------
# Tonality
# ---------------------------------------------------------------------------


def _compute_tonality(
    comp_brand_data: dict,
    project_brand: str,
    orm_campaign_total: int,
    target_with_campaign: int,
    orm_negative_meta: dict,
) -> dict:
    """Compute target-brand tonality reconciled with ORM campaign data.

    Canonical formula:
      positive = ORM campaign total (all campaign placements are positive)
      negative = analytics negative for the target brand (after exclusions)
      neutral = target_with_campaign - positive - negative
    """
    negative_method = "empty_orm_negative_section"
    by_tone = comp_brand_data.get("by_brand_tonality", {})

    # Analytics tonality for target
    analytics_positive = 0
    analytics_neutral = 0
    analytics_negative = 0
    analytics_total = 0
    for name, tone in by_tone.items():
        if _brand_name_matches(name, project_brand):
            analytics_positive = tone.get("positive", 0)
            analytics_neutral = tone.get("neutral", 0)
            analytics_negative = tone.get("negative", 0)
            analytics_total = analytics_positive + analytics_neutral + analytics_negative
            break

    orm_negative_count = 0
    orm_negative_status = orm_negative_meta.get("status", "section_not_found")
    if orm_negative_status == "ready":
        orm_negative_count = max(int(orm_negative_meta.get("negative_count", 0) or 0), 0)
        negative_method = "project_orm_negative_section"

    # Canonical formula
    positive = orm_campaign_total
    # Negative comes from analytics (it's the natural negative tone for the brand)
    negative = orm_negative_count if orm_negative_count > 0 else analytics_negative
    neutral = max(target_with_campaign - positive - negative, 0)

    return {
        "positive": positive,
        "neutral": neutral,
        "negative": negative,
        "negative_method": negative_method,
        "analytics_target_tonality": {
            "positive": analytics_positive,
            "neutral": analytics_neutral,
            "negative": analytics_negative,
            "total": analytics_total,
        },
        "positive_detail": {
            "orm_campaign_positive": orm_campaign_total,
            "organic_positive": max(positive - orm_campaign_total, 0),
        },
        "negative_detail": {
            "orm_negative_rows": orm_negative_count,
            "analytics_negative_total": analytics_negative,
        },
    }


# ---------------------------------------------------------------------------
# Platform scope
# ---------------------------------------------------------------------------


def _compute_platform_scope(comp_platform_scope: dict, analytics_messages: dict) -> dict:
    """Platform scope uses comparative report data, fallback to analytics messages."""
    if comp_platform_scope.get("total", 0) > 0:
        return {
            "method": "comparative_report_platform_types",
            "total": comp_platform_scope["total"],
            "by_platform_type": comp_platform_scope.get("by_platform_type", {}),
            "platform_type_rows": comp_platform_scope.get("platform_type_rows", []),
            "source_rows": comp_platform_scope.get("source_rows", []),
            "other_count": comp_platform_scope.get("other_count", 0),
        }

    by_platform = analytics_messages.get("by_platform_type", {})
    by_source = analytics_messages.get("by_source", {})

    platform_type_rows = []
    total = sum(by_platform.values()) if by_platform else 0
    for pt in SOURCE_TYPE_ORDER:
        count = int(by_platform.get(pt, 0))
        if count > 0:
            platform_type_rows.append({
                "type": pt,
                "messages": count,
                "share": count / total if total else 0,
            })
    for name, count in sorted(by_platform.items(), key=lambda x: x[1], reverse=True):
        if name not in SOURCE_TYPE_ORDER:
            platform_type_rows.append({
                "type": name,
                "messages": int(count),
                "share": count / total if total else 0,
            })

    source_rows = []
    sorted_sources = sorted(by_source.items(), key=lambda x: x[1], reverse=True)
    top_10 = sorted_sources[:10]
    top_count = sum(c for _, c in top_10)
    for source, count in top_10:
        source_rows.append({
            "source": source,
            "label": source,
            "messages": count,
            "share": count / total if total else 0,
        })
    other_count = max(total - top_count, 0)
    if other_count > 0:
        source_rows.append({
            "source": "other",
            "label": "Другие",
            "messages": other_count,
            "share": other_count / total if total else 0,
        })

    return {
        "method": "analytics_messages_fallback",
        "total": total,
        "by_platform_type": {r["type"]: r["messages"] for r in platform_type_rows},
        "platform_type_rows": platform_type_rows,
        "source_rows": source_rows,
        "other_count": other_count,
    }


# ---------------------------------------------------------------------------
# Topic scope
# ---------------------------------------------------------------------------


def _compute_topic_scope(analytics_messages: dict, project_brand: str, competitor_brands: list[str]) -> dict:
    by_topic = analytics_messages.get("by_topic", {})
    total = sum(by_topic.values()) if by_topic else 0

    topics = []
    for label, count in sorted(by_topic.items(), key=lambda x: x[1], reverse=True)[:12]:
        if len(label.split()) < 2 and len(label) < 8:
            continue
        brand_low = _lower(project_brand)
        if brand_low and brand_low in _lower(label):
            continue
        skip_brand = False
        for cb in competitor_brands or []:
            if _lower(cb) in _lower(label):
                skip_brand = True
                break
        if skip_brand:
            continue
        topics.append({
            "label": label,
            "count": count,
            "examples": [],
        })
        if len(topics) >= 8:
            break

    return {
        "method": "analytics_topic_fields_from_messages_sheet",
        "total": total,
        "topics": topics,
    }


# ---------------------------------------------------------------------------
# Ratings scope
# ---------------------------------------------------------------------------


def _compute_ratings_scope(project_orm_path: Path, ratings_sheet: str, current_period: str) -> dict:
    try:
        wb = load_workbook(project_orm_path, data_only=True, read_only=False)
        sheet = ratings_sheet if ratings_sheet and ratings_sheet in wb.sheetnames else "Рейтинги"
        if sheet not in wb.sheetnames:
            return {
                "periods": [],
                "positive_cards": [],
                "negative_cards": [],
                "no_review_cards": [],
                "reviews_count": [],
                "has_current_period": False,
                "status": "missing_ratings_sheet",
            }
        ws = wb[sheet]
        periods = ["start", "current"]
        positive_cards = [0, 0]
        negative_cards = [0, 0]
        no_reviews = [0, 0]
        reviews_counts = [0, 0]
        for row in ws.iter_rows(min_row=3, values_only=True):
            values = list(row)
            if not values or not values[0]:
                continue
            start_pos = _clean_int(values[1] if len(values) > 1 else 0)
            current_pos = _clean_int(values[5] if len(values) > 5 else 0)
            start_neg = _clean_int(values[2] if len(values) > 2 else 0)
            current_neg = _clean_int(values[6] if len(values) > 6 else 0)
            start_no = _clean_int(values[3] if len(values) > 3 else 0)
            current_no = _clean_int(values[7] if len(values) > 7 else 0)
            start_reviews = _clean_int(values[4] if len(values) > 4 else 0)
            current_reviews = _clean_int(values[8] if len(values) > 8 else 0)
            positive_cards[0] += start_pos
            positive_cards[1] += current_pos
            negative_cards[0] += start_neg
            negative_cards[1] += current_neg
            no_reviews[0] += start_no
            no_reviews[1] += current_no
            reviews_counts[0] += start_reviews
            reviews_counts[1] += current_reviews
        return {
            "periods": ["start", "current"],
            "positive_cards": positive_cards,
            "negative_cards": negative_cards,
            "no_review_cards": no_reviews,
            "reviews_count": reviews_counts,
            "has_current_period": True,
            "status": "ready",
        }
    except Exception as exc:
        return {
            "periods": [],
            "positive_cards": [],
            "negative_cards": [],
            "no_review_cards": [],
            "reviews_count": [],
            "has_current_period": False,
            "status": "error",
            "error": str(exc),
        }


# ---------------------------------------------------------------------------
# Media plan
# ---------------------------------------------------------------------------


def _extract_media_plan(project_orm_path: Path, period_hint: str = "") -> dict:
    try:
        from .media_plan_extractor import extract_media_plan_table
        return extract_media_plan_table(project_orm_path, period_hint=period_hint)
    except Exception:
        return {"status": "missing", "slide_table": {"headers": [], "rows": []}}


# ---------------------------------------------------------------------------
# Period helpers
# ---------------------------------------------------------------------------


def _period_from_content(content_ws) -> str:
    try:
        return _period_label_from_content(content_ws)
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# MAIN: build_canonical_model
# ---------------------------------------------------------------------------

REQUIRED_CONFIG_KEYS = [
    "project_orm_path", "analytics_report_path",
    "project_brand", "report_month",
]


def build_canonical_model(config: dict) -> dict:
    """Build the canonical data model from all input sources.

    Config keys:
        project_orm_path (Path): ORM table workbook
        analytics_report_path (Path): main analytics/report workbook
        comparative_report_path (Path, optional): comparative analytics
        media_plan_path (Path, optional): media plan workbook
        template_path (Path, optional): template PPTX
        project_brand (str): target brand name
        competitors (list[str]): competitor brand names
        brand_aliases (dict, optional): brand -> [aliases]
        report_month (str): month sheet name (e.g. "Май")
        source_exclusions (list[str], optional): sources to exclude
        generation_options (dict, optional): extra options
    """
    project_orm_path = Path(config["project_orm_path"])
    analytics_path = Path(config["analytics_report_path"])
    project_brand = str(config["project_brand"]).strip()
    competitors = [str(c).strip() for c in (config.get("competitors") or []) if str(c or "").strip()]
    report_month = str(config.get("report_month") or "Май").strip()
    source_exclusions = [str(s).strip() for s in (config.get("source_exclusions") or []) if str(s or "").strip()]
    generation_options = config.get("generation_options") or {}
    period = config.get("period") or ""

    brand_aliases = config.get("brand_aliases") or {}
    comparative_path = Path(config["comparative_report_path"]) if config.get("comparative_report_path") else None
    media_plan_path = Path(config["media_plan_path"]) if config.get("media_plan_path") else None

    model = {
        "model_version": "0.4.0",
        "project_brand": project_brand,
        "competitors": competitors,
        "report_month": report_month,
        "period": period,
        "source_exclusions_applied": _build_exclusion_note(source_exclusions),
        "orm_campaign": {},
        "analytics_messages": {},
        "campaign_reconciliation": {},
        "target_brand": {},
        "competitive_sov_scope": {},
        "target_tonality": {},
        "platform_scope": {},
        "topic_scope": {},
        "ratings_scope": {},
        "media_plan": {},
        "raw_sources": {},
    }

    # ---- 1. ORM campaign scope ----
    campaign = _extract_orm_campaign(project_orm_path, report_month, period)
    model["orm_campaign"] = campaign
    target_campaign = campaign.get("total", 0)

    # ---- 2. Comparative report brand data (source of truth for brand metrics) ----
    comp_brand_data = _extract_comparative_brand_data(comparative_path, project_brand, competitors)

    # Fallback: if the comparative report (Показатели sheet) has no data, try
    # Brand Analytics Теги sheet directly from the analytics workbook.
    if not comp_brand_data.get("has_data"):
        ba_data = _extract_ba_brand_data(analytics_path, project_brand, competitors)
        if ba_data.get("has_data"):
            comp_brand_data = ba_data
            comp_brand_data["source"] = "ba_tegi_sheet"

    model["comparative_brand_data"] = comp_brand_data

    # ---- 3. Analytics messages (for reconciliation, source, platform data) ----
    analytics_msgs = _extract_analytics_messages(analytics_path)
    model["analytics_messages"] = analytics_msgs

    # ---- 4. Comparative platform scope ----
    comp_platform = _extract_comparative_platform_scope(comparative_path, project_brand)
    model["comparative_platform_scope"] = comp_platform

    # ---- 5. Source exclusions ----
    if source_exclusions:
        removed = 0
        brand_excluded = Counter()
        excluded_norm = {_lower(s) for s in source_exclusions}
        # Filter analytics messages by source
        if "message_cards" in analytics_msgs:
            filtered_cards = []
            for card in analytics_msgs.get("message_cards", []):
                if _lower(card.get("source", "")) not in excluded_norm:
                    filtered_cards.append(card)
                else:
                    removed += 1
                    # Track which brands were affected
                    obj_raw = card.get("objects_raw", "")
                    if obj_raw:
                        for obj in obj_raw.split(","):
                            obj = obj.strip()
                            if obj:
                                brand_excluded[obj] += 1
            analytics_msgs["message_cards"] = filtered_cards
        if "by_source" in analytics_msgs:
            filtered_source = {}
            for source, count in analytics_msgs["by_source"].items():
                if _lower(source) not in excluded_norm:
                    filtered_source[source] = count
                else:
                    removed += 1
            analytics_msgs["by_source"] = filtered_source
        analytics_msgs["source_exclusions_removed"] = removed
        model["source_exclusions_applied"]["removed_messages_count"] = removed

        # Apply exclusions to comparative brand data
        if comp_brand_data.get("has_data"):
            target_brand_excluded = 0
            for obj_name, exc_count in brand_excluded.items():
                norm_obj = _canonical_brand_name(obj_name, project_brand, competitors)
                for brand_name in list(comp_brand_data["by_brand"].keys()):
                    if _brand_name_matches(brand_name, norm_obj):
                        old = comp_brand_data["by_brand"].get(brand_name, 0)
                        comp_brand_data["by_brand"][brand_name] = max(old - exc_count, 0)
                        target_brand_excluded += exc_count
                        break
            model["source_exclusions_applied"]["brand_excluded_counts"] = dict(brand_excluded)
            model["source_exclusions_applied"]["comparative_brand_adjustments"] = True

    model["raw_sources"] = {
        "project_orm": str(project_orm_path),
        "analytics": str(analytics_path),
        "comparative": str(comparative_path) if comparative_path else None,
    }

    # ---- 6. ORM-to-analytics reconciliation ----
    orm_rows = []
    try:
        wb = load_workbook(project_orm_path, data_only=True, read_only=False)
        from .slide08_seeding_metrics_extractor import _extract_rows as _extract_orm_rows
        target_sheet = report_month if report_month and report_month in wb.sheetnames else None
        if not target_sheet:
            for name in wb.sheetnames:
                if report_month and (report_month.lower() in name.lower() or name.lower() in report_month.lower()):
                    target_sheet = name
                    break
        if target_sheet:
            orm_rows, _meta = _extract_orm_rows(wb[target_sheet])
    except Exception:
        pass

    analytics_cards = analytics_msgs.get("message_cards") or []

    # Build brand object filters: match the project brand.
    # Do NOT add active-substance (INN) variants as brand filters,
    # because INNs like "пентоксифиллин" can be a separate tracked
    # object that belongs to a competitor, not to our campaign.
    # Reconciliation uses platform/date/text matching, not brand
    # name matching, to avoid false positives.
    brand_filters = [_lower(project_brand)]

    reconciliation = _reconcile_orm_with_analytics(
        orm_rows, analytics_cards,
        brand_object_filters=brand_filters,
    )
    model["campaign_reconciliation"] = reconciliation

    # ---- 7. Target brand counts ----
    target_analytics = 0
    for name, count in comp_brand_data.get("by_brand", {}).items():
        if _brand_name_matches(name, project_brand):
            target_analytics = count
            break

    # Determine how many ORM placements were matched in analytics.
    # The reconciliation algorithm produces a match count. If the
    # analytics has at least as many mentions as the campaign, all
    # ORM placements are conceptually "matched" (they are campaign).
    # If analytics has fewer, some ORM is invisible to analytics.
    recon_matched = len(reconciliation.get("analytics_matched_indices", []))

    if target_analytics > 0 and target_campaign > 0:
        if target_campaign <= target_analytics:
            # Analytics has enough mentions to cover the campaign.
            # All ORM placements are campaign (conceptually matched).
            recon_matched = target_campaign
        elif recon_matched > 0:
            # ORM exceeds analytics mentions. Use algorithmic match
            # count with coverage adjustment.
            recon_matched = min(recon_matched, target_analytics)
        else:
            recon_matched = target_analytics

    target_organic = max(target_analytics - recon_matched, 0)
    target_with_campaign = target_organic + target_campaign
    model["target_brand"] = {
        "brand": project_brand,
        "organic": target_organic,
        "campaign": target_campaign,
        "with_campaign": target_with_campaign,
        "without_campaign": target_organic,
        "analytics_mentions": target_analytics,
        "reconciled_matched_count": recon_matched,
        "note": "Campaign = ORM total placements; organic = analytics brand mentions minus reconciled matched count",
    }

    # Update reconciliation with computed value
    model["campaign_reconciliation"]["effective_matched"] = recon_matched

    # ---- 8. Competitive SOV ----
    target_organic_for_sov = model["target_brand"]["organic"]
    target_campaign_for_sov = model["target_brand"]["campaign"]
    sov = _compute_competitive_sov(comp_brand_data, project_brand, competitors, 
                                    target_campaign_for_sov, target_organic_for_sov)
    model["competitive_sov_scope"] = sov

    # ---- 9. Tonality ----
    orm_negative_meta = _extract_orm_negative(project_orm_path, report_month)
    tonality = _compute_tonality(
        comp_brand_data, project_brand,
        target_campaign, target_with_campaign, orm_negative_meta,
    )
    model["target_tonality"] = tonality
    model["orm_negative_meta"] = orm_negative_meta

    # ---- 10. Platform scope ----
    platform = _compute_platform_scope(comp_platform, analytics_msgs)
    model["platform_scope"] = platform

    # ---- 11. Topic scope ----
    topics = _compute_topic_scope(analytics_msgs, project_brand, competitors)
    model["topic_scope"] = topics

    # ---- 12. Ratings scope ----
    ratings_sheet = generation_options.get("ratings_sheet", "Рейтинги")
    current_period = generation_options.get("ratings_current_period", "")
    ratings = _compute_ratings_scope(project_orm_path, ratings_sheet, current_period)
    model["ratings_scope"] = ratings

    # ---- 11. Media plan ----
    media_plan = _extract_media_plan(project_orm_path, period)
    model["media_plan"] = media_plan

    # ---- 12. Period info ----
    try:
        wb = open_workbook_safe(analytics_path)
        for sheet_candidate in ["Содержание", wb.sheetnames[0]]:
            if sheet_candidate in wb.sheetnames:
                period_raw = _period_label_from_content(wb[sheet_candidate])
                if period_raw:
                    break
        else:
            period_raw = period
        prep, caption, title = _period_to_caption(period_raw)
        model["period_raw"] = period_raw
        model["month_prepositional"] = prep
        model["month_caption"] = caption
        model["month_title"] = title
    except Exception:
        model["period_raw"] = period
        model["month_prepositional"] = "в отчетном периоде"
        model["month_caption"] = period or "отчетный период"
        model["month_title"] = "ОТЧЕТНЫЙ ПЕРИОД"

    model["status"] = "ready"

    return model


def save_canonical_model(model: dict, output_dir: Path) -> Path:
    """Save the canonical model to JSON."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "canonical_model.json"

    serializable = deepcopy(model)
    # Convert Path objects to strings
    def _make_serializable(obj):
        if isinstance(obj, Path):
            return str(obj)
        if isinstance(obj, dict):
            return {k: _make_serializable(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_make_serializable(i) for i in obj]
        if isinstance(obj, (Counter, defaultdict)):
            return dict(obj)
        return obj

    serializable = _make_serializable(serializable)
    path.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")
    return path
