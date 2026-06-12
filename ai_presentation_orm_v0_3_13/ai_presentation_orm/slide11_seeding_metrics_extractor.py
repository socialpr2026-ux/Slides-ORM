from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse
import re

from openpyxl import load_workbook

from .slide08_seeding_metrics_extractor import _extract_rows as _extract_orm_rows, _month_sheet as _orm_month_sheet
from .text_contracts import slide11_ratings_insight


MONTH_ORDER = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}

MONTH_LOCATIVE = {
    "январь": "в январе",
    "февраль": "в феврале",
    "март": "в марте",
    "апрель": "в апреле",
    "май": "в мае",
    "июнь": "в июне",
    "июль": "в июле",
    "август": "в августе",
    "сентябрь": "в сентябре",
    "октябрь": "в октябре",
    "ноябрь": "в ноябре",
    "декабрь": "в декабре",
}
MONTH_GENITIVE = {
    "январь": "января",
    "февраль": "февраля",
    "март": "марта",
    "апрель": "апреля",
    "май": "мая",
    "июнь": "июня",
    "июль": "июля",
    "август": "августа",
    "сентябрь": "сентября",
    "октябрь": "октября",
    "ноябрь": "ноября",
    "декабрь": "декабря",
}

MONTH_STEMS = [
    ("январ", 1),
    ("феврал", 2),
    ("март", 3),
    ("апрел", 4),
    ("июн", 6),
    ("июл", 7),
    ("август", 8),
    ("сентябр", 9),
    ("октябр", 10),
    ("ноябр", 11),
    ("декабр", 12),
]


def _text(value) -> str:
    return str(value or "").strip()


def _lower(value) -> str:
    return _text(value).lower().replace("ё", "е")


def _header_index(headers: list[str], *needles: str) -> int | None:
    lowered = [_lower(item) for item in headers]
    for needle in needles:
        low_needle = _lower(needle)
        for idx, header in enumerate(lowered):
            if header == low_needle:
                return idx
    for needle in needles:
        low_needle = _lower(needle)
        for idx, header in enumerate(lowered):
            if low_needle and low_needle in header:
                return idx
    return None


def _clean_int(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    text = re.sub(r"[^0-9,.-]", "", str(value)).replace(",", ".")
    if not text:
        return None
    try:
        return int(round(float(text)))
    except Exception:
        return None


def _clean_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9,.-]", "", str(value)).replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _fmt_int(value: int | float | None) -> str:
    if value is None:
        return "нет данных"
    return f"{int(round(float(value))):,}".replace(",", " ")


def _fmt_rating(value: float | None) -> str:
    if value is None:
        return "нет данных"
    return f"{float(value):.1f}".replace(".", ",")


def _cards_word(value: int) -> str:
    number = abs(int(value))
    if 11 <= number % 100 <= 14:
        return "карточек"
    last = number % 10
    if last == 1:
        return "карточка"
    if 2 <= last <= 4:
        return "карточки"
    return "карточек"


def _ratings_sheet(wb, requested: str = ""):
    if requested and requested in wb.sheetnames:
        return wb[requested], requested
    for name in wb.sheetnames:
        if "рейтинг" in _lower(name):
            return wb[name], name
    raise ValueError("Ratings sheet not found")


def _find_header_rows(ws) -> tuple[int, int]:
    for row_idx in range(1, min(ws.max_row or 0, 20) + 1):
        values = [_lower(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column or 0, 30) + 1)]
        if any(value == "старт" for value in values):
            next_values = [_lower(ws.cell(row_idx + 1, col).value) for col in range(1, min(ws.max_column or 0, 30) + 1)]
            if any("отзыв" in value for value in next_values) and any("рейтинг" in value for value in next_values):
                return row_idx, row_idx + 1
    raise ValueError("Ratings header rows not found")


def _period_pairs(ws, period_row: int, metric_row: int) -> list[dict]:
    pairs = []
    last_label = ""
    for col in range(1, (ws.max_column or 0) + 1):
        label = _text(ws.cell(period_row, col).value)
        if label:
            last_label = label
        metric = _lower(ws.cell(metric_row, col).value)
        next_metric = _lower(ws.cell(metric_row, col + 1).value) if col < (ws.max_column or 0) else ""
        if "отзыв" in metric and "рейтинг" in next_metric and last_label:
            pairs.append({"label": last_label, "reviews_col": col, "rating_col": col + 1})
    return pairs


def _period_sort_key(label: str) -> int:
    low = _lower(label)
    if low == "старт":
        return 0
    if re.search(r"\bма[яйе]\b", low):
        return 5
    for stem, order in MONTH_STEMS:
        if stem in low:
            return order
    for name, order in MONTH_ORDER.items():
        if name in low:
            return order
    return 99


def _period_month_key(label: str) -> tuple[int, int] | None:
    low = _lower(label)
    month = None
    if re.search(r"\bма[яйе]\b", low):
        month = 5
    if month is None:
        for stem, order in MONTH_STEMS:
            if stem in low:
                month = order
                break
    for name, order in MONTH_ORDER.items():
        if month is None and name in low:
            month = order
            break
    if month is None:
        return None
    year_match = re.search(r"(20\d{2})", low)
    year = int(year_match.group(1)) if year_match else 0
    return year, month


def _month_label_from_key(key: tuple[int, int] | None) -> str:
    if key is None:
        return "отчетный период"
    year, month_order = key
    month = next((name for name, order in MONTH_ORDER.items() if order == month_order), "")
    return f"{month} {year}".strip()


def _period_phrase(label: str) -> str:
    low = _lower(label)
    for month, phrase in MONTH_LOCATIVE.items():
        if month in low:
            return phrase
    return f"в периоде {str(label or 'текущий период').lower()}"


def _period_genitive(label: str) -> str:
    low = _lower(label)
    for month, value in MONTH_GENITIVE.items():
        if month in low:
            return value
    return str(label or "текущего периода").lower()


def _select_current_period(periods: list[dict], requested: str = "") -> dict | None:
    non_start = [period for period in periods if _lower(period.get("label")) != "старт"]
    if requested:
        req = _lower(requested)
        matches = []
        for period in non_start:
            if req in _lower(period.get("label")) or _lower(period.get("label")) in req:
                matches.append(period)
        if matches:
            def requested_sort_key(item: dict) -> tuple[int, int, int]:
                key = _period_month_key(item.get("label", ""))
                if key:
                    return key[0], key[1], 0
                return 0, _period_sort_key(item.get("label", "")), 0
            return sorted(matches, key=requested_sort_key)[-1]
    ready = [
        period
        for period in non_start
        if period.get("data_status") == "ready" and (period.get("reviews_total") or period.get("filled_rows", 0) > 0)
    ]
    def sort_key(item: dict) -> tuple[int, int, int]:
        key = _period_month_key(item.get("label", ""))
        if key:
            return key[0], key[1], 0
        return 0, _period_sort_key(item.get("label", "")), 0
    return sorted(ready or non_start, key=sort_key)[-1] if (ready or non_start) else None


def _platform_label(url: str) -> str:
    host = urlparse(str(url or "")).netloc.lower().replace("www.", "")
    mapping = {
        "otzovik.com": "Отзовик",
        "irecommend.ru": "IRecommend",
        "reviews.yandex.ru": "Яндекс",
        "vseotzyvy.ru": "ВсеОтзывы",
        "otzyv.pro": "Отзыв.pro",
        "otzyvru.com": "ОтзывРу",
        "apteka.ru": "Apteka.ru",
        "eapteka.ru": "ЕАПТЕКА",
        "asna.ru": "АСНА",
        "rigla.ru": "Ригла",
        "apteka-april.ru": "Апрель",
        "zdravcity.ru": "Здравсити",
        "budzdorov.ru": "Будь здоров",
        "uteka.ru": "Ютека",
        "megapteka.ru": "Мегаптека",
        "polza.ru": "Polza.ru",
        "megamarket.ru": "Мегамаркет",
        "wildberries.ru": "Wildberries",
        "ozon.ru": "Ozon",
        "market.yandex.ru": "Яндекс.Маркет",
        "msk.vapteke.ru": "В Аптеке",
        "farmlend.ru": "Фармленд",
        "medum.ru": "Medum",
        "med-otzyv.ru": "Мед-Отзыв",
    }
    return mapping.get(host, host or _text(url)[:40] or "Не указано")


MARKETPLACE_PLATFORMS = {
    "Ozon",
    "Wildberries",
    "Яндекс.Маркет",
    "Мегамаркет",
}

PHARMACY_PLATFORMS = {
    "Apteka.ru",
    "ЕАПТЕКА",
    "АСНА",
    "Ригла",
    "Апрель",
    "Здравсити",
    "Будь здоров",
    "Ютека",
    "Мегаптека",
    "Polza.ru",
    "В Аптеке",
    "Фармленд",
    "Medum",
}


def _platform_group(platform: str) -> str:
    label = _text(platform)
    if label in MARKETPLACE_PLATFORMS:
        return "Маркетплейсы"
    return "Аптеки и отзовики"


def _platform_display(platform: str) -> str:
    label = _text(platform)
    mapping = {
        "Apteka.ru": "www.apteka.ru",
        "Ютека": "www.uteka.ru",
        "IRecommend": "www.irecommend.ru",
        "Отзовик": "www.otzovik.com",
        "Яндекс": "www.reviews.yandex.ru",
        "Отзыв.pro": "www.otzyv.pro",
        "ВсеОтзывы": "www.vseotzyvy.ru",
        "Ozon": "www.ozon.ru",
        "Wildberries": "www.wildberries.ru",
        "Яндекс.Маркет": "market.yandex.ru",
        "Мегаптека": "www.megapteka.ru",
        "ЕАПТЕКА": "www.eapteka.ru",
    }
    return mapping.get(label, label or "Не указано")


def _row_kind(rating: float | None, reviews: int | None) -> str:
    if not reviews:
        return "no_reviews"
    if rating is None:
        return "no_reviews"
    if rating > 4:
        return "positive_cards"
    if rating < 4:
        return "negative_cards"
    return "neutral_cards"


def _summarize_period(rows: list[dict], period: dict) -> dict:
    label = period.get("label", "")
    reviews_key = f"reviews_{label}"
    rating_key = f"rating_{label}"
    cards_total = 0
    cards_with_reviews = 0
    positive_cards = 0
    neutral_cards = 0
    negative_cards = 0
    no_reviews = 0
    reviews_total = 0
    weighted_rating_sum = 0.0
    filled_rows = 0
    for row in rows:
        reviews = row.get(reviews_key)
        rating = row.get(rating_key)
        has_value = reviews is not None or rating is not None
        if has_value:
            filled_rows += 1
        if not has_value and _lower(label) != "старт":
            continue
        cards_total += 1
        if reviews:
            cards_with_reviews += 1
            reviews_total += int(reviews)
            if rating is not None:
                weighted_rating_sum += int(reviews) * float(rating)
        kind = _row_kind(rating, reviews)
        if kind == "positive_cards":
            positive_cards += 1
        elif kind == "negative_cards":
            negative_cards += 1
        elif kind == "neutral_cards":
            neutral_cards += 1
        else:
            no_reviews += 1
    denominator = cards_total or len(rows)
    return {
        "label": label,
        "cards_total": denominator,
        "cards_with_reviews": cards_with_reviews,
        "positive_cards": positive_cards,
        "neutral_cards": neutral_cards,
        "negative_cards": negative_cards,
        "no_reviews": no_reviews,
        "reviews_total": reviews_total if filled_rows or _lower(label) == "старт" else None,
        "weighted_rating": weighted_rating_sum / reviews_total if reviews_total else None,
        "positive_share": positive_cards / denominator if denominator else 0,
        "neutral_share": neutral_cards / denominator if denominator else 0,
        "negative_share": negative_cards / denominator if denominator else 0,
        "no_reviews_share": no_reviews / denominator if denominator else 0,
        "filled_rows": filled_rows,
        "data_status": "ready" if filled_rows or _lower(label) == "старт" else "missing",
    }


def _summary_block(ws, periods: list[dict]) -> dict[str, dict]:
    labels = {}
    for row_idx in range(1, (ws.max_row or 0) + 1):
        label = _lower(ws.cell(row_idx, 1).value)
        if "отзывов всего" in label:
            labels["reviews_total"] = row_idx
        elif "рейтингом >4" in label:
            labels["positive"] = row_idx
        elif "рейтингом <4" in label:
            labels["negative"] = row_idx
        elif "без отзыв" in label:
            labels["no_reviews"] = row_idx
        if {"reviews_total", "positive", "negative", "no_reviews"}.issubset(labels):
            break
    if not {"reviews_total", "positive", "negative", "no_reviews"}.issubset(labels):
        return {}

    out = {}
    for period in periods:
        label = period["label"]
        count_col = period["reviews_col"]
        share_col = period["rating_col"]
        positive = _clean_int(ws.cell(labels["positive"], count_col).value) or 0
        negative = _clean_int(ws.cell(labels["negative"], count_col).value) or 0
        no_reviews = _clean_int(ws.cell(labels["no_reviews"], count_col).value) or 0
        cards_total = positive + negative + no_reviews
        reviews_total = _clean_int(ws.cell(labels["reviews_total"], count_col).value)
        positive_share = _clean_float(ws.cell(labels["positive"], share_col).value)
        negative_share = _clean_float(ws.cell(labels["negative"], share_col).value)
        no_reviews_share = _clean_float(ws.cell(labels["no_reviews"], share_col).value)
        out[label] = {
            "summary_source": "ratings_sheet_summary_block",
            "summary_rows": labels,
            "cards_total": cards_total,
            "cards_with_reviews": positive + negative,
            "positive_cards": positive,
            "neutral_cards": 0,
            "negative_cards": negative,
            "no_reviews": no_reviews,
            "reviews_total": reviews_total,
            "positive_share": positive_share if positive_share is not None else (positive / cards_total if cards_total else 0),
            "neutral_share": 0,
            "negative_share": negative_share if negative_share is not None else (negative / cards_total if cards_total else 0),
            "no_reviews_share": no_reviews_share if no_reviews_share is not None else (no_reviews / cards_total if cards_total else 0),
            "data_status": "ready" if cards_total > 0 or int(reviews_total or 0) > 0 or _lower(label) == "старт" else "missing",
        }
    return out


def _extract_rows(ws, metric_row: int, periods: list[dict]) -> list[dict]:
    rows = []
    for row_idx in range(metric_row + 1, (ws.max_row or 0) + 1):
        first = _text(ws.cell(row_idx, 1).value)
        if not first:
            continue
        low = _lower(first)
        if "отзывов всего" in low:
            break
        if not first.lower().startswith(("http://", "https://")):
            continue
        row = {
            "row": row_idx,
            "url": first,
            "platform": _platform_label(first),
            "product": _text(ws.cell(row_idx, 2).value),
        }
        for period in periods:
            label = period["label"]
            row[f"reviews_{label}"] = _clean_int(ws.cell(row_idx, period["reviews_col"]).value)
            row[f"rating_{label}"] = _clean_float(ws.cell(row_idx, period["rating_col"]).value)
        if any(row.get(f"reviews_{period['label']}") is not None or row.get(f"rating_{period['label']}") is not None for period in periods):
            rows.append(row)
    return rows


def _top_platform_rows(rows: list[dict], start_label: str, current_label: str, limit: int = 10) -> list[dict]:
    grouped = defaultdict(lambda: {"cards": 0, "start_reviews": 0, "current_reviews": 0, "start_weighted": 0.0, "current_weighted": 0.0})
    for row in rows:
        item = grouped[row["platform"]]
        item["cards"] += 1
        for prefix, label in [("start", start_label), ("current", current_label)]:
            reviews = row.get(f"reviews_{label}")
            rating = row.get(f"rating_{label}")
            if reviews:
                item[f"{prefix}_reviews"] += int(reviews)
                if rating is not None:
                    item[f"{prefix}_weighted"] += int(reviews) * float(rating)
    out = []
    for platform, item in grouped.items():
        out.append({
            "platform": platform,
            "group": _platform_group(platform),
            "cards": item["cards"],
            "start_reviews": item["start_reviews"],
            "current_reviews": item["current_reviews"],
            "start_rating": item["start_weighted"] / item["start_reviews"] if item["start_reviews"] else None,
            "current_rating": item["current_weighted"] / item["current_reviews"] if item["current_reviews"] else None,
        })
    return sorted(out, key=lambda row: (row["current_reviews"], row["start_reviews"]), reverse=True)[:limit]


def _platform_group_rows(rows: list[dict], start_label: str, current_label: str) -> list[dict]:
    grouped = defaultdict(lambda: {"cards": 0, "start_reviews": 0, "current_reviews": 0})
    for row in rows:
        group = _platform_group(row.get("platform", ""))
        item = grouped[group]
        item["cards"] += 1
        item["start_reviews"] += int(row.get(f"reviews_{start_label}") or 0)
        item["current_reviews"] += int(row.get(f"reviews_{current_label}") or 0)
    ordered = []
    for group in ["Маркетплейсы", "Аптеки и отзовики"]:
        values = grouped.get(group, {})
        ordered.append({
            "group": group,
            "cards": int(values.get("cards", 0) or 0),
            "start_reviews": int(values.get("start_reviews", 0) or 0),
            "current_reviews": int(values.get("current_reviews", 0) or 0),
        })
    return ordered


def _placement_sheet(wb, requested: str = ""):
    candidates = []
    if requested:
        candidates.append(requested)
    candidates.extend(["Май", "Апрель", "Июнь"])
    candidates.extend(wb.sheetnames)
    seen = set()
    for name in candidates:
        if not name or name in seen or name not in wb.sheetnames:
            continue
        seen.add(name)
        ws = wb[name]
        try:
            header_row, headers = _find_placement_header(ws)
        except Exception:
            continue
        return ws, name, header_row, headers
    return None, "", 0, []


def _find_placement_header(ws):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 60), values_only=True), start=1):
        headers = [_text(value) for value in row]
        lowered = [_lower(value) for value in headers]
        if any("тип размещ" in value for value in lowered) and any("площад" in value or "ссылка" in value for value in lowered):
            return row_idx, headers
    raise ValueError("Placement header not found")


def _placement_summary(wb, *, month_sheet: str = "") -> dict:
    try:
        ws, sheet_name = _orm_month_sheet(wb, month_sheet or "Май")
        rows, _meta = _extract_orm_rows(ws)
    except Exception:
        return {"status": "missing", "sheet": month_sheet, "placed_reviews": 0, "platforms": []}
    review_rows = [row for row in rows if row.get("material_group") == "Отзывы" or "отзыв" in _lower(row.get("placement_type"))]
    platforms = Counter(row.get("platform") or "Не указано" for row in review_rows)
    groups = Counter(_platform_group(row.get("platform") or "") for row in review_rows)
    return {
        "status": "ready" if review_rows else "empty",
        "sheet": sheet_name,
        "placed_reviews": len(review_rows),
        "platforms": [
            {
                "platform": key,
                "display": _platform_display(key),
                "group": _platform_group(key),
                "reviews": int(value),
            }
            for key, value in platforms.most_common()
        ],
        "groups": [{"group": key, "reviews": int(value)} for key, value in groups.most_common()],
    }


def _largest_growth_platforms(top_platforms: list[dict], limit: int = 3) -> list[dict]:
    rows = []
    for row in top_platforms:
        growth = int(row.get("current_reviews", 0) or 0) - int(row.get("start_reviews", 0) or 0)
        if growth <= 0:
            continue
        item = dict(row)
        item["review_growth"] = growth
        rows.append(item)
    return sorted(rows, key=lambda item: item["review_growth"], reverse=True)[:limit]


def _summary_text(data: dict) -> str:
    return "• " + slide11_ratings_insight(data)

    brand = data.get("project_brand") or "бренда"
    start = data.get("start_summary") or {}
    current = data.get("current_summary") or {}
    current_label = current.get("label") or "текущий период"
    current_genitive = _period_genitive(current_label)
    caveat = data.get("ratings_period_caveat") or ""
    caveat_prefix = f"{caveat} " if caveat else ""
    if current.get("data_status") == "missing":
        return (
            f"• {caveat_prefix}На старте по бренду {brand} зафиксировано {_fmt_int(start.get('reviews_total', 0))} отзывов. "
            f"Доля карточек с рейтингом выше 4 баллов составляет {start.get('positive_share', 0) * 100:.0f}%.\n"
            "• В текущем периоде нет заполненных рейтинговых значений, поэтому динамика относительно старта не рассчитывается."
        )
    negative_cards = int(current.get("negative_cards") or 0)
    if negative_cards:
        negative_sentence = f"Карточек с рейтингом ниже 4 баллов зафиксировано: {negative_cards}."
    else:
        negative_sentence = "Негативных карточек не зафиксировано."
    start_reviews = int(start.get("reviews_total") or 0)
    current_reviews = int(current.get("reviews_total") or 0)
    placement = data.get("agency_review_placements") or {}
    placed_reviews = int(placement.get("placed_reviews") or 0)
    placement_platforms = ", ".join(row.get("platform", "") for row in (placement.get("platforms") or [])[:3] if row.get("platform"))
    growth_sources = data.get("largest_review_growth_platforms") or []
    growth_text = ", ".join(f"{row.get('platform')} (+{_fmt_int(row.get('review_growth'))})" for row in growth_sources[:3])
    if current_reviews > start_reviews:
        reviews_sentence = f"Количество отзывов на сайтах в топе выдачи на конец {current_genitive} увеличилось с {_fmt_int(start_reviews)} до {_fmt_int(current_reviews)} сообщений."
    elif current_reviews < start_reviews:
        reviews_sentence = f"Количество отзывов на сайтах в топе выдачи на конец {current_genitive} снизилось с {_fmt_int(start_reviews)} до {_fmt_int(current_reviews)} сообщений."
    else:
        reviews_sentence = f"Количество отзывов на сайтах в топе выдачи на конец {current_genitive} не изменилось и составило {_fmt_int(current_reviews)} сообщений."
    start_positive = start.get("positive_share", 0) * 100
    current_positive = current.get("positive_share", 0) * 100
    positive_delta = current_positive - start_positive
    if positive_delta > 0:
        delta_text = f"{positive_delta:.0f}".replace(".", ",")
        positive_sentence = f"Доля позитивных карточек выше старта на {delta_text} п.п."
    elif positive_delta < 0:
        delta_text = f"{abs(positive_delta):.0f}".replace(".", ",")
        positive_sentence = f"Доля позитивных карточек ниже старта на {delta_text} п.п."
    else:
        positive_sentence = "Доля позитивных карточек сохранилась на уровне старта"
    no_reviews_cards = int(current.get("no_reviews") or 0)
    if negative_cards:
        risk_sentence = f"Зона контроля — {negative_cards} {_cards_word(negative_cards)} с рейтингом ниже 4 баллов."
    elif no_reviews_cards:
        risk_sentence = f"Зона роста — {no_reviews_cards} {_cards_word(no_reviews_cards)} без отзывов, где рейтинг может быть усилен накоплением пользовательских оценок."
    else:
        risk_sentence = "Критических рейтинговых зон в текущем периоде не выделено."
    return (
        f"• {brand} сохраняет сильную рейтинговую базу: {current_positive:.0f}% карточек имеют рейтинг выше 4 баллов. "
        f"{positive_sentence} {negative_sentence}\n"
        f"• {reviews_sentence} За период добавлено {_fmt_int(placed_reviews)} свежих отзывов.\n"
        f"• {risk_sentence} Основной пользовательский прирост дают площадки: {growth_text or placement_platforms}."
    )


def extract_slide11_seeding_metrics(
    *,
    ratings_path: Path,
    project_brand: str,
    period: str = "",
    ratings_sheet: str = "",
    current_period: str = "",
    placement_month_sheet: str = "",
    source_system: str = "проектная таблица ORM",
) -> dict:
    project_period = period
    wb = load_workbook(ratings_path, data_only=True, read_only=False)
    ws, sheet_name = _ratings_sheet(wb, ratings_sheet)
    period_row, metric_row = _find_header_rows(ws)
    periods = _period_pairs(ws, period_row, metric_row)
    if not periods:
        raise ValueError("Rating period columns not found")
    start_period = next((item for item in periods if _lower(item.get("label")) == "старт"), periods[0])
    rows = _extract_rows(ws, metric_row, periods)
    period_summaries = [_summarize_period(rows, item) for item in periods]
    summary_by_period = _summary_block(ws, periods)
    if summary_by_period:
        merged_summaries = []
        for summary in period_summaries:
            override = summary_by_period.get(summary["label"])
            if override:
                item = dict(summary)
                for key, value in override.items():
                    if key == "reviews_total" and value is None and item.get("reviews_total") is not None:
                        continue
                    item[key] = value
                merged_summaries.append(item)
            else:
                merged_summaries.append(summary)
        period_summaries = merged_summaries
    for period_item, summary in zip(periods, period_summaries):
        period_item.update(summary)
    current = _select_current_period(periods, current_period)
    if current is None:
        current = start_period
    start_summary = next((item for item in period_summaries if item["label"] == start_period["label"]), {})
    current_summary = next((item for item in period_summaries if item["label"] == current["label"]), {})
    top_platforms = _top_platform_rows(rows, start_period["label"], current["label"])
    placement_summary = _placement_summary(wb, month_sheet=placement_month_sheet)
    largest_growth_platforms = _largest_growth_platforms(top_platforms)
    platform_group_rows = _platform_group_rows(rows, start_period["label"], current["label"])
    report_period_key = _period_month_key(project_period)
    current_period_key = _period_month_key(current.get("label", ""))
    ratings_matches_report_period = False
    if report_period_key and current_period_key:
        report_year, report_month = report_period_key
        current_year, current_month = current_period_key
        ratings_matches_report_period = report_month == current_month and (
            report_year == current_year or not report_year or not current_year
        )
    ratings_period_caveat = ""
    if report_period_key and current_period_key and not ratings_matches_report_period:
        ratings_period_caveat = (
            f"Нет рейтинговых данных за {_month_label_from_key(report_period_key)}, "
            f"поэтому динамика показана по последнему доступному периоду — {current.get('label', '').lower()}."
        )
    data = {
        "source_file": ratings_path.name,
        "source_system": source_system,
        "source_label": "проектная таблица ORM",
        "project_brand": project_brand,
        "period_raw": project_period,
        "sheet": sheet_name,
        "header_rows": {"period_row": period_row, "metric_row": metric_row},
        "selected_periods": {"start": start_period["label"], "current": current["label"]},
        "ratings_report_period_key": list(report_period_key) if report_period_key else None,
        "ratings_current_period_key": list(current_period_key) if current_period_key else None,
        "ratings_matches_report_period": ratings_matches_report_period,
        "ratings_period_caveat": ratings_period_caveat,
        "period_summaries": period_summaries,
        "start_summary": start_summary,
        "current_summary": current_summary,
        "rating_rows_count": len(rows),
        "top_platform_rows": top_platforms,
        "review_platform_groups": platform_group_rows,
        "largest_review_growth_platforms": largest_growth_platforms,
        "agency_review_placements": placement_summary,
        "review_totals": {
            "start": int(start_summary.get("reviews_total") or 0),
            "current": int(current_summary.get("reviews_total") or 0) if current_summary.get("reviews_total") is not None else None,
            "delta": (
                int(current_summary.get("reviews_total") or 0) - int(start_summary.get("reviews_total") or 0)
                if current_summary.get("reviews_total") is not None and start_summary.get("reviews_total") is not None
                else None
            ),
            "agency_reviews": int(placement_summary.get("placed_reviews") or 0),
        },
        "sample_rows": rows[:12],
        "methodology": {
            "row_filter": "Rows with URL in the first column before the summary block.",
            "review_count": "Review count is read from the period column whose second header is 'Отзывы'.",
            "rating": "Rating is read from the adjacent period column whose second header is 'Рейтинг'.",
            "card_status": "Rating > 4 is positive, rating < 4 is negative, empty or zero reviews are counted as no reviews.",
            "chart_totals": "If the sheet has summary rows for total reviews and card rating groups, these rows are used as authoritative chart totals.",
            "current_period_selection": "Requested period if present, otherwise latest non-start period with filled data.",
        },
        "methodology_status": "ready" if rows and start_summary else "blocked",
        "input_role_note": "Slide 11 uses the ratings workbook configured for this run.",
    }
    data["insight_text"] = _summary_text(data)
    return data
